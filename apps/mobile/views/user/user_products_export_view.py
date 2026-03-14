"""
Vue optimisée pour l'export CSV/Excel de produits en streaming pour Flutter/Drift.
Optimisé pour gérer jusqu'à 1M d'articles avec streaming et compression.
"""
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from django.http import StreamingHttpResponse, HttpResponse
from django.db import connection
import csv
import io
import gzip
from typing import Iterator, Optional
import logging

from apps.mobile.services.user_service import UserService
from apps.mobile.utils import error_response
from apps.mobile.exceptions import (
    UserNotFoundException,
    AccountNotFoundException,
    ProductNotFoundException,
    DataValidationException,
    DatabaseConnectionException
)

logger = logging.getLogger(__name__)


class UserProductsExportView(APIView):
    """
    API optimisée pour exporter les produits en CSV/Excel avec streaming.
    
    Optimisations pour 1M+ d'articles :
    - Streaming pour éviter la charge mémoire
    - Compression gzip optionnelle
    - Pagination par chunks
    - Format CSV optimisé pour Drift upsert
    
    URL: /mobile/api/products/export/
    
    Paramètres query :
    - format: 'csv' (défaut) ou 'excel'
    - compressed: 'true' pour activer gzip (défaut: false)
    - chunk_size: Nombre de lignes par chunk (défaut: 10000)
    """
    permission_classes = [IsAuthenticated]
    
    # Configuration optimale pour 1M d'articles
    DEFAULT_CHUNK_SIZE = 10000  # 10K lignes par chunk
    MAX_CHUNK_SIZE = 50000      # Maximum pour éviter timeout
    
    def get(self, request):
        """
        Exporte les produits en streaming CSV/Excel.
        
        Args:
            request: Requête GET avec paramètres optionnels :
                - format: 'csv' ou 'excel' (défaut: 'csv')
                - compressed: 'true' pour gzip (défaut: 'false')
                - chunk_size: Taille des chunks (défaut: 10000)
        """
        try:
            user_id = request.user.id
            
            # Paramètres de la requête
            export_format = request.GET.get('format', 'csv').lower()
            compressed = request.GET.get('compressed', 'false').lower() == 'true'
            chunk_size = min(
                int(request.GET.get('chunk_size', self.DEFAULT_CHUNK_SIZE)),
                self.MAX_CHUNK_SIZE
            )
            
            # Validation du format
            if export_format not in ['csv', 'excel']:
                return error_response(
                    message="Format non supporté. Utilisez 'csv' ou 'excel'",
                    status_code=status.HTTP_400_BAD_REQUEST,
                    error_type='VALIDATION_ERROR'
                )
            
            # Récupérer le compte de l'utilisateur
            user_service = UserService()
            
            # Vérifier que l'utilisateur existe
            try:
                from apps.users.models import UserApp
                user = UserApp.objects.get(id=user_id)
                account = user.compte
                if not account:
                    raise AccountNotFoundException(f"Aucun compte associé à l'utilisateur {user_id}")
            except UserApp.DoesNotExist:
                raise UserNotFoundException(f"Utilisateur avec l'ID {user_id} non trouvé")
            
            # Générer le nom du fichier
            from datetime import datetime
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            extension = 'csv.gz' if (compressed and export_format == 'csv') else export_format
            filename = f"produits_export_{user_id}_{timestamp}.{extension}"
            
            # Générer le stream selon le format
            if export_format == 'csv':
                return self._stream_csv(user_id, account.id, chunk_size, compressed, filename)
            else:
                # Pour Excel, on utilise une approche différente (plus lourd)
                return self._stream_excel(user_id, account.id, chunk_size, filename)
                
        except (UserNotFoundException, AccountNotFoundException) as e:
            return error_response(
                message=str(e),
                status_code=status.HTTP_404_NOT_FOUND,
                error_type='NOT_FOUND'
            )
        except Exception as e:
            logger.error(f"Erreur lors de l'export: {str(e)}", exc_info=True)
            return error_response(
                message="Erreur interne du serveur",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                error_type='INTERNAL_ERROR'
            )
    
    def _stream_csv(self, user_id: int, account_id: int, chunk_size: int, 
                   compressed: bool, filename: str) -> StreamingHttpResponse:
        """
        Génère un CSV en streaming optimisé pour Drift upsert.
        
        Format optimisé :
        - Colonnes minimales nécessaires
        - Pas de formatage complexe
        - UTF-8 avec BOM pour Excel
        - Prêt pour import direct dans Drift
        """
        def generate_csv_chunks() -> Iterator[bytes]:
            """Générateur qui produit les chunks CSV"""
            # Buffer pour les données CSV
            buffer = io.StringIO()
            writer = None
            
            # En-têtes CSV optimisés pour Drift
            headers = [
                'web_id',              # Clé primaire pour upsert
                'product_name',
                'designation',          # Short_Description
                'product_code',
                'internal_product_code',
                'family_name',
                'is_variant',
                'n_lot',
                'n_serie',            # Booléen: produit supporte les numéros de série
                'dlc',
                'numeros_serie',      # JSON stringifié: liste des numéros de série (un article = une ligne)
                'created_at',
                'updated_at'
            ]
            
            # Écrire le BOM UTF-8 pour Excel
            yield '\ufeff'.encode('utf-8')
            
            # Écrire les en-têtes
            header_line = ','.join(headers) + '\n'
            yield header_line.encode('utf-8')
            
            # Récupérer les produits par chunks depuis la DB
            offset = 0
            total_processed = 0
            
            while True:
                # Récupérer un chunk de produits directement depuis la DB
                products = self._get_products_chunk(account_id, offset, chunk_size)
                
                if not products:
                    break
                
                # Écrire chaque produit
                for product in products:
                    row = self._format_product_row(product)
                    yield (row + '\n').encode('utf-8')
                    total_processed += 1
                
                offset += chunk_size
                
                # Log de progression pour les gros exports
                if total_processed % 100000 == 0:
                    logger.info(f"Export CSV: {total_processed} produits traités")
            
            logger.info(f"Export CSV terminé: {total_processed} produits exportés")
        
        # Créer la réponse streaming
        content_type = 'text/csv; charset=utf-8'
        if compressed:
            content_type = 'application/gzip'
            # Wrapper avec gzip
            response = StreamingHttpResponse(
                self._gzip_wrapper(generate_csv_chunks()),
                content_type=content_type
            )
            response['Content-Encoding'] = 'gzip'
        else:
            response = StreamingHttpResponse(
                generate_csv_chunks(),
                content_type=content_type
            )
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['X-Total-Processed'] = 'streaming'  # On ne connaît pas le total à l'avance
        
        return response
    
    def _gzip_wrapper(self, iterator: Iterator[bytes]) -> Iterator[bytes]:
        """Wrapper pour compresser le stream avec gzip"""
        buffer = io.BytesIO()
        compressor = gzip.GzipFile(mode='wb', fileobj=buffer)
        
        try:
            for chunk in iterator:
                compressor.write(chunk)
                # Flush périodique pour éviter les gros buffers
                if buffer.tell() > 1024 * 1024:  # 1MB
                    compressor.flush()
                    data = buffer.getvalue()
                    buffer.seek(0)
                    buffer.truncate(0)
                    yield data
        finally:
            # Flush final
            compressor.close()
            remaining = buffer.getvalue()
            if remaining:
                yield remaining
    
    def _get_products_chunk(self, account_id: int, offset: int, limit: int) -> list:
        """
        Récupère un chunk de produits directement depuis la DB.
        Optimisé pour éviter de charger tout en mémoire.
        Inclut les numéros de série sans dupliquer les articles.
        """
        from apps.masterdata.models import Product, NSerie
        import json
        
        try:
            # Récupérer les produits avec prefetch pour optimiser les numéros de série
            products = Product.objects.select_related('Product_Family').prefetch_related(
                'numeros_serie'  # Prefetch des numéros de série
            ).filter(
                Product_Family__compte_id=account_id,
                Product_Status='ACTIVE'
            ).order_by('id')[offset:offset + limit]
            
            # Formater les données
            formatted = []
            for product in products:
                # Récupérer les numéros de série actifs pour ce produit
                numeros_serie_list = []
                if getattr(product, 'n_serie', False):
                    # Récupérer les numéros de série actifs
                    nseries = product.numeros_serie.filter(status='ACTIVE').order_by('n_serie')
                    for nserie in nseries:
                        numeros_serie_list.append({
                            'id': nserie.id,
                            'n_serie': nserie.n_serie,
                            'reference': nserie.reference,
                            'status': nserie.status,
                        })
                
                # Un seul article = une ligne, même avec plusieurs numéros de série
                formatted.append({
                    'web_id': product.id,
                    'product_name': product.Short_Description or '',
                    'designation': product.Short_Description or '',
                    'product_code': product.Product_Code or '',
                    'internal_product_code': product.Internal_Product_Code or '',
                    'family_name': product.Product_Family.Family_Name if product.Product_Family else '',
                    'is_variant': getattr(product, 'is_variant', False),
                    'n_lot': getattr(product, 'n_lot', False),
                    'n_serie': getattr(product, 'n_serie', False),
                    'dlc': getattr(product, 'dlc', False),
                    'numeros_serie': json.dumps(numeros_serie_list, ensure_ascii=False),  # JSON stringifié
                    'created_at': product.created_at.isoformat() if hasattr(product, 'created_at') and product.created_at else '',
                    'updated_at': product.updated_at.isoformat() if hasattr(product, 'updated_at') and product.updated_at else '',
                })
            
            return formatted
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des produits: {str(e)}")
            return []
    
    def _format_product_row(self, product: dict) -> str:
        """
        Formate une ligne de produit pour CSV.
        Échappe les valeurs et gère les caractères spéciaux.
        """
        def escape_csv_value(value):
            """Échappe une valeur pour CSV"""
            if value is None:
                return ''
            value_str = str(value)
            # Si contient virgule, guillemets ou saut de ligne, entourer de guillemets
            if ',' in value_str or '"' in value_str or '\n' in value_str:
                # Échapper les guillemets doubles
                value_str = value_str.replace('"', '""')
                return f'"{value_str}"'
            return value_str
        
        # Convertir les booléens en 0/1 pour Drift
        is_variant = '1' if product.get('is_variant') else '0'
        n_lot = '1' if product.get('n_lot') else '0'
        n_serie = '1' if product.get('n_serie') else '0'
        dlc = '1' if product.get('dlc') else '0'
        
        row = [
            escape_csv_value(product.get('web_id', '')),
            escape_csv_value(product.get('product_name', '')),
            escape_csv_value(product.get('designation', '')),
            escape_csv_value(product.get('product_code', '')),
            escape_csv_value(product.get('internal_product_code', '')),
            escape_csv_value(product.get('family_name', '')),
            is_variant,
            n_lot,
            n_serie,
            dlc,
            escape_csv_value(product.get('numeros_serie', '[]')),
            escape_csv_value(product.get('created_at', '')),
            escape_csv_value(product.get('updated_at', '')),
        ]
        
        return ','.join(row)
    
    def _stream_excel(self, user_id: int, account_id: int, chunk_size: int, filename: str):
        """
        Export Excel avec streaming (moins optimal que CSV pour 1M+).
        Note: Excel a des limitations (1M lignes max), donc on recommande CSV.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.writer.excel import save_virtual_workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            return error_response(
                message="openpyxl est requis pour l'export Excel. Installez-le avec: pip install openpyxl",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                error_type='DEPENDENCY_ERROR'
            )
        
        def generate_excel_stream() -> Iterator[bytes]:
            """Générateur qui produit le fichier Excel par chunks"""
            # Créer le workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Produits"
            
            # En-têtes
            headers = [
                'ID', 'Nom du produit', 'Désignation', 'Code produit', 'Code produit interne',
                'Famille', 'Est variante', 'N° Lot', 'N° Série', 'DLC',
                'Numéros de série', 'Date de création', 'Date de mise à jour'
            ]
            
            # Écrire les en-têtes
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = cell.font.copy(bold=True)
            
            # Récupérer et écrire les produits par chunks
            offset = 0
            total_processed = 0
            row_num = 2
            
            while True:
                products = self._get_products_chunk(account_id, offset, chunk_size)
                
                if not products:
                    break
                
                # Écrire chaque produit
                for product in products:
                    ws.cell(row=row_num, column=1, value=product.get('web_id'))
                    ws.cell(row=row_num, column=2, value=product.get('product_name'))
                    ws.cell(row=row_num, column=3, value=product.get('designation'))
                    ws.cell(row=row_num, column=4, value=product.get('product_code'))
                    ws.cell(row=row_num, column=5, value=product.get('internal_product_code'))
                    ws.cell(row=row_num, column=6, value=product.get('family_name'))
                    ws.cell(row=row_num, column=7, value='Oui' if product.get('is_variant') else 'Non')
                    ws.cell(row=row_num, column=8, value='Oui' if product.get('n_lot') else 'Non')
                    ws.cell(row=row_num, column=9, value='Oui' if product.get('n_serie') else 'Non')
                    ws.cell(row=row_num, column=10, value='Oui' if product.get('dlc') else 'Non')
                    ws.cell(row=row_num, column=11, value=str(product.get('numeros_serie', '[]')))
                    ws.cell(row=row_num, column=12, value=product.get('created_at', ''))
                    ws.cell(row=row_num, column=13, value=product.get('updated_at', ''))
                    
                    row_num += 1
                    total_processed += 1
                    
                    # Avertissement si on approche la limite Excel
                    if total_processed >= 1000000:
                        logger.warning("Limite Excel atteinte (1M lignes). Utilisez CSV pour plus d'articles.")
                        break
                
                offset += chunk_size
                
                # Log de progression
                if total_processed % 100000 == 0:
                    logger.info(f"Export Excel: {total_processed} produits traités")
            
            # Ajuster la largeur des colonnes
            column_widths = {
                'A': 10, 'B': 30, 'C': 30, 'D': 20, 'E': 20,
                'F': 20, 'G': 15, 'H': 10, 'I': 10, 'J': 10,
                'K': 40, 'L': 20, 'M': 20
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            logger.info(f"Export Excel terminé: {total_processed} produits exportés")
            
            # Sauvegarder dans un buffer
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Yielder le contenu par chunks pour le streaming
            chunk_size_bytes = 1024 * 1024  # 1MB chunks
            while True:
                chunk = excel_buffer.read(chunk_size_bytes)
                if not chunk:
                    break
                yield chunk
        
        # Créer la réponse streaming
        response = StreamingHttpResponse(
            generate_excel_stream(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

