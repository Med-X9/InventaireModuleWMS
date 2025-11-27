from rest_framework import status
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from django.http import HttpResponse
import io
from datetime import datetime

from apps.mobile.services.user_service import UserService
from apps.mobile.utils import error_response
from apps.mobile.exceptions import (
    UserNotFoundException,
    AccountNotFoundException,
    ProductNotFoundException,
    DataValidationException,
    DatabaseConnectionException
)


class UserProductsView(APIView):
    """
    API pour exporter les produits du même compte qu'un utilisateur en format Excel.
    
    Permet d'exporter la liste des produits appartenant au même compte
    de l'utilisateur connecté sous forme de fichier Excel (.xlsx).
    Utile pour la gestion des inventaires et l'export des produits disponibles.
    
    URL: /mobile/api/products/
    
    Fonctionnalités:
    - Export des produits du même compte en format Excel
    - L'utilisateur est récupéré automatiquement depuis le token d'authentification
    - Filtrage par compte associé à l'utilisateur connecté
    - Génération d'un fichier Excel avec formatage et en-têtes stylisés
    - Gestion des erreurs spécifiques et cas d'absence de produits
    
    Réponses:
    - 200: Fichier Excel généré avec succès
    - 400: Erreur de validation
    - 401: Non authentifié
    - 404: Utilisateur ou compte non trouvé
    - 500: Erreur interne du serveur ou erreur de base de données
    """
    permission_classes = [IsAuthenticated]
    
    @swagger_auto_schema(
        operation_summary="Export Excel des produits utilisateur mobile",
        operation_description="Exporte la liste des produits appartenant au même compte de l'utilisateur connecté en format Excel (.xlsx)",
        responses={
            200: openapi.Response(
                description="Fichier Excel généré avec succès",
                content={
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': {
                        'schema': openapi.Schema(
                            type=openapi.TYPE_FILE,
                            description="Fichier Excel contenant les produits"
                        )
                    }
                }
            ),
            400: openapi.Response(
                description="Erreur de validation",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN, example=False),
                        'error': openapi.Schema(type=openapi.TYPE_STRING, example='Erreur de validation'),
                        'error_type': openapi.Schema(type=openapi.TYPE_STRING, example='VALIDATION_ERROR')
                    }
                )
            ),
            401: openapi.Response(
                description="Non authentifié",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'detail': openapi.Schema(type=openapi.TYPE_STRING, example='Authentication credentials were not provided.')
                    }
                )
            ),
            404: openapi.Response(
                description="Utilisateur ou compte non trouvé",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN, example=False),
                        'error': openapi.Schema(type=openapi.TYPE_STRING, example='Utilisateur non trouvé'),
                        'error_type': openapi.Schema(type=openapi.TYPE_STRING, example='USER_NOT_FOUND')
                    }
                )
            ),
            500: openapi.Response(
                description="Erreur interne du serveur",
                schema=openapi.Schema(
                    type=openapi.TYPE_OBJECT,
                    properties={
                        'success': openapi.Schema(type=openapi.TYPE_BOOLEAN, example=False),
                        'error': openapi.Schema(type=openapi.TYPE_STRING, example='Erreur interne du serveur'),
                        'error_type': openapi.Schema(type=openapi.TYPE_STRING, example='INTERNAL_ERROR')
                    }
                )
            )
        },
        security=[{'Bearer': []}],
        tags=['Utilisateur Mobile']
    )
    def get(self, request):
        """
        Exporte les produits du même compte de l'utilisateur connecté en fichier Excel.
        
        Args:
            request: Requête GET
            - L'utilisateur est récupéré automatiquement depuis le token d'authentification
            
        Returns:
            HttpResponse: Fichier Excel contenant les produits
        """
        try:
            # Récupérer l'utilisateur depuis le token d'authentification
            user_id = request.user.id
            print(f"user_id depuis token: {user_id}")
            
            user_service = UserService()
            
            response_data = user_service.get_user_products(user_id)
            
            # Extraire seulement les informations demandées de chaque produit
            products = []
            if response_data and 'data' in response_data and 'products' in response_data['data']:
                for product in response_data['data']['products']:
                    products.append({
                        'web_id': product.get('web_id'),
                        'product_name': product.get('product_name'),
                        'product_code': product.get('product_code'),
                        'internal_product_code': product.get('internal_product_code'),
                        'family_name': product.get('family_name'),
                        'is_variant': product.get('is_variant'),
                        'n_lot': product.get('n_lot'),
                        'n_serie': product.get('n_serie'),
                        'dlc': product.get('dlc'),
                        'numeros_serie': product.get('numeros_serie', []),
                        'created_at': product.get('created_at'),
                        'updated_at': product.get('updated_at')
                    })
            
            # Générer le fichier Excel
            excel_buffer = self._generate_excel(products)
            
            # Définir le nom du fichier avec timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"produits_utilisateur_{user_id}_{timestamp}.xlsx"
            
            # Créer la réponse HTTP avec le fichier Excel
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except UserNotFoundException as e:
            return error_response(
                message=str(e),
                status_code=status.HTTP_404_NOT_FOUND,
                error_type='USER_NOT_FOUND'
            )
        except AccountNotFoundException as e:
            return error_response(
                message=str(e),
                status_code=status.HTTP_404_NOT_FOUND,
                error_type='ACCOUNT_NOT_FOUND'
            )
        except ProductNotFoundException as e:
            # Générer un fichier Excel vide avec seulement les en-têtes
            excel_buffer = self._generate_excel([])
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"produits_utilisateur_{request.user.id}_{timestamp}.xlsx"
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except DataValidationException as e:
            return error_response(
                message=str(e),
                status_code=status.HTTP_400_BAD_REQUEST,
                error_type='VALIDATION_ERROR'
            )
        except DatabaseConnectionException as e:
            return error_response(
                message=str(e),
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                error_type='DATABASE_ERROR'
            )
        except ValueError as e:
            return error_response(
                message=f'Erreur de validation: {str(e)}',
                status_code=status.HTTP_400_BAD_REQUEST,
                error_type='VALIDATION_ERROR'
            )
        except ImportError as e:
            return error_response(
                message=str(e),
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                error_type='DEPENDENCY_ERROR'
            )
        except Exception as e:
            print(f"Erreur inattendue dans UserProductsView: {str(e)}")
            import traceback
            print(f"Traceback complet: {traceback.format_exc()}")
            return error_response(
                message="Erreur interne du serveur",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                error_type='INTERNAL_ERROR'
            )
    
    def _generate_excel(self, products: list) -> io.BytesIO:
        """
        Génère un fichier Excel à partir de la liste des produits.
        
        Args:
            products: Liste des produits à exporter
            
        Returns:
            BytesIO: Buffer contenant le fichier Excel
            
        Raises:
            ImportError: Si openpyxl n'est pas installé
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError(
                "openpyxl est requis pour l'export Excel. "
                "Installez-le avec: pip install openpyxl"
            )
        
        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Produits"
        
        # Définir les en-têtes
        headers = [
            'ID',
            'Nom du produit',
            'Code produit',
            'Code produit interne',
            'Famille',
            'Est variante',
            'N° Lot',
            'N° Série',
            'DLC',
            'Numéros de série',
            'Date de création',
            'Date de mise à jour'
        ]
        
        # Style pour les en-têtes
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Écrire les en-têtes
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Écrire les données
        for row_num, product in enumerate(products, 2):
            # Convertir les valeurs booléennes en texte
            is_variant = "Oui" if product.get('is_variant') else "Non"
            n_lot = "Oui" if product.get('n_lot') else "Non"
            n_serie = "Oui" if product.get('n_serie') else "Non"
            dlc = "Oui" if product.get('dlc') else "Non"
            
            # Convertir la liste des numéros de série en chaîne
            numeros_serie = product.get('numeros_serie', [])
            numeros_serie_str = ', '.join([str(ns) for ns in numeros_serie]) if numeros_serie else ''
            
            # Formater les dates
            created_at = product.get('created_at', '')
            updated_at = product.get('updated_at', '')
            if created_at:
                try:
                    if isinstance(created_at, str):
                        created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M:%S')
                except:
                    pass
            if updated_at:
                try:
                    if isinstance(updated_at, str):
                        updated_at = datetime.fromisoformat(updated_at.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M:%S')
                except:
                    pass
            
            ws.cell(row=row_num, column=1, value=product.get('web_id'))
            ws.cell(row=row_num, column=2, value=product.get('product_name'))
            ws.cell(row=row_num, column=3, value=product.get('product_code'))
            ws.cell(row=row_num, column=4, value=product.get('internal_product_code'))
            ws.cell(row=row_num, column=5, value=product.get('family_name'))
            ws.cell(row=row_num, column=6, value=is_variant)
            ws.cell(row=row_num, column=7, value=n_lot)
            ws.cell(row=row_num, column=8, value=n_serie)
            ws.cell(row=row_num, column=9, value=dlc)
            ws.cell(row=row_num, column=10, value=numeros_serie_str)
            ws.cell(row=row_num, column=11, value=created_at)
            ws.cell(row=row_num, column=12, value=updated_at)
        
        # Ajuster la largeur des colonnes
        column_widths = {
            'A': 10,  # ID
            'B': 30,  # Nom du produit
            'C': 20,  # Code produit
            'D': 20,  # Code produit interne
            'E': 20,  # Famille
            'F': 15,  # Est variante
            'G': 10,  # N° Lot
            'H': 10,  # N° Série
            'I': 10,  # DLC
            'J': 40,  # Numéros de série
            'K': 20,  # Date de création
            'L': 20,  # Date de mise à jour
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Sauvegarder dans un buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
