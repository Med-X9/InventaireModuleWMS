"""
Commande Django pour générer un template Excel pour l'import de produits
Usage: python manage.py generate_product_import_template
"""
from django.core.management.base import BaseCommand
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from django.conf import settings


class Command(BaseCommand):
    help = 'Génère un template Excel pour l\'import de produits'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output',
            type=str,
            default='template_import_produits.xlsx',
            help='Nom du fichier de sortie (défaut: template_import_produits.xlsx)'
        )

    def handle(self, *args, **options):
        output_file = options['output']
        
        # Définir les colonnes selon ProductResource
        columns = [
            'internal product code',  # OBLIGATOIRE - identifiant unique
            'short description',      # Optionnel
            'barcode',                # Optionnel
            'product group',          # Optionnel
            'stock unit',             # Requis dans le modèle
            'product status',         # Optionnel (défaut: ACTIVE)
            'product family',         # OBLIGATOIRE - doit exister dans la base
            'is variant',             # Boolean (True/False, 1/0, Oui/Non)
            'n lot',                  # Boolean (True/False, 1/0, Oui/Non)
            'n serie',                # Boolean
            'dlc',                    # Boolean
        ]
        
        # Données d'exemple
        example_data = [
            {
                'internal product code': 'PROD-001',
                'short description': 'Produit exemple 1',
                'barcode': '1234567890123',
                'product group': 'GRP01',
                'stock unit': 'PIECE',
                'product status': 'ACTIVE',
                'product family': 'Famille Exemple 1',  # ⚠️ Doit exister dans la base
                'is variant': False,
                'n lot': True,
                'n serie': False,
                'dlc': True,
            },
            {
                'internal product code': 'PROD-002',
                'short description': 'Produit exemple 2',
                'barcode': '1234567890124',
                'product group': 'GRP02',
                'stock unit': 'KG',
                'product status': 'ACTIVE',
                'product family': 'Famille Exemple 2',  # ⚠️ Doit exister dans la base
                'is variant': True,
                'n lot': False,
                'n serie': True,
                'dlc': False,
            },
            {
                'internal product code': 'PROD-003',
                'short description': 'Produit exemple 3',
                'barcode': '',
                'product group': '',
                'stock unit': 'LITRE',
                'product status': 'INACTIVE',
                'product family': 'Famille Exemple 1',  # ⚠️ Doit exister dans la base
                'is variant': False,
                'n lot': False,
                'n serie': False,
                'dlc': False,
            },
        ]
        
        # Créer le DataFrame
        df = pd.DataFrame(example_data, columns=columns)
        
        # Sauvegarder en Excel
        df.to_excel(output_file, index=False, sheet_name='Produits')
        
        # Formater le fichier Excel
        wb = load_workbook(output_file)
        ws = wb['Produits']
        
        # Styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        required_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        optional_fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Colonnes obligatoires
        required_columns = ['internal product code', 'product family', 'stock unit']
        
        # Formater les en-têtes
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            
            # Ajouter un indicateur pour les colonnes obligatoires
            if col_name in required_columns:
                cell.comment = f"⚠️ OBLIGATOIRE: {col_name}"
        
        # Formater les cellules de données
        for row_idx in range(2, len(example_data) + 2):
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                
                # Colorer les colonnes obligatoires
                if col_name in required_columns:
                    cell.fill = required_fill
                else:
                    cell.fill = optional_fill
        
        # Ajuster la largeur des colonnes
        for col_idx, col_name in enumerate(columns, start=1):
            max_length = max(
                len(str(col_name)),
                max([len(str(example_data[i].get(col_name, ''))) for i in range(len(example_data))])
            )
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Ajouter une feuille d'instructions
        ws_instructions = wb.create_sheet('Instructions', 0)
        
        instructions = [
            ['INSTRUCTIONS POUR L\'IMPORT DE PRODUITS'],
            [''],
            ['COLONNES OBLIGATOIRES:'],
            ['- internal product code: Code unique du produit (identifiant)'],
            ['- product family: Nom de la famille (doit exister dans la base)'],
            ['- stock unit: Unité de stock (ex: PIECE, KG, LITRE, etc.)'],
            [''],
            ['COLONNES OPTIONNELLES:'],
            ['- short description: Description courte du produit'],
            ['- barcode: Code-barres'],
            ['- product group: Groupe de produit'],
            ['- product status: Statut (ACTIVE, INACTIVE, OBSOLETE) - Défaut: ACTIVE'],
            ['- is variant: Indique si le produit est une variante (True/False, 1/0, Oui/Non)'],
            ['- n lot: Numéro de lot requis (True/False, 1/0, Oui/Non)'],
            ['- n serie: Numéro de série requis (True/False, 1/0, Oui/Non)'],
            ['- dlc: Date limite de consommation (True/False, 1/0, Oui/Non)'],
            [''],
            ['IMPORTANT:'],
            ['1. Le champ "product family" doit correspondre exactement au nom d\'une famille existante'],
            ['2. Le "internal product code" doit être unique'],
            ['3. Les valeurs booléennes acceptent: True/False, 1/0, Oui/Non, Yes/No'],
            ['4. Le statut par défaut est ACTIVE si non spécifié'],
            ['5. L\'import fonctionne en mode "tout ou rien" - une erreur annule tout l\'import'],
            [''],
            ['FORMAT DU FICHIER:'],
            ['- Format Excel (.xlsx) recommandé'],
            ['- La première ligne doit contenir les en-têtes de colonnes'],
            ['- Les lignes suivantes contiennent les données'],
            [''],
            ['EXEMPLES DE VALEURS:'],
            ['- product status: ACTIVE, INACTIVE, OBSOLETE'],
            ['- stock unit: PIECE, KG, LITRE, M2, M3, etc.'],
            ['- is variant / n lot / n serie / dlc: True, False, 1, 0, Oui, Non'],
        ]
        
        for row_idx, instruction in enumerate(instructions, start=1):
            cell = ws_instructions.cell(row=row_idx, column=1)
            cell.value = instruction[0] if instruction else ''
            
            # Formater le titre
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color='366092')
            elif 'OBLIGATOIRES' in str(instruction) or 'OPTIONNELLES' in str(instruction) or 'IMPORTANT' in str(instruction) or 'FORMAT' in str(instruction) or 'EXEMPLES' in str(instruction):
                cell.font = Font(bold=True, size=11)
        
        # Ajuster la largeur de la colonne d'instructions
        ws_instructions.column_dimensions['A'].width = 80
        
        # Sauvegarder
        wb.save(output_file)
        
        self.stdout.write(
            self.style.SUCCESS(
                f'✅ Template Excel créé avec succès: {output_file}'
            )
        )
        self.stdout.write(
            f'📋 Le fichier contient {len(example_data)} lignes d\'exemple'
        )
        self.stdout.write(
            f'📝 N\'oubliez pas de remplacer "Famille Exemple 1" et "Famille Exemple 2" par des familles existantes dans votre base de données'
        )

