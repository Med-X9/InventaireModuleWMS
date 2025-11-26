import os
import pandas as pd
from django.db import transaction
from import_export import resources, exceptions
from ..models import ImportTask, ImportError, Product, Family
import logging
from datetime import datetime
import tempfile

logger = logging.getLogger(__name__)


class ProductImportService:
    """Service pour importer des produits en mode 'tout ou rien'"""
    
    CHUNK_SIZE = 10000  # 10k lignes par chunk
    
    def __init__(self, resource_class):
        self.resource_class = resource_class
        self.family_cache = {}
        self._load_family_cache()
    
    def _load_family_cache(self):
        """Charge toutes les familles en mémoire"""
        self.family_cache = {
            f.family_name: f 
            for f in Family.objects.only('id', 'family_name')
        }
        logger.info(f"Cache de {len(self.family_cache)} familles chargé")
    
    def process_file_chunked(self, file_path: str, import_task: ImportTask):
        """
        Traite un fichier en mode 'tout ou rien'
        
        LOGIQUE:
        1. Valide TOUTES les lignes AVANT d'importer
        2. Détecte les doublons DANS le fichier
        3. Si une seule erreur → ANNULER (aucun import)
        4. Si toutes valides → Importer TOUT dans une transaction
        5. Si erreur pendant import → Rollback automatique
        """
        try:
            logger.info(f"=== DÉBUT IMPORT TÂCHE {import_task.id} ===")
            logger.info(f"Fichier: {file_path}")
            logger.info(f"Statut initial: {import_task.status}")
            
            # PHASE 1: VALIDATION COMPLÈTE
            try:
                import_task.status = 'VALIDATING'
                import_task.save()
                logger.info(f"Statut mis à VALIDATING pour l'import {import_task.id}")
                
                total_rows = self._count_file_rows(file_path)
                import_task.total_rows = total_rows
                import_task.save()
                logger.info(f"Nombre total de lignes détecté: {total_rows}")
                
                logger.info(f"Début validation complète ({total_rows} lignes)")
                
                # Lire le fichier Excel complet
                logger.info(f"Lecture du fichier Excel: {file_path}")
                df_full = pd.read_excel(file_path, engine='openpyxl')
                logger.info(f"Fichier Excel lu: {len(df_full)} lignes")
                
                # PHASE 1.1: Détecter les doublons DANS le fichier
                logger.info("Début détection des doublons dans le fichier")
                duplicate_errors = self._detect_duplicates_in_file(df_full)
                logger.info(f"Détection des doublons terminée: {len(duplicate_errors)} doublon(s) trouvé(s)")
                validation_errors = []
                
                # Convertir les erreurs de doublons en ImportError
                for dup_error in duplicate_errors:
                    error_obj = ImportError(
                        import_task=import_task,
                        row_number=dup_error['row_number'],
                        error_type=dup_error['error_type'],
                        error_message=dup_error['error_message'],
                        field_name=dup_error.get('field_name'),
                        field_value=dup_error.get('field_value'),
                        row_data=dup_error.get('row_data', {}),
                    )
                    validation_errors.append(error_obj)
                
                # PHASE 1.2: Valider toutes les lignes (traiter par chunks)
                logger.info("Début validation des lignes par chunks")
                row_number = 2  # Ligne 1 = header
                total_chunks = (len(df_full) + self.CHUNK_SIZE - 1) // self.CHUNK_SIZE
                logger.info(f"Validation de {total_chunks} chunk(s)")
                
                for chunk_idx, i in enumerate(range(0, len(df_full), self.CHUNK_SIZE), 1):
                    logger.info(f"Validation chunk {chunk_idx}/{total_chunks} (lignes {i} à {min(i + self.CHUNK_SIZE, len(df_full))})")
                    chunk_df = df_full.iloc[i:i + self.CHUNK_SIZE]
                    chunk_errors = self._validate_chunk(chunk_df, import_task, row_number)
                    validation_errors.extend(chunk_errors)
                    
                    import_task.validated_rows = min(row_number + len(chunk_df) - 1, total_rows)
                    import_task.save()
                    logger.debug(f"Progression validation: {import_task.validated_rows}/{total_rows}")
                    
                    row_number += len(chunk_df)
                
                logger.info(f"Validation terminée: {len(validation_errors)} erreur(s) détectée(s)")
                
            except Exception as validation_exception:
                # Erreur spécifique pendant la validation
                logger.error(f"Erreur pendant la phase de validation: {str(validation_exception)}", exc_info=True)
                import_task.status = 'FAILED'
                import_task.error_message = (
                    f"Erreur lors de la validation: {str(validation_exception)} "
                    f"(Type: {type(validation_exception).__name__})"
                )
                import_task.save()
                logger.error(f"Statut mis à FAILED pour l'import {import_task.id} (erreur validation)")
                raise
            
            # VÉRIFICATION: Si erreurs → ANNULER TOUT
            if validation_errors:
                import_task.status = 'CANCELLED'
                import_task.error_count = len(validation_errors)
                import_task.error_message = (
                    f"Import annulé: {len(validation_errors)} erreur(s) détectée(s) "
                    f"lors de la validation. Aucun produit n'a été importé."
                )
                
                # Sauvegarder les erreurs
                ImportError.objects.bulk_create(validation_errors, batch_size=1000)
                
                # Générer fichier d'erreurs
                errors_file_path = self._generate_errors_file(import_task, validation_errors)
                import_task.errors_file_path = errors_file_path
                import_task.save()
                
                logger.warning(f"Import annulé: {len(validation_errors)} erreurs détectées")
                return  # ⭐ ARRÊT: Rien n'est importé
            
            # PHASE 2: IMPORT (seulement si validation OK)
            logger.info("Validation OK, début import dans transaction")
            import_task.status = 'PROCESSING'
            import_task.save()
            logger.info(f"Statut mis à jour à PROCESSING pour l'import {import_task.id}")
            
            # ⭐ IMPORT DANS UNE TRANSACTION (tout ou rien)
            try:
                logger.info(f"Début de la transaction d'import pour {total_rows} lignes")
                with transaction.atomic():  # Transaction globale
                    total_imported = 0
                    total_updated = 0
                    
                    row_number = 2
                    # Lire le fichier Excel et traiter par chunks
                    logger.info(f"Lecture du fichier Excel: {file_path}")
                    df_full = pd.read_excel(file_path, engine='openpyxl')
                    logger.info(f"Fichier Excel lu: {len(df_full)} lignes")
                    
                    # Traiter par chunks
                    total_chunks = (len(df_full) + self.CHUNK_SIZE - 1) // self.CHUNK_SIZE
                    logger.info(f"Début traitement de {total_chunks} chunks")
                    
                    for chunk_idx, i in enumerate(range(0, len(df_full), self.CHUNK_SIZE), 1):
                        logger.info(f"Traitement chunk {chunk_idx}/{total_chunks} (lignes {i} à {min(i + self.CHUNK_SIZE, len(df_full))})")
                        chunk_df = df_full.iloc[i:i + self.CHUNK_SIZE]
                        chunk_data = self._dataframe_to_dataset(chunk_df)
                        logger.info(f"Chunk converti en dataset: {len(chunk_data)} lignes")
                        
                        result = self._import_chunk(chunk_data, import_task, row_number)
                        logger.info(f"Chunk importé: {result.get('imported', 0)} importés, {result.get('updated', 0)} mis à jour")
                        
                        total_imported += result.get('imported', 0)
                        total_updated += result.get('updated', 0)
                        
                        import_task.processed_rows = min(row_number + len(chunk_df) - 1, total_rows)
                        import_task.imported_count = total_imported
                        import_task.updated_count = total_updated
                        import_task.save()
                        logger.info(f"Progression sauvegardée: {import_task.processed_rows}/{total_rows}")
                        
                        row_number += len(chunk_df)
                    
                    # Succès complet
                    logger.info(f"Tous les chunks traités. Finalisation...")
                    import_task.status = 'COMPLETED'
                    import_task.processed_rows = total_rows
                    import_task.save()
                    
                    logger.info(
                        f"Import terminé avec succès: {total_imported} importés, "
                        f"{total_updated} mis à jour"
                    )
                    
            except Exception as e:
                # ⭐ Erreur pendant import → rollback automatique
                logger.error(f"Exception lors de l'import: {str(e)}", exc_info=True)
                import_task.status = 'FAILED'
                import_task.error_message = (
                    f"Erreur lors de l'import: {str(e)}. "
                    f"Aucun produit n'a été importé (rollback automatique)."
                )
                import_task.save()
                logger.error(f"Statut mis à jour à FAILED pour l'import {import_task.id}")
                raise  # Transaction annulée automatiquement
            
        except Exception as e:
            logger.error(f"=== ERREUR CRITIQUE IMPORT TÂCHE {import_task.id} ===", exc_info=True)
            if import_task.status not in ['CANCELLED', 'FAILED']:
                import_task.status = 'FAILED'
                import_task.error_message = f"{str(e)} (Type: {type(e).__name__})"
                import_task.save()
                logger.error(f"Statut mis à jour à FAILED pour l'import {import_task.id}")
            logger.error(f"=== FIN IMPORT TÂCHE {import_task.id} (ÉCHEC) ===")
            raise
        finally:
            # S'assurer que le statut est toujours mis à jour même en cas d'erreur silencieuse
            try:
                import_task.refresh_from_db()
                if import_task.status == 'VALIDATING':
                    logger.warning(f"⚠️ Statut toujours VALIDATING après traitement, mise à jour à FAILED")
                    import_task.status = 'FAILED'
                    import_task.error_message = "Erreur silencieuse: le traitement s'est arrêté après la validation."
                    import_task.save()
            except Exception as e:
                logger.error(f"Erreur lors de la mise à jour finale du statut: {str(e)}")
    
    def _detect_duplicates_in_file(self, df: pd.DataFrame) -> list:
        """Détecte les doublons dans le fichier Excel lui-même"""
        errors = []
        
        try:
            if df.empty:
                logger.warning("Le DataFrame est vide, aucun doublon à détecter")
                return errors
            
            # Normaliser les colonnes (gérer les variations de casse et espaces)
            df.columns = df.columns.str.strip().str.lower()
            
            # Vérifier les doublons de Internal_Product_Code dans le fichier
            if 'internal product code' in df.columns:
                try:
                    internal_code_col = 'internal product code'
                    df[internal_code_col] = df[internal_code_col].astype(str).str.strip()
                    
                    # Trouver tous les codes dupliqués
                    value_counts = df[internal_code_col].value_counts()
                    duplicates = value_counts[value_counts > 1].index.tolist()
                    
                    for internal_code in duplicates:
                        if internal_code and internal_code != 'nan' and internal_code.lower() != 'nan':
                            # Trouver toutes les lignes avec ce code
                            duplicate_indices = df[df[internal_code_col] == internal_code].index.tolist()
                            row_numbers = [r + 2 for r in duplicate_indices]  # +2 car ligne 1 = header, index 0-based
                            
                            # Créer une erreur pour chaque ligne dupliquée
                            for idx in duplicate_indices:
                                try:
                                    row = df.iloc[idx]
                                    errors.append({
                                        'row_number': idx + 2,
                                        'error_type': 'DUPLICATE_ERROR',
                                        'error_message': f"Le code produit interne '{internal_code}' apparaît plusieurs fois dans le fichier (lignes: {', '.join(map(str, row_numbers))}).",
                                        'field_name': 'internal product code',
                                        'field_value': internal_code,
                                        'row_data': row.to_dict(),
                                    })
                                except Exception as e:
                                    logger.error(f"Erreur lors de la détection de doublon pour la ligne {idx + 2}: {str(e)}")
                except Exception as e:
                    logger.error(f"Erreur lors de la détection des doublons de code produit interne: {str(e)}", exc_info=True)
            
            # Vérifier les doublons de Barcode dans le fichier (si présent et non vide)
            if 'barcode' in df.columns:
                try:
                    barcode_col = 'barcode'
                    df[barcode_col] = df[barcode_col].astype(str).str.strip()
                    # Filtrer les valeurs vides et 'nan'
                    df_barcode = df[(df[barcode_col] != '') & (df[barcode_col] != 'nan') & (df[barcode_col].notna())]
                    
                    # Trouver tous les barcodes dupliqués
                    if not df_barcode.empty:
                        value_counts = df_barcode[barcode_col].value_counts()
                        duplicates = value_counts[value_counts > 1].index.tolist()
                        
                        for barcode in duplicates:
                            if barcode and barcode != 'nan' and barcode.lower() != 'nan':
                                # Trouver toutes les lignes avec ce barcode
                                duplicate_indices = df_barcode[df_barcode[barcode_col] == barcode].index.tolist()
                                row_numbers = [r + 2 for r in duplicate_indices]
                                
                                # Créer une erreur pour chaque ligne dupliquée
                                for idx in duplicate_indices:
                                    try:
                                        row = df_barcode.iloc[idx]
                                        errors.append({
                                            'row_number': idx + 2,
                                            'error_type': 'DUPLICATE_ERROR',
                                            'error_message': f"Le code-barres '{barcode}' apparaît plusieurs fois dans le fichier (lignes: {', '.join(map(str, row_numbers))}).",
                                            'field_name': 'barcode',
                                            'field_value': barcode,
                                            'row_data': row.to_dict(),
                                        })
                                    except Exception as e:
                                        logger.error(f"Erreur lors de la détection de doublon pour la ligne {idx + 2}: {str(e)}")
                except Exception as e:
                    logger.error(f"Erreur lors de la détection des doublons de barcode: {str(e)}", exc_info=True)
        
        except Exception as e:
            logger.error(f"Erreur critique lors de la détection des doublons: {str(e)}", exc_info=True)
        
        return errors
    
    def _validate_chunk(self, df: pd.DataFrame, import_task: ImportTask, start_row: int) -> list:
        """Valide un chunk et retourne la liste des erreurs"""
        errors = []
        
        # Utiliser un compteur séquentiel au lieu de l'index du DataFrame
        # pour garantir que row_number est correct même si le DataFrame a un index non séquentiel
        row_offset = 0
        for index, row_data in df.iterrows():
            row_number = start_row + row_offset
            row_dict = row_data.to_dict()
            
            try:
                validation_error = self._validate_row(row_dict, row_number)
                if validation_error:
                    # Si c'est un dict d'erreur (format standard)
                    if isinstance(validation_error, dict):
                        error_obj = ImportError(
                            import_task=import_task,
                            row_number=row_number,
                            error_type=validation_error['error_type'],
                            error_message=validation_error['error_message'],
                            field_name=validation_error.get('field_name'),
                            field_value=validation_error.get('field_value'),
                            row_data=row_dict,
                        )
                        errors.append(error_obj)
                    # Si c'est déjà un ImportError (depuis _detect_duplicates_in_file)
                    elif hasattr(validation_error, 'row_number'):
                        validation_error.import_task = import_task
                        errors.append(validation_error)
            except Exception as e:
                # Gérer les erreurs lors de la validation d'une ligne
                logger.error(f"Erreur lors de la validation de la ligne {row_number}: {str(e)}", exc_info=True)
                error_obj = ImportError(
                    import_task=import_task,
                    row_number=row_number,
                    error_type='VALIDATION_ERROR',
                    error_message=f"Erreur lors de la validation: {str(e)}",
                    field_name=None,
                    field_value=None,
                    row_data=row_dict,
                )
                errors.append(error_obj)
            
            row_offset += 1
        
        return errors
    
    def _validate_row(self, row_data: dict, row_number: int) -> dict:
        """Valide une ligne et retourne un dict d'erreur si invalide"""
        try:
            # Helper pour convertir en string de manière sécurisée
            def safe_str(value, default=''):
                if value is None:
                    return default
                try:
                    return str(value).strip()
                except Exception:
                    return default
            
            # Validation famille
            family_name = safe_str(row_data.get('product family', ''))
            if not family_name:
                return {
                    'error_type': 'VALIDATION_ERROR',
                    'error_message': "La colonne 'product family' est obligatoire.",
                    'field_name': 'product family',
                    'field_value': family_name,
                }
            
            if family_name not in self.family_cache:
                return {
                    'error_type': 'VALIDATION_ERROR',
                    'error_message': f"La famille '{family_name}' n'existe pas dans la base de données.",
                    'field_name': 'product family',
                    'field_value': family_name,
                }
            
            # Validation barcode
            barcode = safe_str(row_data.get('barcode', ''))
            if not barcode:
                return {
                    'error_type': 'VALIDATION_ERROR',
                    'error_message': "La colonne 'barcode' est obligatoire.",
                    'field_name': 'barcode',
                    'field_value': barcode,
                }
            
            # Validation code produit interne
            internal_code = safe_str(row_data.get('internal product code', ''))
            if not internal_code:
                return {
                    'error_type': 'VALIDATION_ERROR',
                    'error_message': "La colonne 'internal product code' est obligatoire.",
                    'field_name': 'internal product code',
                    'field_value': internal_code,
                }
            
            # Vérifier unicité code produit interne dans la base de données
            try:
                if Product.objects.filter(Internal_Product_Code=internal_code).exists():
                    return {
                        'error_type': 'VALIDATION_ERROR',
                        'error_message': f"Le code produit interne '{internal_code}' existe déjà dans la base de données.",
                        'field_name': 'internal product code',
                        'field_value': internal_code,
                    }
            except Exception as e:
                logger.error(f"Erreur lors de la vérification de l'unicité du code produit interne '{internal_code}': {str(e)}")
                return {
                    'error_type': 'VALIDATION_ERROR',
                    'error_message': f"Erreur lors de la vérification de l'unicité du code produit interne: {str(e)}",
                    'field_name': 'internal product code',
                    'field_value': internal_code,
                }
            
            # Vérifier unicité barcode dans la base de données (si présent)
            if barcode:
                try:
                    existing_product = Product.objects.filter(Barcode=barcode).first()
                    if existing_product:
                        return {
                            'error_type': 'VALIDATION_ERROR',
                            'error_message': f"Le code-barres '{barcode}' existe déjà dans la base de données (produit: {existing_product.Short_Description}).",
                            'field_name': 'barcode',
                            'field_value': barcode,
                        }
                except Exception as e:
                    logger.error(f"Erreur lors de la vérification de l'unicité du barcode '{barcode}': {str(e)}")
                    return {
                        'error_type': 'VALIDATION_ERROR',
                        'error_message': f"Erreur lors de la vérification de l'unicité du code-barres: {str(e)}",
                        'field_name': 'barcode',
                        'field_value': barcode,
                    }
            
            # Validation description
            short_desc = safe_str(row_data.get('short description', ''))
            if not short_desc:
                return {
                    'error_type': 'VALIDATION_ERROR',
                    'error_message': "La colonne 'short description' est obligatoire.",
                    'field_name': 'short description',
                    'field_value': short_desc,
                }
            
            # Validation unité de stock
            stock_unit = safe_str(row_data.get('stock unit', ''))
            if not stock_unit:
                return {
                    'error_type': 'VALIDATION_ERROR',
                    'error_message': "La colonne 'stock unit' est obligatoire.",
                    'field_name': 'stock unit',
                    'field_value': stock_unit,
                }
            
            return None  # Pas d'erreur
            
        except Exception as e:
            # Erreur inattendue lors de la validation
            logger.error(f"Erreur inattendue lors de la validation de la ligne {row_number}: {str(e)}", exc_info=True)
            return {
                'error_type': 'VALIDATION_ERROR',
                'error_message': f"Erreur inattendue lors de la validation: {str(e)}",
                'field_name': None,
                'field_value': None,
            }
    
    def _import_chunk(self, dataset, import_task: ImportTask, start_row: int):
        """Importe un chunk dans la transaction"""
        resource = self.resource_class()
        imported = 0
        updated = 0
        
        logger.info(f"Début import chunk: {len(dataset)} lignes à partir de la ligne {start_row}")
        
        for index, row_data in enumerate(dataset.dict):
            try:
                # Normaliser les valeurs booléennes avant l'import
                boolean_fields = ['is variant', 'n lot', 'n serie', 'dlc']
                for field in boolean_fields:
                    if field in row_data:
                        value = row_data[field]
                        # Normaliser: None, '', 'None', 'null' -> False
                        if value is None or value == '' or str(value).strip().lower() in ('none', 'null', 'n/a', 'na'):
                            row_data[field] = False
                        elif isinstance(value, bool):
                            row_data[field] = value
                        elif isinstance(value, str):
                            value_lower = value.strip().lower()
                            if value_lower in ('true', '1', 'yes', 'oui', 'o', 'y', 't'):
                                row_data[field] = True
                            elif value_lower in ('false', '0', 'no', 'non', 'n', 'f'):
                                row_data[field] = False
                            else:
                                row_data[field] = False
                        elif isinstance(value, (int, float)):
                            row_data[field] = bool(value)
                        else:
                            row_data[field] = False
                    else:
                        # Si la colonne n'existe pas, définir à False
                        row_data[field] = False
                
                # Créer un dataset avec une seule ligne
                from tablib import Dataset
                single_row_dataset = Dataset()
                
                # S'assurer que les headers sont définis
                if not single_row_dataset.headers and dataset.headers:
                    single_row_dataset.headers = dataset.headers
                
                # Ajouter la ligne avec les valeurs normalisées
                row_values = [row_data.get(header, '') for header in dataset.headers]
                single_row_dataset.append(row_values)
                
                logger.debug(f"Import ligne {start_row + index}: {row_data.get('internal product code', 'N/A')}")
                
                result = resource.import_data(
                    single_row_dataset,
                    dry_run=False,
                    raise_errors=True,  # ⭐ Lever exception si erreur
                    use_transactions=False,  # Transaction gérée au niveau supérieur
                )
                
                if result.totals.get('new', 0) > 0:
                    imported += 1
                elif result.totals.get('update', 0) > 0:
                    updated += 1
                
            except Exception as e:
                # ⭐ Toute erreur annule la transaction
                logger.error(f"Erreur ligne {start_row + index}: {str(e)}", exc_info=True)
                raise exceptions.ImportError(
                    f"Erreur ligne {start_row + index}: {str(e)}"
                )
        
        logger.info(f"Chunk importé: {imported} importés, {updated} mis à jour")
        return {
            'imported': imported,
            'updated': updated,
        }
    
    def _count_file_rows(self, file_path: str) -> int:
        """Compte le nombre de lignes dans le fichier"""
        try:
            # Lire uniquement pour compter les lignes (sans charger toutes les données)
            df = pd.read_excel(file_path, engine='openpyxl', nrows=0)  # Lire seulement les headers
            # Utiliser openpyxl directement pour compter les lignes sans charger tout en mémoire
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            # Compter les lignes non vides (sans la ligne d'en-tête)
            row_count = sum(1 for row in ws.iter_rows(min_row=2) if any(cell.value for cell in row))
            wb.close()
            return row_count
        except Exception as e:
            logger.warning(f"Impossible de compter les lignes avec openpyxl, tentative avec pandas: {str(e)}")
            try:
                # Fallback: lire le fichier complet (peut être lent pour gros fichiers)
                df = pd.read_excel(file_path, engine='openpyxl')
                return len(df)
            except Exception as e2:
                logger.error(f"Impossible de compter les lignes: {str(e2)}")
                return 0
    
    def _dataframe_to_dataset(self, df):
        """Convertit un DataFrame pandas en Dataset import_export"""
        from tablib import Dataset
        
        # Convertir DataFrame en liste de dictionnaires
        records = df.to_dict('records')
        
        # Créer le Dataset avec les en-têtes et les données
        if len(records) > 0:
            dataset = Dataset()
            # Définir les en-têtes depuis les clés du premier enregistrement
            dataset.headers = list(records[0].keys())
            # Ajouter les données
            for record in records:
                dataset.append([record.get(header, '') for header in dataset.headers])
        else:
            dataset = Dataset()
        
        return dataset
    
    def _generate_errors_file(self, import_task: ImportTask, errors: list) -> str:
        """Génère un fichier Excel avec toutes les erreurs"""
        error_data = []
        for error in errors:
            error_data.append({
                'Ligne': error.row_number,
                'Type d\'erreur': error.error_type,
                'Champ concerné': error.field_name or '',
                'Valeur du champ': str(error.field_value) if error.field_value else '',
                'Message d\'erreur': error.error_message,
                'Barcode': error.row_data.get('barcode', ''),
                'Code produit interne': error.row_data.get('internal product code', ''),
                'Description': error.row_data.get('short description', ''),
            })
        
        df_errors = pd.DataFrame(error_data)
        
        temp_dir = tempfile.gettempdir()
        errors_file_path = os.path.join(
            temp_dir,
            f'errors_import_{import_task.id}_{int(datetime.now().timestamp())}.xlsx'
        )
        
        df_errors.to_excel(errors_file_path, index=False, engine='openpyxl')
        logger.info(f"Fichier d'erreurs généré: {errors_file_path}")
        return errors_file_path
