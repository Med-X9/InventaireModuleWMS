"""
Service pour l'import/export des utilisateurs
Contient la logique métier
"""
from typing import List, Dict, Any, Optional
import logging
from io import BytesIO
from ..repositories.user_import_export_repository import UserImportExportRepository
from django.contrib.auth import get_user_model
from apps.masterdata.models import Account

User = get_user_model()
logger = logging.getLogger(__name__)


class UserImportExportService:
    """
    Service pour l'import/export des utilisateurs
    Contient la logique métier
    """
    
    def __init__(self):
        self.repository = UserImportExportRepository()
    
    def export_users_to_excel(self, account_id: Optional[int] = None) -> BytesIO:
        """
        Exporte les utilisateurs vers un fichier Excel
        
        Args:
            account_id: ID du compte (optionnel). Si fourni, exporte uniquement les utilisateurs du compte
            
        Returns:
            BytesIO: Le contenu du fichier Excel en mémoire
            
        Raises:
            ValueError: Si une erreur survient lors de l'export
        """
        try:
            # Vérifier que pandas et openpyxl sont disponibles
            try:
                import pandas as pd
            except ImportError:
                raise ValueError(
                    "pandas est requis pour l'export Excel. "
                    "Installez-le avec: pip install pandas"
                )
            
            try:
                import openpyxl
            except ImportError:
                raise ValueError(
                    "openpyxl est requis pour l'export Excel. "
                    "Installez-le avec: pip install openpyxl"
                )
            
            # Récupérer les utilisateurs
            if account_id:
                users = self.repository.get_users_by_account(account_id)
            else:
                users = self.repository.get_all_users()
            
            if not users:
                raise ValueError("Aucun utilisateur trouvé à exporter")
            
            # Construire les données pour le DataFrame
            rows = []
            for user in users:
                row = {
                    'Nom d\'utilisateur': user.username,
                    'Email': user.email or '',
                    'Nom': user.nom,
                    'Prénom': user.prenom,
                    'Type': user.type,
                    'Compte': user.compte.account_name if user.compte else '',
                    'Actif': 'Oui' if user.is_active else 'Non',
                    'Administrateur': 'Oui' if user.is_staff else 'Non',
                }
                rows.append(row)
            
            # Créer le DataFrame
            df = pd.DataFrame(rows)
            
            # Générer le fichier Excel
            excel_buffer = BytesIO()
            
            try:
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Utilisateurs')
                    
                    # Récupérer la feuille pour appliquer des formats
                    worksheet = writer.sheets['Utilisateurs']
                    
                    # Appliquer un style aux en-têtes
                    from openpyxl.styles import Font, PatternFill, Alignment
                    from openpyxl.utils import get_column_letter
                    
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Ajuster la largeur des colonnes
                    for idx, col in enumerate(df.columns, start=1):
                        try:
                            col_data = df[col].fillna('').astype(str)
                            
                            if len(col_data) > 0:
                                max_data_length = col_data.map(len).max()
                            else:
                                max_data_length = 0
                            
                            max_length = max(max_data_length, len(str(col)))
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
                        except Exception:
                            worksheet.column_dimensions[get_column_letter(idx)].width = 15
                    
                    # Figer la première ligne (en-têtes)
                    worksheet.freeze_panes = 'A2'
                    
            except Exception as e:
                logger.error(f"Erreur lors de la génération du fichier Excel: {str(e)}", exc_info=True)
                raise ValueError(f"Impossible de générer le fichier Excel: {str(e)}")
            
            excel_buffer.seek(0)
            logger.info(f"Fichier Excel généré avec succès: {len(users)} utilisateurs")
            return excel_buffer
            
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Erreur inattendue lors de l'export des utilisateurs: {str(e)}", exc_info=True)
            raise ValueError(f"Erreur lors de l'export des utilisateurs: {str(e)}")
    
    def import_users_from_excel(self, excel_file, account_id: Optional[int] = None) -> Dict[str, Any]:
        """
        Importe des utilisateurs depuis un fichier Excel
        
        Args:
            excel_file: Fichier Excel à importer
            account_id: ID du compte (optionnel). Si fourni, assigne tous les utilisateurs à ce compte
            
        Returns:
            Dictionnaire contenant les résultats de l'importation
            
        Raises:
            ValueError: Si le fichier est invalide ou si une erreur survient
        """
        try:
            # Vérifier que pandas est disponible
            try:
                import pandas as pd
            except ImportError:
                raise ValueError(
                    "pandas est requis pour l'import Excel. "
                    "Installez-le avec: pip install pandas"
                )
            
            # Lire le fichier Excel
            try:
                df = pd.read_excel(excel_file)
            except Exception as e:
                raise ValueError(f"Impossible de lire le fichier Excel: {str(e)}")
            
            # Colonnes requises
            required_columns = ['Nom d\'utilisateur', 'Nom', 'Prénom', 'Type']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                raise ValueError(
                    f"Colonnes manquantes dans le fichier Excel: {', '.join(missing_columns)}"
                )
            
            # Vérifier le compte si fourni
            compte = None
            if account_id:
                try:
                    compte = Account.objects.get(id=account_id)
                except Account.DoesNotExist:
                    raise ValueError(f"Compte avec l'ID {account_id} non trouvé")
            
            # Résultats de l'importation
            results = {
                'total_rows': len(df),
                'success': 0,
                'errors': 0,
                'created': 0,
                'updated': 0,
                'errors_details': []
            }
            
            # Traiter chaque ligne
            for index, row in df.iterrows():
                try:
                    # Préparer les données de l'utilisateur
                    user_data = {
                        'username': str(row['Nom d\'utilisateur']).strip(),
                        'nom': str(row['Nom']).strip(),
                        'prenom': str(row['Prénom']).strip(),
                        'type': str(row['Type']).strip(),
                    }
                    
                    # Validation du username
                    if not user_data['username']:
                        results['errors'] += 1
                        results['errors_details'].append({
                            'row': index + 2,  # +2 car Excel commence à 1 et on a l'en-tête
                            'username': user_data['username'],
                            'error': 'Le nom d\'utilisateur est requis'
                        })
                        continue
                    
                    # Validation du type
                    valid_types = ['Web', 'Mobile']
                    if user_data['type'] not in valid_types:
                        results['errors'] += 1
                        results['errors_details'].append({
                            'row': index + 2,
                            'username': user_data['username'],
                            'error': f"Type invalide. Les types valides sont: {', '.join(valid_types)}"
                        })
                        continue
                    
                    # Champs optionnels
                    if 'Email' in df.columns and pd.notna(row.get('Email')):
                        user_data['email'] = str(row['Email']).strip()
                    
                    if account_id or ('Compte' in df.columns and pd.notna(row.get('Compte'))):
                        if account_id:
                            user_data['compte'] = compte
                        elif 'Compte' in df.columns and pd.notna(row.get('Compte')):
                            compte_nom = str(row['Compte']).strip()
                            try:
                                user_data['compte'] = Account.objects.get(account_name=compte_nom)
                            except Account.DoesNotExist:
                                results['errors'] += 1
                                results['errors_details'].append({
                                    'row': index + 2,
                                    'username': user_data['username'],
                                    'error': f"Compte '{compte_nom}' non trouvé"
                                })
                                continue
                    
                    if 'Actif' in df.columns and pd.notna(row.get('Actif')):
                        actif_value = str(row['Actif']).strip().lower()
                        user_data['is_active'] = actif_value in ['oui', 'yes', 'true', '1', 'o']
                    
                    if 'Administrateur' in df.columns and pd.notna(row.get('Administrateur')):
                        admin_value = str(row['Administrateur']).strip().lower()
                        user_data['is_staff'] = admin_value in ['oui', 'yes', 'true', '1', 'o']
                    
                    # Vérifier si l'utilisateur existe déjà
                    existing_user = self.repository.get_user_by_username(user_data['username'])
                    
                    if existing_user:
                        # Mise à jour de l'utilisateur existant
                        self.repository.update_user(existing_user, user_data)
                        results['updated'] += 1
                        logger.info(f"Utilisateur mis à jour: {user_data['username']}")
                    else:
                        # Création d'un nouvel utilisateur
                        # Mot de passe par défaut si non fourni
                        if 'Mot de passe' in df.columns and pd.notna(row.get('Mot de passe')):
                            user_data['password'] = str(row['Mot de passe']).strip()
                        else:
                            # Générer un mot de passe par défaut
                            import secrets
                            import string
                            alphabet = string.ascii_letters + string.digits
                            user_data['password'] = ''.join(secrets.choice(alphabet) for i in range(12))
                        
                        self.repository.create_user(user_data)
                        results['created'] += 1
                        logger.info(f"Utilisateur créé: {user_data['username']}")
                    
                    results['success'] += 1
                    
                except Exception as e:
                    results['errors'] += 1
                    username = str(row.get('Nom d\'utilisateur', 'N/A')).strip()
                    results['errors_details'].append({
                        'row': index + 2,
                        'username': username,
                        'error': str(e)
                    })
                    logger.error(f"Erreur lors du traitement de la ligne {index + 2}: {str(e)}", exc_info=True)
            
            logger.info(
                f"Import terminé: {results['success']} succès, "
                f"{results['errors']} erreurs ({results['created']} créés, {results['updated']} mis à jour)"
            )
            
            return results
            
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"Erreur inattendue lors de l'import des utilisateurs: {str(e)}", exc_info=True)
            raise ValueError(f"Erreur lors de l'import des utilisateurs: {str(e)}")

