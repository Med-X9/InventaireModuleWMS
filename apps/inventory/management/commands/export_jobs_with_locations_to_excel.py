"""
Commande Django pour exporter tous les jobs avec leurs emplacements liés dans un fichier Excel.

Format Excel généré:
- job_id: ID du job
- job_reference: Référence du job (ex: JOB-0001)
- job_status: Statut du job
- inventory_reference: Référence de l'inventaire
- warehouse_name: Nom de l'entrepôt
- location_id: ID de l'emplacement
- location_reference: Référence de l'emplacement
- location_sous_zone: Nom de la sous-zone
- location_zone: Nom de la zone
- job_detail_reference: Référence du JobDetail
- job_detail_status: Statut du JobDetail
- counting_order: Ordre du comptage (si applicable)

Exemple d'utilisation:
    python manage.py export_jobs_with_locations_to_excel --file path/to/export.xlsx
    python manage.py export_jobs_with_locations_to_excel --file path/to/export.xlsx --inventory-id 1
    python manage.py export_jobs_with_locations_to_excel --file path/to/export.xlsx --job-id 123
"""
from django.core.management.base import BaseCommand, CommandError
from django.db.models import Prefetch
import logging
from typing import Optional

from apps.inventory.models import Job, JobDetail
from apps.masterdata.models import Location

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Exporte tous les jobs avec leurs emplacements liés dans un fichier Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            required=True,
            help='Chemin vers le fichier Excel de sortie (.xlsx)',
        )
        parser.add_argument(
            '--inventory-id',
            type=int,
            help='ID de l\'inventaire spécifique (optionnel)',
        )
        parser.add_argument(
            '--job-id',
            type=int,
            help='ID du job spécifique (optionnel)',
        )
        parser.add_argument(
            '--warehouse-id',
            type=int,
            help='ID de l\'entrepôt spécifique (optionnel)',
        )

    def handle(self, *args, **options):
        file_path = options['file']
        inventory_id = options.get('inventory_id')
        job_id = options.get('job_id')
        warehouse_id = options.get('warehouse_id')
        
        self.stdout.write(self.style.SUCCESS(f'📊 Export des jobs avec emplacements vers Excel: {file_path}'))
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise CommandError(
                "openpyxl n'est pas installé. Installez-le avec: pip install openpyxl"
            )
        
        # Récupérer les jobs avec leurs JobDetails et emplacements optimisés
        queryset = Job.objects.select_related(
            'inventory',
            'warehouse'
        ).prefetch_related(
            Prefetch(
                'jobdetail_set',
                queryset=JobDetail.objects.select_related(
                    'location',
                    'location__sous_zone',
                    'location__sous_zone__zone',
                    'counting'
                ).order_by('location__location_reference')
            )
        ).order_by('reference')
        
        # Appliquer les filtres
        if job_id:
            queryset = queryset.filter(id=job_id)
        
        if inventory_id:
            queryset = queryset.filter(inventory_id=inventory_id)
        
        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)
        
        jobs = queryset.all()
        total_jobs = jobs.count()
        
        self.stdout.write(f'  📦 Nombre de jobs à exporter: {total_jobs}')
        
        if total_jobs == 0:
            self.stdout.write(self.style.WARNING('⚠️  Aucun job trouvé'))
            return
        
        # Compter le total de lignes (jobs + emplacements)
        total_rows = 0
        for job in jobs:
            total_rows += job.jobdetail_set.count() or 1  # Au moins une ligne par job même sans emplacement
        
        self.stdout.write(f'  📍 Nombre total de lignes (jobs × emplacements): {total_rows}')
        
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jobs et Emplacements"
        
        # En-têtes
        headers = [
            'job_id',
            'job_reference',
            'job_status',
            'inventory_reference',
            'inventory_label',
            'warehouse_name',
            'location_id',
            'location_reference',
            'location_sous_zone',
            'location_zone',
            'job_detail_reference',
            'job_detail_status',
            'counting_order',
            'job_detail_created_at',
            'job_detail_termine_date'
        ]
        ws.append(headers)
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Parcourir les jobs et leurs emplacements
        jobs_exported = 0
        locations_exported = 0
        
        for job in jobs:
            job_details = job.jobdetail_set.all()
            
            # Si le job n'a pas d'emplacements, exporter quand même le job avec des valeurs vides
            if not job_details.exists():
                row_data = [
                    job.id,
                    job.reference,
                    job.status,
                    job.inventory.reference if job.inventory else '',
                    job.inventory.label if job.inventory else '',
                    job.warehouse.warehouse_name if job.warehouse else '',
                    '',  # location_id
                    '',  # location_reference
                    '',  # location_sous_zone
                    '',  # location_zone
                    '',  # job_detail_reference
                    '',  # job_detail_status
                    '',  # counting_order
                    '',  # job_detail_created_at
                    ''   # job_detail_termine_date
                ]
                ws.append(row_data)
                jobs_exported += 1
            else:
                # Exporter chaque emplacement du job
                for job_detail in job_details:
                    location = job_detail.location
                    
                    # Récupérer les informations de la sous-zone et zone
                    sous_zone_name = ''
                    zone_name = ''
                    if location and location.sous_zone:
                        sous_zone_name = location.sous_zone.sous_zone_name or ''
                        if location.sous_zone.zone:
                            zone_name = location.sous_zone.zone.zone_name or ''
                    
                    # Récupérer l'ordre du comptage si disponible
                    counting_order = ''
                    if job_detail.counting:
                        counting_order = job_detail.counting.order or ''
                    
                    row_data = [
                        job.id,
                        job.reference,
                        job.status,
                        job.inventory.reference if job.inventory else '',
                        job.inventory.label if job.inventory else '',
                        job.warehouse.warehouse_name if job.warehouse else '',
                        location.id if location else '',
                        location.location_reference if location else '',
                        sous_zone_name,
                        zone_name,
                        job_detail.reference,
                        job_detail.status,
                        counting_order,
                        job_detail.created_at.strftime('%Y-%m-%d %H:%M:%S') if job_detail.created_at else '',
                        job_detail.termine_date.strftime('%Y-%m-%d %H:%M:%S') if job_detail.termine_date else ''
                    ]
                    ws.append(row_data)
                    locations_exported += 1
                
                jobs_exported += 1
        
        # Ajuster la largeur des colonnes
        column_widths = {
            'A': 10,  # job_id
            'B': 20,  # job_reference
            'C': 15,  # job_status
            'D': 20,  # inventory_reference
            'E': 30,  # inventory_label
            'F': 25,  # warehouse_name
            'G': 10,  # location_id
            'H': 25,  # location_reference
            'I': 20,  # location_sous_zone
            'J': 20,  # location_zone
            'K': 20,  # job_detail_reference
            'L': 15,  # job_detail_status
            'M': 12,  # counting_order
            'N': 20,  # job_detail_created_at
            'O': 20   # job_detail_termine_date
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Sauvegarder le fichier
        try:
            wb.save(file_path)
            self.stdout.write(self.style.SUCCESS(f'\n✅ Export terminé avec succès!'))
            self.stdout.write(f'  📄 Fichier: {file_path}')
            self.stdout.write(f'  📊 Jobs exportés: {jobs_exported}')
            self.stdout.write(f'  📍 Emplacements exportés: {locations_exported}')
        except Exception as e:
            raise CommandError(f"Erreur lors de la sauvegarde du fichier Excel: {str(e)}")

