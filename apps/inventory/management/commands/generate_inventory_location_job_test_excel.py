"""
Commande Django pour générer un fichier Excel de test pour l'API d'import InventoryLocationJob.

Ce fichier contient tous les cas de test possibles :
- Cas valides (active=true avec job, session_1, session_2)
- Cas valides (active=false sans job/sessions)
- Cas d'erreurs (warehouse inexistant, emplacement inexistant, format invalide, etc.)
- Validation d'incrémentation des jobs
- Emplacements non liés au warehouse
- Sessions hors plage

Exemple d'utilisation:
    python manage.py generate_inventory_location_job_test_excel --inventory-id 1 --file test_import.xlsx
"""
from django.core.management.base import BaseCommand, CommandError
from django.db.models import Q
import logging

from apps.inventory.models import Inventory, Setting
from apps.masterdata.models import Warehouse, Location

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Génère un fichier Excel de test pour l\'API d\'import InventoryLocationJob avec tous les cas de test'

    def add_arguments(self, parser):
        parser.add_argument(
            '--inventory-id',
            type=int,
            required=True,
            help='ID de l\'inventaire à utiliser pour les tests',
        )
        parser.add_argument(
            '--file',
            type=str,
            default='test_inventory_location_job_import.xlsx',
            help='Chemin vers le fichier Excel de sortie (défaut: test_inventory_location_job_import.xlsx)',
        )

    def handle(self, *args, **options):
        inventory_id = options['inventory_id']
        file_path = options['file']
        
        self.stdout.write(self.style.SUCCESS(f'Generation du fichier Excel de test: {file_path}'))
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise CommandError(
                "openpyxl n'est pas installe. Installez-le avec: pip install openpyxl"
            )
        
        # Récupérer l'inventaire
        try:
            inventory = Inventory.objects.get(id=inventory_id, is_deleted=False)
            self.stdout.write(f'  Inventaire: {inventory.label} (ID: {inventory.id})')
        except Inventory.DoesNotExist:
            raise CommandError(f"L'inventaire avec l'ID {inventory_id} n'existe pas.")
        
        # Récupérer les warehouses associés à cet inventaire
        settings = Setting.objects.filter(inventory=inventory).select_related('warehouse')
        warehouses = [setting.warehouse for setting in settings]
        
        if not warehouses:
            raise CommandError(f"Aucun warehouse associe a l'inventaire {inventory.label}")
        
        self.stdout.write(f'  Warehouses associes: {len(warehouses)}')
        for wh in warehouses:
            self.stdout.write(f'     - {wh.warehouse_name} ({wh.reference})')
        
        # Récupérer les emplacements pour chaque warehouse
        all_locations = []
        for warehouse in warehouses:
            locations = Location.objects.filter(
                sous_zone__zone__warehouse=warehouse,
                is_deleted=False
            ).select_related('sous_zone__zone__warehouse')[:20]  # Limiter à 20 par warehouse pour le test
            all_locations.extend(locations)
        
        if not all_locations:
            raise CommandError(f"Aucun emplacement trouve pour les warehouses de l'inventaire")
        
        self.stdout.write(f'  Emplacements trouves: {len(all_locations)}')
        
        # Récupérer un warehouse qui n'appartient pas à l'inventaire (pour les tests d'erreur)
        other_warehouses = Warehouse.objects.exclude(
            id__in=[wh.id for wh in warehouses]
        ).filter(is_deleted=False)[:2]
        
        # Récupérer des emplacements d'autres warehouses (pour les tests d'erreur)
        other_locations = []
        if other_warehouses.exists():
            for wh in other_warehouses:
                locs = Location.objects.filter(
                    sous_zone__zone__warehouse=wh,
                    is_deleted=False
                )[:3]
                other_locations.extend(locs)
        
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Import"
        
        # En-têtes
        headers = ['warehouse', 'emplacement', 'active', 'job', 'session_1', 'session_2', 'commentaire']
        ws.append(headers)
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        row_num = 2
        
        # ============================================
        # SECTION 1: CAS VALIDES - Active = True
        # ============================================
        ws.append(['', '', '', '', '', '', '=== CAS VALIDES - Active = True ==='])
        row_num += 1
        
        # Cas 1-5: Lignes valides avec active=true, jobs incrémentés correctement
        valid_warehouse = warehouses[0]
        valid_locations = all_locations[:5] if len(all_locations) >= 5 else all_locations
        
        for idx, location in enumerate(valid_locations, 1):
            job_num = f"{idx:02d}"
            ws.append([
                valid_warehouse.reference,  # warehouse
                location.location_reference,  # emplacement
                True,  # active
                f'job-{job_num}',  # job
                f'equipe-{1000 + idx}',  # session_1 (1000-1999)
                f'equipe-{2000 + idx}',  # session_2 (2000-2999)
                f'Cas valide {idx}: active=true avec job-{job_num}'
            ])
            row_num += 1
        
        # ============================================
        # SECTION 2: CAS VALIDES - Active = False
        # ============================================
        ws.append(['', '', '', '', '', '', '=== CAS VALIDES - Active = False ==='])
        row_num += 1
        
        # Cas 6-7: Lignes valides avec active=false (job et sessions peuvent être vides)
        if len(all_locations) >= 7:
            for location in all_locations[5:7]:
                ws.append([
                    valid_warehouse.reference,  # warehouse
                    location.location_reference,  # emplacement
                    False,  # active
                    '',  # job (vide car active=false)
                    '',  # session_1 (vide car active=false)
                    '',  # session_2 (vide car active=false)
                    'Cas valide: active=false, job et sessions vides'
                ])
                row_num += 1
        
        # ============================================
        # SECTION 3: ERREURS - Warehouse
        # ============================================
        ws.append(['', '', '', '', '', '', '=== ERREURS - Warehouse ==='])
        row_num += 1
        
        # Cas 8: Warehouse inexistant
        ws.append([
            'WAREHOUSE_INEXISTANT',  # warehouse
            valid_locations[0].location_reference if valid_locations else 'LOC001',  # emplacement
            True,
            'job-01',
            'equipe-1000',
            'equipe-2000',
            'ERREUR: Warehouse inexistant'
        ])
        row_num += 1
        
        # Cas 9: Warehouse n'appartient pas à l'inventaire
        if other_warehouses.exists():
            other_wh = other_warehouses[0]
            ws.append([
                other_wh.reference,  # warehouse
                valid_locations[0].location_reference if valid_locations else 'LOC001',  # emplacement
                True,
                'job-01',
                'equipe-1000',
                'equipe-2000',
                f'ERREUR: Warehouse {other_wh.reference} n\'appartient pas à l\'inventaire'
            ])
            row_num += 1
        
        # Cas 10: Warehouse vide
        ws.append([
            '',  # warehouse vide
            valid_locations[0].location_reference if valid_locations else 'LOC001',  # emplacement
            True,
            'job-01',
            'equipe-1000',
            'equipe-2000',
            'ERREUR: Warehouse vide (obligatoire)'
        ])
        row_num += 1
        
        # ============================================
        # SECTION 4: ERREURS - Emplacement
        # ============================================
        ws.append(['', '', '', '', '', '', '=== ERREURS - Emplacement ==='])
        row_num += 1
        
        # Cas 11: Emplacement inexistant
        ws.append([
            valid_warehouse.reference,  # warehouse
            'EMPLACEMENT_INEXISTANT',  # emplacement
            True,
            'job-01',
            'equipe-1000',
            'equipe-2000',
            'ERREUR: Emplacement inexistant'
        ])
        row_num += 1
        
        # Cas 12: Emplacement vide
        ws.append([
            valid_warehouse.reference,  # warehouse
            '',  # emplacement vide
            True,
            'job-01',
            'equipe-1000',
            'equipe-2000',
            'ERREUR: Emplacement vide (obligatoire)'
        ])
        row_num += 1
        
        # Cas 13: Emplacement n'appartient pas au warehouse
        if other_locations:
            other_loc = other_locations[0]
            ws.append([
                valid_warehouse.reference,  # warehouse
                other_loc.location_reference,  # emplacement d'un autre warehouse
                True,
                'job-01',
                'equipe-1000',
                'equipe-2000',
                f'ERREUR: Emplacement {other_loc.location_reference} n\'appartient pas au warehouse {valid_warehouse.reference}'
            ])
            row_num += 1
        
        # ============================================
        # SECTION 5: ERREURS - Format Job
        # ============================================
        ws.append(['', '', '', '', '', '', '=== ERREURS - Format Job ==='])
        row_num += 1
        
        # Cas 14: Format job invalide (sans préfixe)
        if valid_locations:
            ws.append([
                valid_warehouse.reference,
                valid_locations[0].location_reference,
                True,
                '01',  # Format invalide
                'equipe-1000',
                'equipe-2000',
                'ERREUR: Format job invalide (attendu: job-XX)'
            ])
            row_num += 1
        
        # Cas 15: Format job invalide (mauvais préfixe)
        if valid_locations:
            ws.append([
                valid_warehouse.reference,
                valid_locations[0].location_reference,
                True,
                'travail-01',  # Format invalide
                'equipe-1000',
                'equipe-2000',
                'ERREUR: Format job invalide (attendu: job-XX)'
            ])
            row_num += 1
        
        # Cas 16: Job vide alors que active=true
        if valid_locations:
            ws.append([
                valid_warehouse.reference,
                valid_locations[0].location_reference,
                True,
                '',  # Job vide alors que active=true
                'equipe-1000',
                'equipe-2000',
                'ERREUR: Job obligatoire lorsque active=true'
            ])
            row_num += 1
        
        # ============================================
        # SECTION 6: ERREURS - Format Session_1
        # ============================================
        ws.append(['', '', '', '', '', '', '=== ERREURS - Format Session_1 ==='])
        row_num += 1
        
        # Cas 17: Format session_1 invalide
        if valid_locations:
            ws.append([
                valid_warehouse.reference,
                valid_locations[0].location_reference,
                True,
                'job-01',
                'team-1000',  # Format invalide
                'equipe-2000',
                'ERREUR: Format session_1 invalide (attendu: equipe-XXXX)'
            ])
            row_num += 1
        
        # Cas 18: Session_1 hors plage (< 1000)
        if valid_locations:
            ws.append([
                valid_warehouse.reference,
                valid_locations[0].location_reference,
                True,
                'job-01',
                'equipe-999',  # Hors plage
                'equipe-2000',
                'ERREUR: Session_1 hors plage (attendu: 1000-1999)'
            ])
            row_num += 1
        
        # Cas 19: Session_1 hors plage (> 1999)
        if valid_locations:
            ws.append([
                valid_warehouse.reference,
                valid_locations[0].location_reference,
                True,
                'job-01',
                'equipe-2000',  # Hors plage (doit être dans 1000-1999)
                'equipe-2000',
                'ERREUR: Session_1 hors plage (attendu: 1000-1999)'
            ])
            row_num += 1
        
        # Cas 20: Session_1 vide alors que active=true
        if valid_locations:
            ws.append([
                valid_warehouse.reference,
                valid_locations[0].location_reference,
                True,
                'job-01',
                '',  # Session_1 vide alors que active=true
                'equipe-2000',
                'ERREUR: Session_1 obligatoire lorsque active=true'
            ])
            row_num += 1
        
        # ============================================
        # SECTION 7: ERREURS - Format Session_2
        # ============================================
        ws.append(['', '', '', '', '', '', '=== ERREURS - Format Session_2 ==='])
        row_num += 1
        
        # Cas 21: Format session_2 invalide
        if valid_locations:
            ws.append([
                valid_warehouse.reference,
                valid_locations[0].location_reference,
                True,
                'job-01',
                'equipe-1000',
                'team-2000',  # Format invalide
                'ERREUR: Format session_2 invalide (attendu: equipe-XXXX)'
            ])
            row_num += 1
        
        # Cas 22: Session_2 hors plage (< 2000)
        if valid_locations:
            ws.append([
                valid_warehouse.reference,
                valid_locations[0].location_reference,
                True,
                'job-01',
                'equipe-1000',
                'equipe-1999',  # Hors plage
                'ERREUR: Session_2 hors plage (attendu: 2000-2999)'
            ])
            row_num += 1
        
        # Cas 23: Session_2 hors plage (> 2999)
        if valid_locations:
            ws.append([
                valid_warehouse.reference,
                valid_locations[0].location_reference,
                True,
                'job-01',
                'equipe-1000',
                'equipe-3000',  # Hors plage
                'ERREUR: Session_2 hors plage (attendu: 2000-2999)'
            ])
            row_num += 1
        
        # Cas 24: Session_2 vide alors que active=true
        if valid_locations:
            ws.append([
                valid_warehouse.reference,
                valid_locations[0].location_reference,
                True,
                'job-01',
                'equipe-1000',
                '',  # Session_2 vide alors que active=true
                'ERREUR: Session_2 obligatoire lorsque active=true'
            ])
            row_num += 1
        
        # ============================================
        # SECTION 8: ERREURS - Incrémentation Jobs
        # ============================================
        ws.append(['', '', '', '', '', '', '=== ERREURS - Incrémentation Jobs ==='])
        row_num += 1
        
        # Cas 25: Jobs ne commencent pas à job-01
        if len(valid_locations) >= 2:
            ws.append([
                valid_warehouse.reference,
                valid_locations[0].location_reference,
                True,
                'job-02',  # Commence à job-02 au lieu de job-01
                'equipe-1000',
                'equipe-2000',
                'ERREUR: Jobs doivent commencer à job-01'
            ])
            row_num += 1
        
        # Cas 26: Rupture dans l\'incrémentation des jobs
        if len(valid_locations) >= 3:
            # Ajouter job-01
            ws.append([
                valid_warehouse.reference,
                valid_locations[0].location_reference,
                True,
                'job-01',
                'equipe-1000',
                'equipe-2000',
                'Cas valide pour test rupture'
            ])
            row_num += 1
            
            # Ajouter job-03 (saut de job-02)
            ws.append([
                valid_warehouse.reference,
                valid_locations[1].location_reference,
                True,
                'job-03',  # Rupture: job-02 manquant
                'equipe-1001',
                'equipe-2001',
                'ERREUR: Rupture d\'incrémentation (job-02 manquant)'
            ])
            row_num += 1
        
        # ============================================
        # SECTION 9: CAS VALIDES - Variantes Active
        # ============================================
        ws.append(['', '', '', '', '', '', '=== CAS VALIDES - Variantes Active ==='])
        row_num += 1
        
        # Cas 27-28: Différentes valeurs pour active (normalisation booléenne)
        if len(valid_locations) >= 2:
            # Cas 27: active = "true" (string)
            ws.append([
                valid_warehouse.reference,
                valid_locations[0].location_reference if len(valid_locations) > 0 else 'LOC001',
                'true',  # String
                'job-01',
                'equipe-1000',
                'equipe-2000',
                'Cas valide: active="true" (string)'
            ])
            row_num += 1
            
            # Cas 28: active = "false" (string)
            ws.append([
                valid_warehouse.reference,
                valid_locations[1].location_reference if len(valid_locations) > 1 else 'LOC002',
                'false',  # String
                '',  # Vide car active=false
                '',  # Vide car active=false
                '',  # Vide car active=false
                'Cas valide: active="false" (string)'
            ])
            row_num += 1
        
        # ============================================
        # SECTION 10: CAS VALIDES - Multi-warehouse
        # ============================================
        if len(warehouses) >= 2:
            ws.append(['', '', '', '', '', '', '=== CAS VALIDES - Multi-warehouse ==='])
            row_num += 1
            
            # Cas 29-30: Lignes avec différents warehouses
            warehouse_2 = warehouses[1]
            locations_wh2 = Location.objects.filter(
                sous_zone__zone__warehouse=warehouse_2,
                is_deleted=False
            )[:2]
            
            for idx, location in enumerate(locations_wh2, 1):
                ws.append([
                    warehouse_2.reference,  # Autre warehouse
                    location.location_reference,
                    True,
                    f'job-{idx:02d}',  # Jobs recommencent à 01 pour ce warehouse
                    f'equipe-{1000 + idx}',
                    f'equipe-{2000 + idx}',
                    f'Cas valide: Warehouse {warehouse_2.reference}'
                ])
                row_num += 1
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 25  # warehouse
        ws.column_dimensions['B'].width = 25  # emplacement
        ws.column_dimensions['C'].width = 12  # active
        ws.column_dimensions['D'].width = 15  # job
        ws.column_dimensions['E'].width = 18  # session_1
        ws.column_dimensions['F'].width = 18  # session_2
        ws.column_dimensions['G'].width = 60  # commentaire
        
        # Ajouter des couleurs pour les sections
        for row in ws.iter_rows(min_row=2, max_row=row_num):
            comment_cell = row[6]  # Colonne commentaire
            if comment_cell.value and 'ERREUR' in str(comment_cell.value):
                # Colorier en rouge pour les erreurs
                for cell in row:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            elif comment_cell.value and '=== CAS VALIDES' in str(comment_cell.value):
                # Colorier en vert pour les sections valides
                for cell in row:
                    cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
            elif comment_cell.value and '=== ERREURS' in str(comment_cell.value):
                # Colorier en orange pour les sections d'erreurs
                for cell in row:
                    cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        
        # Sauvegarder le fichier
        try:
            wb.save(file_path)
            self.stdout.write(self.style.SUCCESS(f'Fichier Excel genere avec succes!'))
            self.stdout.write(f'  Fichier: {file_path}')
            self.stdout.write(f'  Nombre de lignes: {row_num - 1}')
            self.stdout.write(f'  Inventaire utilise: {inventory.label} (ID: {inventory.id})')
            self.stdout.write(f'  Warehouses utilises: {", ".join([wh.reference for wh in warehouses])}')
            self.stdout.write('')
            self.stdout.write(self.style.WARNING('IMPORTANT: Ce fichier contient des cas valides ET des cas d\'erreur.'))
            self.stdout.write('   Pour tester uniquement les cas valides, supprimez les lignes avec "ERREUR" dans la colonne commentaire.')
        except Exception as e:
            raise CommandError(f"Erreur lors de la sauvegarde du fichier Excel: {str(e)}")
