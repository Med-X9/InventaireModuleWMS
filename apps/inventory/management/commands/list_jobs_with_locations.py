"""
Commande Django pour lister les jobs avec leurs emplacements, sous-zones et zones.

Cette commande :
- Récupère tous les jobs pour un inventory-id et warehouse-id donnés
- Pour chaque job, récupère les emplacements (via JobDetail)
- Affiche les informations : Job, Emplacement, Sous-zone, Zone
- Peut exporter les résultats en Excel ou afficher dans la console

Exemple d'utilisation:
    python manage.py list_jobs_with_locations --inventory-id 2 --warehouse-id 1
    python manage.py list_jobs_with_locations --inventory-id 2 --warehouse-id 1 --output-file jobs_locations.xlsx
    python manage.py list_jobs_with_locations --inventory-id 2 --warehouse-id 1 --format console
"""
from django.core.management.base import BaseCommand, CommandError
from django.db.models import Q
import logging
from typing import Optional, List, Dict
from collections import defaultdict

from apps.inventory.models import Job, JobDetail
from apps.masterdata.models import Location, SousZone, Zone

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Liste les jobs avec leurs emplacements, sous-zones et zones'

    def add_arguments(self, parser):
        parser.add_argument(
            '--inventory-id',
            type=int,
            required=True,
            help='ID de l\'inventaire',
        )
        parser.add_argument(
            '--warehouse-id',
            type=int,
            required=True,
            help='ID de l\'entrepôt',
        )
        parser.add_argument(
            '--output-file',
            type=str,
            help='Chemin vers le fichier Excel de sortie (optionnel)',
        )
        parser.add_argument(
            '--format',
            type=str,
            choices=['console', 'excel'],
            default='console',
            help='Format de sortie : console ou excel (défaut: console)',
        )

    def handle(self, *args, **options):
        inventory_id = options['inventory_id']
        warehouse_id = options['warehouse_id']
        output_file = options.get('output_file')
        output_format = options.get('format', 'console')
        
        # Si un fichier de sortie est fourni, forcer le format Excel
        if output_file:
            output_format = 'excel'
        
        self.stdout.write(self.style.SUCCESS('📋 Liste des jobs avec leurs emplacements, sous-zones et zones'))
        self.stdout.write(f'  🔍 Inventaire ID: {inventory_id}')
        self.stdout.write(f'  🔍 Entrepôt ID: {warehouse_id}')
        
        # Récupérer les jobs
        jobs = Job.objects.filter(
            inventory_id=inventory_id,
            warehouse_id=warehouse_id
        ).select_related('inventory', 'warehouse').order_by('reference')
        
        if not jobs.exists():
            self.stdout.write(self.style.WARNING('⚠️  Aucun job trouvé'))
            return
        
        self.stdout.write(f'  ✅ {jobs.count()} job(s) trouvé(s)')
        
        # Récupérer les données pour chaque job
        jobs_data = []
        
        for job in jobs:
            # Récupérer les JobDetail avec leurs locations, sous-zones et zones
            job_details = JobDetail.objects.filter(
                job=job
            ).select_related(
                'location',
                'location__sous_zone',
                'location__sous_zone__zone'
            ).order_by('location__location_reference', '-id')
            
            # Grouper par location pour éviter les doublons
            locations_data = []
            seen_locations = set()
            
            for job_detail in job_details:
                location = job_detail.location
                if location.id in seen_locations:
                    continue
                seen_locations.add(location.id)
                
                sous_zone = location.sous_zone if location.sous_zone else None
                zone = sous_zone.zone if sous_zone else None
                
                locations_data.append({
                    'location_id': location.id,
                    'location_reference': location.location_reference,
                    'location_code': location.reference,
                    'sous_zone_id': sous_zone.id if sous_zone else None,
                    'sous_zone_name': sous_zone.sous_zone_name if sous_zone else 'N/A',
                    'sous_zone_reference': sous_zone.reference if sous_zone else 'N/A',
                    'zone_id': zone.id if zone else None,
                    'zone_name': zone.zone_name if zone else 'N/A',
                    'zone_reference': zone.reference if zone else 'N/A',
                })
            
            jobs_data.append({
                'job_id': job.id,
                'job_reference': job.reference,
                'job_status': job.status,
                'locations': locations_data
            })
        
        # Afficher ou exporter selon le format
        if output_format == 'excel':
            self._export_to_excel(jobs_data, output_file or 'jobs_locations.xlsx')
        else:
            self._display_console(jobs_data)
    
    def _display_console(self, jobs_data: List[Dict]):
        """
        Affiche les données dans la console.
        
        Args:
            jobs_data: Liste des données des jobs avec leurs emplacements
        """
        self.stdout.write('\n' + '='*80)
        self.stdout.write(self.style.SUCCESS('📊 RÉSULTATS'))
        self.stdout.write('='*80)
        
        total_locations = 0
        
        for job_data in jobs_data:
            job_ref = job_data['job_reference']
            job_status = job_data['job_status']
            locations = job_data['locations']
            total_locations += len(locations)
            
            self.stdout.write(f'\n📦 Job: {job_ref} (Statut: {job_status})')
            self.stdout.write(f'  📍 {len(locations)} emplacement(s)')
            
            if locations:
                # Afficher un tableau
                self.stdout.write('  ' + '-'*76)
                self.stdout.write(f'  {"Emplacement":<20} {"Sous-zone":<20} {"Zone":<20}')
                self.stdout.write('  ' + '-'*76)
                
                for loc in locations:
                    location_ref = loc['location_reference'][:18] if len(loc['location_reference']) > 18 else loc['location_reference']
                    sous_zone_name = loc['sous_zone_name'][:18] if len(loc['sous_zone_name']) > 18 else loc['sous_zone_name']
                    zone_name = loc['zone_name'][:18] if len(loc['zone_name']) > 18 else loc['zone_name']
                    
                    self.stdout.write(
                        f'  {location_ref:<20} {sous_zone_name:<20} {zone_name:<20}'
                    )
        
        self.stdout.write('\n' + '='*80)
        self.stdout.write(f'  Total: {len(jobs_data)} job(s), {total_locations} emplacement(s)')
        self.stdout.write('='*80)
    
    def _export_to_excel(self, jobs_data: List[Dict], output_file: str):
        """
        Exporte les données vers un fichier Excel.
        
        Args:
            jobs_data: Liste des données des jobs avec leurs emplacements
            output_file: Chemin vers le fichier Excel de sortie
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise CommandError(
                "openpyxl n'est pas installé. Installez-le avec: pip install openpyxl"
            )
        
        try:
            # Créer un nouveau workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Jobs et Emplacements"
            
            # En-têtes
            headers = [
                'Job ID',
                'Référence Job',
                'Statut Job',
                'Emplacement ID',
                'Référence Emplacement',
                'Code Emplacement',
                'Sous-zone ID',
                'Nom Sous-zone',
                'Référence Sous-zone',
                'Zone ID',
                'Nom Zone',
                'Référence Zone'
            ]
            
            # Style pour les en-têtes
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Écrire les en-têtes
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Écrire les données
            row_num = 2
            for job_data in jobs_data:
                job_id = job_data['job_id']
                job_ref = job_data['job_reference']
                job_status = job_data['job_status']
                locations = job_data['locations']
                
                if not locations:
                    # Si pas d'emplacement, écrire quand même le job
                    ws.cell(row=row_num, column=1, value=job_id)
                    ws.cell(row=row_num, column=2, value=job_ref)
                    ws.cell(row=row_num, column=3, value=job_status)
                    row_num += 1
                else:
                    # Écrire chaque emplacement
                    for loc in locations:
                        ws.cell(row=row_num, column=1, value=job_id)
                        ws.cell(row=row_num, column=2, value=job_ref)
                        ws.cell(row=row_num, column=3, value=job_status)
                        ws.cell(row=row_num, column=4, value=loc['location_id'])
                        ws.cell(row=row_num, column=5, value=loc['location_reference'])
                        ws.cell(row=row_num, column=6, value=loc['location_code'])
                        ws.cell(row=row_num, column=7, value=loc['sous_zone_id'])
                        ws.cell(row=row_num, column=8, value=loc['sous_zone_name'])
                        ws.cell(row=row_num, column=9, value=loc['sous_zone_reference'])
                        ws.cell(row=row_num, column=10, value=loc['zone_id'])
                        ws.cell(row=row_num, column=11, value=loc['zone_name'])
                        ws.cell(row=row_num, column=12, value=loc['zone_reference'])
                        row_num += 1
            
            # Ajuster la largeur des colonnes
            column_widths = {
                'A': 10,  # Job ID
                'B': 15,  # Référence Job
                'C': 12,  # Statut Job
                'D': 12,  # Emplacement ID
                'E': 25,  # Référence Emplacement
                'F': 20,  # Code Emplacement
                'G': 12,  # Sous-zone ID
                'H': 25,  # Nom Sous-zone
                'I': 20,  # Référence Sous-zone
                'J': 10,  # Zone ID
                'K': 25,  # Nom Zone
                'L': 20,  # Référence Zone
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Sauvegarder le fichier
            wb.save(output_file)
            
            self.stdout.write(
                self.style.SUCCESS(
                    f'\n✅ Fichier Excel créé avec succès: {output_file}'
                )
            )
            self.stdout.write(f'  📊 {len(jobs_data)} job(s) exporté(s)')
            
            total_locations = sum(len(job['locations']) for job in jobs_data)
            self.stdout.write(f'  📍 {total_locations} emplacement(s) exporté(s)')
            
        except Exception as e:
            import traceback
            logger.error(f"Erreur lors de l'export Excel: {str(e)}\n{traceback.format_exc()}")
            raise CommandError(f"Erreur lors de l'export Excel: {str(e)}")

