"""
Commande Django pour lancer un nouveau comptage (3e ou ultérieur) pour une liste de jobs.

Cette commande :
- Récupère les jobs par leurs références
- Trouve automatiquement les emplacements avec écart pour chaque job
- Lance un nouveau comptage pour ces emplacements
- Utilise la même logique que l'API CountingLaunchView

Exemple d'utilisation:
    python manage.py launch_counting_for_jobs --username user123  # Utilise STATIC_JOB_REFERENCES par défaut
    python manage.py launch_counting_for_jobs --username user123 --job-refs JOB-0001 JOB-0002  # Surcharge avec des jobs spécifiques
    python manage.py launch_counting_for_jobs --file jobs_sessions.txt  # Lit format Job + Session (séparé par tab)
    python manage.py launch_counting_for_jobs --file jobs_sessions.xlsx  # Lit format Excel avec colonnes Job + Session
"""
from django.core.management.base import BaseCommand, CommandError
import logging
from typing import List, Dict
from collections import defaultdict

from apps.inventory.models import Job
from apps.inventory.services.counting_launch_service import CountingLaunchService
from apps.inventory.exceptions.counting_exceptions import (
    CountingValidationError,
    CountingNotFoundError,
    CountingCreationError,
)
from apps.users.models import UserApp

logger = logging.getLogger(__name__)

# Liste statique des références de jobs pour lesquels lancer les comptages
STATIC_JOB_REFERENCES = [
    'JOB-0182',
    'JOB-0594',
    'JOB-0595',
    'JOB-0596',
    'JOB-0597',
    'JOB-0600',
]


class Command(BaseCommand):
    help = 'Lance un nouveau comptage pour une liste de jobs (emplacements avec écart)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--username',
            type=str,
            help='Username de la session (équipe mobile) à affecter (obligatoire si --file non fourni)',
        )
        parser.add_argument(
            '--job-refs',
            nargs='+',
            type=str,
            help='Liste des références de jobs (ex: JOB-0001 JOB-0002)',
        )
        parser.add_argument(
            '--file',
            type=str,
            help='Chemin vers un fichier Excel (.xlsx) ou texte (.txt). Format texte: Job + Session (séparé par tab). Format Excel: colonnes Job + Session',
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default=0,
            help='Nom ou index de la feuille Excel à lire (défaut: première feuille)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Mode test : affiche ce qui sera fait sans lancer les comptages',
        )
        parser.add_argument(
            '--debug',
            action='store_true',
            help='Mode debug : affiche des informations détaillées sur les emplacements avec écart',
        )

    def handle(self, *args, **options):
        username = options.get('username')
        job_refs_input = options.get('job_refs', [])
        file_path = options.get('file')
        sheet = options.get('sheet', 0)
        dry_run = options.get('dry_run', False)
        debug = options.get('debug', False)
        
        # Activer le logging debug si demandé
        if debug:
            logging.getLogger('apps.inventory.services.counting_launch_service').setLevel(logging.DEBUG)
            logging.getLogger('apps.inventory.management.commands.launch_counting_for_jobs').setLevel(logging.DEBUG)
        
        # Dictionnaire pour grouper les jobs par session
        # Format: {session_username: [job_ref1, job_ref2, ...]}
        jobs_by_session: Dict[str, List[str]] = defaultdict(list)
        
        # Si un fichier est fourni, lire le format Job + Session
        if file_path:
            try:
                if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                    # Fichier Excel avec format Job + Session
                    jobs_by_session = self.read_excel_file_with_sessions(file_path, sheet)
                else:
                    # Fichier texte avec format Job + Session (séparé par tab)
                    jobs_by_session = self.read_text_file_with_sessions(file_path)
            except FileNotFoundError:
                raise CommandError(f"Fichier non trouvé: {file_path}")
            except CommandError:
                raise
            except Exception as e:
                raise CommandError(f"Erreur lors de la lecture du fichier: {str(e)}")
        
        # Si pas de fichier, utiliser username + job-refs ou STATIC_JOB_REFERENCES
        if not jobs_by_session:
            # Validation : username est obligatoire si pas de fichier
            if not username:
                raise CommandError("Vous devez fournir --username ou --file")
            
            # Initialiser avec STATIC_JOB_REFERENCES par défaut
            job_refs = STATIC_JOB_REFERENCES.copy()
            using_static = True
            
            # Surcharger avec --job-refs si fourni
            if job_refs_input:
                job_refs = job_refs_input.copy()
                using_static = False
            
            # Grouper tous les jobs sous le même username
            jobs_by_session[username] = job_refs
            
            # Afficher la source des références
            if using_static:
                self.stdout.write(self.style.SUCCESS(f'📋 Utilisation de STATIC_JOB_REFERENCES: {len(job_refs)} job(s)'))
            else:
                self.stdout.write(self.style.SUCCESS(f'📋 Utilisation de références personnalisées: {len(job_refs)} job(s)'))
        
        # Afficher le résumé
        total_jobs = sum(len(jobs) for jobs in jobs_by_session.values())
        self.stdout.write(self.style.SUCCESS(f'📋 Lancement de comptage pour {total_jobs} job(s) répartis sur {len(jobs_by_session)} session(s)'))
        
        if dry_run:
            self.stdout.write(self.style.WARNING('⚠️  MODE DRY-RUN : Aucun comptage ne sera lancé'))
        
        # Traiter chaque session
        all_results = []
        for session_username, job_refs in jobs_by_session.items():
            # Supprimer les doublons pour cette session
            job_refs = list(set(job_refs))
            
            if not job_refs:
                continue
            
            self.stdout.write(f'\n{"="*60}')
            self.stdout.write(f'📦 SESSION: {session_username} ({len(job_refs)} job(s))')
            self.stdout.write(f'{"="*60}')
            
            # Convertir username en session_id
            try:
                session_id = self._get_session_id_by_username(session_username)
                self.stdout.write(f'  👤 Username: {session_username} → Session ID: {session_id}')
            except CommandError as e:
                self.stdout.write(self.style.ERROR(f'  ❌ {str(e)}'))
                continue
            
            # Récupérer les jobs par leurs références
            jobs = Job.objects.filter(reference__in=job_refs).select_related('inventory', 'warehouse')
            
            found_refs = set(jobs.values_list('reference', flat=True))
            missing_refs = set(job_refs) - found_refs
            
            if missing_refs:
                self.stdout.write(self.style.ERROR(
                    f'  ❌ Jobs non trouvés: {", ".join(sorted(missing_refs))}'
                ))
            
            if not jobs.exists():
                self.stdout.write(self.style.ERROR('  ❌ Aucun job trouvé pour cette session'))
                continue
            
            self.stdout.write(f'  ✅ Jobs trouvés: {jobs.count()}')
            
            # Vérifier les emplacements avec écart pour chaque job AVANT de lancer
            service = CountingLaunchService()
            jobs_with_locations = {}
            total_locations_with_ecart = 0
            
            self.stdout.write('\n  🔍 Vérification des emplacements avec écart:')
            for job in jobs:
                try:
                    location_ids = service.get_locations_with_discrepancy_for_job(job.id)
                    jobs_with_locations[job.id] = {
                        'job': job,
                        'location_ids': location_ids
                    }
                    total_locations_with_ecart += len(location_ids)
                    
                    if len(location_ids) > 0:
                        self.stdout.write(
                            self.style.SUCCESS(
                                f'    ✅ {job.reference}: {len(location_ids)} emplacement(s) avec écart'
                            )
                        )
                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(
                            f'    ❌ {job.reference}: Erreur lors de la vérification - {str(e)}'
                        )
                    )
                    logger.error(f"Erreur lors de la vérification pour job {job.reference}: {str(e)}", exc_info=True)
                    jobs_with_locations[job.id] = {
                        'job': job,
                        'location_ids': [],
                        'error': str(e)
                    }
            
            # Si aucun emplacement avec écart n'a été trouvé, passer à la session suivante
            if total_locations_with_ecart == 0:
                self.stdout.write(
                    self.style.WARNING(
                        '  ⚠️  Aucun emplacement avec écart trouvé pour les jobs de cette session.'
                    )
                )
                continue
            
            self.stdout.write(
                self.style.SUCCESS(
                    f'  📊 Total: {total_locations_with_ecart} emplacement(s) avec écart à traiter'
                )
            )
            
            # Récupérer les IDs des jobs qui ont des emplacements avec écart
            job_ids = [job_id for job_id, data in jobs_with_locations.items() if len(data['location_ids']) > 0]
            
            if dry_run:
                self.stdout.write(f'  📋 Jobs à traiter: {len(job_ids)}')
                self.stdout.write(f'  📍 Emplacements avec écart à traiter: {total_locations_with_ecart}')
                continue
            
            # Lancer les comptages pour cette session
            try:
                result = service.launch_counting_for_jobs(job_ids, session_id)
                result['session_username'] = session_username
                result['session_id'] = session_id
                all_results.append(result)
                
                # Afficher les résultats pour cette session
                self.stdout.write(f'\n  📊 RÉSULTATS pour {session_username}:')
                self.stdout.write(f"    Emplacements avec écart trouvés: {result['total_locations_found']}")
                self.stdout.write(f"    Emplacements traités avec succès: {result['total_locations_processed']}")
                self.stdout.write(f"    Emplacements en échec: {result['total_locations_failed']}")
                
            except Exception as e:
                self.stdout.write(
                    self.style.ERROR(
                        f'  ❌ Erreur lors du lancement pour {session_username}: {str(e)}'
                    )
                )
                logger.error(f"Erreur lors du lancement pour session {session_username}: {str(e)}", exc_info=True)
        
        # Afficher le résumé global
        if not dry_run and all_results:
            self.stdout.write('\n' + '='*60)
            self.stdout.write(self.style.SUCCESS('📊 RÉSUMÉ GLOBAL'))
            self.stdout.write('='*60)
            total_found = sum(r['total_locations_found'] for r in all_results)
            total_processed = sum(r['total_locations_processed'] for r in all_results)
            total_failed = sum(r['total_locations_failed'] for r in all_results)
            
            self.stdout.write(f"  Sessions traitées: {len(all_results)}")
            self.stdout.write(f"  Emplacements avec écart trouvés: {total_found}")
            self.stdout.write(f"  Emplacements traités avec succès: {total_processed}")
            self.stdout.write(f"  Emplacements en échec: {total_failed}")
            
            if total_processed == total_found:
                self.stdout.write(
                    self.style.SUCCESS(
                        '\n✅ Tous les emplacements ont été traités avec succès!'
                    )
                )
            elif total_failed > 0:
                self.stdout.write(
                    self.style.WARNING(
                        f'\n⚠️  {total_failed} emplacement(s) sur {total_found} ont échoué.'
                    )
                )
    
    def _get_session_id_by_username(self, username: str) -> int:
        """
        Récupère l'ID de la session (UserApp de type Mobile) par username.
        Convertit le username en lowercase avant la recherche dans la BDD.
        
        Args:
            username: Username de la session
            
        Returns:
            ID de la session
            
        Raises:
            CommandError: Si la session n'est pas trouvée
        """
        if not username:
            raise CommandError("Le username ne peut pas être vide")
        
        # Convertir en lowercase avant la recherche
        username_lower = username.strip().lower()
        
        try:
            session = UserApp.objects.get(
                username__iexact=username_lower,
                type='Mobile',
                is_active=True
            )
            return session.id
        except UserApp.DoesNotExist:
            raise CommandError(
                f"Session avec le username '{username_lower}' non trouvée (type Mobile, actif)"
            )
        except UserApp.MultipleObjectsReturned:
            # Si plusieurs sessions trouvées, prendre la première
            session = UserApp.objects.filter(
                username__iexact=username_lower,
                type='Mobile',
                is_active=True
            ).first()
            self.stdout.write(
                self.style.WARNING(
                    f"  ⚠ Plusieurs sessions trouvées pour '{username_lower}', utilisation de la première: {session.username} (ID: {session.id})"
                )
            )
            return session.id
    
    def read_text_file_with_sessions(self, file_path: str) -> Dict[str, List[str]]:
        """
        Lit un fichier texte et extrait les références de jobs avec leurs sessions.
        Format attendu: Job\tSession (séparé par tabulation)
        
        Args:
            file_path: Chemin vers le fichier texte
            
        Returns:
            Dictionnaire {session_username: [job_ref1, job_ref2, ...]}
        """
        jobs_by_session = defaultdict(list)
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                
                # Ignorer la première ligne si c'est un en-tête
                start_line = 0
                if lines and ('Référence' in lines[0] or 'Job' in lines[0] or 'SESSION' in lines[0]):
                    start_line = 1
                
                for line_num, line in enumerate(lines[start_line:], start=start_line + 1):
                    line = line.strip()
                    if not line:
                        continue
                    
                    # Séparer par tabulation ou espace multiple
                    parts = line.split('\t')
                    if len(parts) < 2:
                        # Essayer avec des espaces multiples
                        parts = line.split()
                    
                    if len(parts) >= 2:
                        job_ref = parts[0].strip()
                        session = parts[1].strip().lower()
                        
                        if job_ref and session:
                            jobs_by_session[session].append(job_ref)
                    else:
                        self.stdout.write(
                            self.style.WARNING(
                                f'  ⚠️  Ligne {line_num} ignorée (format invalide): {line}'
                            )
                        )
            
            return dict(jobs_by_session)
            
        except Exception as e:
            import traceback
            logger.error(f"Erreur lors de la lecture du fichier texte {file_path}: {str(e)}\n{traceback.format_exc()}")
            raise CommandError(f"Erreur lors de la lecture du fichier texte: {str(e)}")
    
    def read_excel_file_with_sessions(self, file_path: str, sheet_name_or_index=0) -> Dict[str, List[str]]:
        """
        Lit un fichier Excel et extrait les références de jobs avec leurs sessions.
        Format attendu: Colonne A = Job, Colonne B = Session
        
        Args:
            file_path: Chemin vers le fichier Excel
            sheet_name_or_index: Nom ou index de la feuille (défaut: première feuille)
            
        Returns:
            Dictionnaire {session_username: [job_ref1, job_ref2, ...]}
        """
        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                "openpyxl n'est pas installé. Installez-le avec: pip install openpyxl"
            )
        
        jobs_by_session = defaultdict(list)
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Obtenir la feuille
            try:
                if isinstance(sheet_name_or_index, int):
                    if sheet_name_or_index >= len(wb.worksheets):
                        raise CommandError(
                            f"Index de feuille {sheet_name_or_index} invalide. "
                            f"Le fichier contient {len(wb.worksheets)} feuille(s)."
                        )
                    sheet = wb.worksheets[sheet_name_or_index]
                else:
                    if sheet_name_or_index not in wb.sheetnames:
                        raise CommandError(
                            f"Feuille '{sheet_name_or_index}' non trouvée. "
                            f"Feuilles disponibles: {', '.join(wb.sheetnames)}"
                        )
                    sheet = wb[sheet_name_or_index]
            except (IndexError, KeyError) as e:
                raise CommandError(f"Erreur lors de l'accès à la feuille: {str(e)}")
            
            # Vérifier que la feuille a au moins une ligne
            if sheet.max_row < 1:
                raise CommandError("La feuille Excel est vide (aucune ligne trouvée)")
            
            # Lire les données depuis les colonnes A (Job) et B (Session)
            # Ignorer la première ligne si c'est un en-tête
            start_row = 1
            first_row_value = sheet.cell(row=1, column=1).value
            if first_row_value and ('Référence' in str(first_row_value) or 'Job' in str(first_row_value)):
                start_row = 2
            
            for row_num in range(start_row, sheet.max_row + 1):
                job_cell = sheet.cell(row=row_num, column=1)
                session_cell = sheet.cell(row=row_num, column=2)
                
                job_ref = str(job_cell.value).strip() if job_cell.value else None
                session = str(session_cell.value).strip().lower() if session_cell.value else None
                
                if job_ref and session and job_ref.lower() not in ['none', 'null', '']:
                    jobs_by_session[session].append(job_ref)
            
            return dict(jobs_by_session)
            
        except CommandError:
            raise
        except Exception as e:
            import traceback
            logger.error(f"Erreur lors de la lecture du fichier Excel {file_path}: {str(e)}\n{traceback.format_exc()}")
            raise CommandError(f"Erreur lors de la lecture du fichier Excel: {str(e)}")
    
    def read_excel_file(self, file_path: str, sheet_name_or_index=0) -> List[str]:
        """
        Lit un fichier Excel et extrait les références de jobs depuis la première colonne.
        
        Args:
            file_path: Chemin vers le fichier Excel
            sheet_name_or_index: Nom ou index de la feuille (défaut: première feuille)
            
        Returns:
            Liste des références de jobs
        """
        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                "openpyxl n'est pas installé. Installez-le avec: pip install openpyxl"
            )
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Obtenir la feuille
            try:
                if isinstance(sheet_name_or_index, int):
                    if sheet_name_or_index >= len(wb.worksheets):
                        raise CommandError(
                            f"Index de feuille {sheet_name_or_index} invalide. "
                            f"Le fichier contient {len(wb.worksheets)} feuille(s)."
                        )
                    sheet = wb.worksheets[sheet_name_or_index]
                else:
                    if sheet_name_or_index not in wb.sheetnames:
                        raise CommandError(
                            f"Feuille '{sheet_name_or_index}' non trouvée. "
                            f"Feuilles disponibles: {', '.join(wb.sheetnames)}"
                        )
                    sheet = wb[sheet_name_or_index]
            except (IndexError, KeyError) as e:
                raise CommandError(f"Erreur lors de l'accès à la feuille: {str(e)}")
            
            # Vérifier que la feuille a au moins une ligne
            if sheet.max_row < 1:
                raise CommandError("La feuille Excel est vide (aucune ligne trouvée)")
            
            # Lire les données depuis la première colonne (colonne A, index 0)
            job_refs = []
            for row in sheet.iter_rows(min_row=1, values_only=False):
                if len(row) > 0:
                    cell_value = row[0].value
                    if cell_value is not None:
                        job_ref = str(cell_value).strip()
                        if job_ref:
                            job_refs.append(job_ref)
            
            return job_refs if job_refs else []
            
        except CommandError:
            raise
        except Exception as e:
            import traceback
            logger.error(f"Erreur lors de la lecture du fichier Excel {file_path}: {str(e)}\n{traceback.format_exc()}")
            raise CommandError(f"Erreur lors de la lecture du fichier Excel: {str(e)}")

