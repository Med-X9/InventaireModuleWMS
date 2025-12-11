"""
Commande Django pour clôturer plusieurs jobs en même temps en utilisant leurs références.

Cette commande suit la même logique que l'API de clôture mais :
- Ignore la validation des personnes
- Force les JobDetail non terminés à TERMINE
- Synchronise les CountingDetail entre les countings 1 et 2
- Vérifie les EcartComptage avant de clôturer le job

Exemple d'utilisation:
    python manage.py close_jobs_by_reference --job-refs JOB-0001 JOB-0002 JOB-0003
    python manage.py close_jobs_by_reference --file jobs_list.txt
    python manage.py close_jobs_by_reference --use-static
    python manage.py close_jobs_by_reference  # Utilise la liste statique par défaut
"""
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from collections import defaultdict
from typing import Optional, List, Tuple, Dict, Any
import logging

from apps.inventory.models import Job, Assigment, JobDetail, CountingDetail, EcartComptage, Counting

logger = logging.getLogger(__name__)

# Liste statique des références de jobs à clôturer
STATIC_JOB_REFERENCES = [
    'JOB-0039',
    'JOB-0269',
    'JOB-0270',
    'JOB-0276',
    'JOB-0277',
    'JOB-0278',
    'JOB-0279',
    'JOB-0280',
    'JOB-0295',
    'JOB-0296',
    'JOB-0319',
    'JOB-0320',
    'JOB-0321',
    'JOB-0322',
    'JOB-0323',
    'JOB-0324',
    'JOB-0325',
    'JOB-0326',
    'JOB-0327',
    'JOB-0328',
    'JOB-0341',
    'JOB-0342',
    'JOB-0353',
    'JOB-0355',
    'JOB-0356',
    'JOB-0359',
    'JOB-0360',
    'JOB-0361',
    'JOB-0362',
    'JOB-0363',
    'JOB-0418',
    'JOB-0420',
    'JOB-0421',
    'JOB-0422',
    'JOB-0442',
    'JOB-0444',
    'JOB-0445',
    'JOB-0446',
    'JOB-0473',
    'JOB-0474',
    'JOB-0476',
    'JOB-0503',
    'JOB-0546',
    'JOB-0547',
    'JOB-0548',
    'JOB-0549',
    'JOB-0550',
    'JOB-0555',
    'JOB-0556',
    'JOB-0557',
]


class Command(BaseCommand):
    help = 'Clôture plusieurs jobs en même temps en utilisant leurs références'

    def add_arguments(self, parser):
        parser.add_argument(
            '--job-refs',
            nargs='+',
            type=str,
            help='Liste des références de jobs à clôturer (ex: JOB-0001 JOB-0002)',
        )
        parser.add_argument(
            '--file',
            type=str,
            help='Chemin vers un fichier Excel (.xlsx) ou texte (.txt) contenant les références de jobs',
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default=0,
            help='Nom ou index de la feuille Excel à lire (défaut: première feuille)',
        )
        parser.add_argument(
            '--use-static',
            action='store_true',
            help='Utiliser la liste statique de références de jobs définie dans le code',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Mode test : affiche ce qui sera fait sans modifier la base de données',
        )

    def handle(self, *args, **options):
        job_refs = options.get('job_refs', [])
        file_path = options.get('file')
        sheet = options.get('sheet', 0)
        dry_run = options.get('dry_run', False)
        use_static = options.get('use_static', False)
        
        # Utiliser la liste statique si demandé
        if use_static:
            job_refs = STATIC_JOB_REFERENCES.copy()
            self.stdout.write(self.style.SUCCESS(f'📋 Utilisation de la liste statique: {len(job_refs)} job(s)'))
        
        # Récupérer les références depuis le fichier si fourni
        if file_path:
            file_refs = []
            try:
                if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                    # Fichier Excel
                    file_refs = self.read_excel_file(file_path, sheet)
                    # S'assurer que file_refs est toujours une liste
                    if file_refs is None or not isinstance(file_refs, list):
                        file_refs = []
                else:
                    # Fichier texte
                    with open(file_path, 'r', encoding='utf-8') as f:
                        file_refs = [line.strip() for line in f if line.strip()]
                        # S'assurer que file_refs est toujours une liste
                        if file_refs is None or not isinstance(file_refs, list):
                            file_refs = []
            except FileNotFoundError:
                raise CommandError(f"Fichier non trouvé: {file_path}")
            except CommandError:
                # Re-raise les CommandError telles quelles
                raise
            except Exception as e:
                # En cas d'erreur, file_refs reste une liste vide
                file_refs = []
                raise CommandError(f"Erreur lors de la lecture du fichier: {str(e)}")
            finally:
                # S'assurer que file_refs est toujours une liste avant d'utiliser extend
                if file_refs is None or not isinstance(file_refs, list):
                    file_refs = []
                if file_refs:  # Si la liste n'est pas vide
                    job_refs.extend(file_refs)
        
        # Si aucune référence n'a été fournie, utiliser la liste statique par défaut
        if not job_refs:
            job_refs = STATIC_JOB_REFERENCES.copy()
            self.stdout.write(self.style.SUCCESS(f'📋 Utilisation de la liste statique par défaut: {len(job_refs)} job(s)'))
        
        if not job_refs:
            raise CommandError(
                "Aucune référence de job trouvée. Utilisez --job-refs, --file, ou --use-static"
            )
        
        # Supprimer les doublons
        job_refs = list(set(job_refs))
        
        self.stdout.write(self.style.SUCCESS(f'📋 Clôture de {len(job_refs)} job(s)'))
        if dry_run:
            self.stdout.write(self.style.WARNING('⚠️  MODE DRY-RUN : Aucune modification ne sera effectuée'))
        
        # Récupérer les jobs par leurs références
        jobs = Job.objects.filter(reference__in=job_refs).select_related('inventory', 'warehouse')
        
        found_refs = set(jobs.values_list('reference', flat=True))
        missing_refs = set(job_refs) - found_refs
        
        if missing_refs:
            self.stdout.write(self.style.ERROR(
                f'❌ Jobs non trouvés: {", ".join(sorted(missing_refs))}'
            ))
        
        if not jobs.exists():
            self.stdout.write(self.style.ERROR('❌ Aucun job trouvé'))
            return
        
        self.stdout.write(f'  ✅ Jobs trouvés: {jobs.count()}')
        
        # Statistiques
        stats = {
            'jobs_processed': 0,
            'jobs_closed': 0,
            'assignments_closed': 0,
            'job_details_closed': 0,
            'duplicates_found': 0,
            'duplicates_removed': 0,
            'counting_details_synced': 0,
            'jobs_skipped_ecarts': 0,
            'jobs_skipped_non_terminated': 0,
            'errors': []
        }
        
        if dry_run:
            self.stdout.write('\n' + '='*60)
            self.stdout.write('SIMULATION - Ce qui sera fait:')
            self.stdout.write('='*60)
        
        # Traiter chaque job
        for job in jobs:
            try:
                result = self.process_job(job, dry_run)
                stats['jobs_processed'] += 1
                stats['jobs_closed'] += result['job_closed']
                stats['assignments_closed'] += result['assignments_closed']
                stats['job_details_closed'] += result['job_details_closed']
                stats['duplicates_found'] += result['duplicates_found']
                stats['duplicates_removed'] += result['duplicates_removed']
                stats['counting_details_synced'] += result['counting_details_synced']
                stats['jobs_skipped_ecarts'] += result['job_skipped_ecarts']
                stats['jobs_skipped_non_terminated'] += result['job_skipped_non_terminated']
            except Exception as e:
                error_msg = f"Erreur lors du traitement du job {job.reference}: {str(e)}"
                stats['errors'].append(error_msg)
                self.stdout.write(self.style.ERROR(f'  ❌ {error_msg}'))
                logger.error(error_msg, exc_info=True)
        
        # Afficher le résumé
        self.stdout.write('\n' + '='*60)
        self.stdout.write(self.style.SUCCESS('📊 RÉSUMÉ'))
        self.stdout.write('='*60)
        self.stdout.write(f"  Jobs traités: {stats['jobs_processed']}")
        self.stdout.write(f"  Jobs clôturés: {stats['jobs_closed']}")
        self.stdout.write(f"  Jobs ignorés (écarts): {stats['jobs_skipped_ecarts']}")
        self.stdout.write(f"  Jobs ignorés (JobDetail non TERMINE): {stats['jobs_skipped_non_terminated']}")
        self.stdout.write(f"  Assignments clôturés: {stats['assignments_closed']}")
        self.stdout.write(f"  JobDetails clôturés: {stats['job_details_closed']}")
        self.stdout.write(f"  CountingDetails synchronisés: {stats['counting_details_synced']}")
        self.stdout.write(f"  Doublons détectés: {stats['duplicates_found']}")
        self.stdout.write(f"  Doublons supprimés: {stats['duplicates_removed']}")
        
        if stats['errors']:
            self.stdout.write(self.style.ERROR(f"\n  ❌ Erreurs: {len(stats['errors'])}"))
            for error in stats['errors']:
                self.stdout.write(self.style.ERROR(f"    - {error}"))
        
        if dry_run:
            self.stdout.write(self.style.WARNING('\n⚠️  MODE DRY-RUN : Aucune modification n\'a été effectuée'))
        else:
            self.stdout.write(self.style.SUCCESS('\n✅ Traitement terminé avec succès!'))
    
    @transaction.atomic
    def process_job(self, job: Job, dry_run: bool = False) -> dict:
        """
        Traite un job en suivant la logique de l'API de clôture :
        0. Vérifie que tous les JobDetail sont TERMINE (sinon ignore le job)
        1. Gère les doublons JobDetail
        2. Force les JobDetail à TERMINE
        3. Clôture les Assignments
        4. Synchronise les CountingDetail entre countings 1 et 2
        5. Vérifie les EcartComptage avant de clôturer le job
        
        Args:
            job: Le job à traiter
            dry_run: Si True, ne fait que simuler les actions
            
        Returns:
            Dictionnaire avec les statistiques du traitement
        """
        result = {
            'job_closed': 0,
            'assignments_closed': 0,
            'job_details_closed': 0,
            'duplicates_found': 0,
            'duplicates_removed': 0,
            'counting_details_synced': 0,
            'job_skipped_ecarts': 0,
            'job_skipped_non_terminated': 0
        }
        
        self.stdout.write(f'\n📦 Job: {job.reference} (ID: {job.id})')
        
        # 0. Vérifier que tous les JobDetail sont TERMINE, sinon ignorer le job
        all_job_details = JobDetail.objects.filter(job=job).select_related('location', 'counting')
        non_terminated_job_details = all_job_details.exclude(status='TERMINE')
        
        if non_terminated_job_details.exists():
            non_terminated_count = non_terminated_job_details.count()
            total_count = all_job_details.count()
            result['job_skipped_non_terminated'] = 1
            self.stdout.write(
                self.style.WARNING(
                    f'  ⚠️  Job {job.reference} ignoré : {non_terminated_count}/{total_count} JobDetail(s) non TERMINE'
                )
            )
            # Afficher les JobDetail non terminés pour information
            for jd in non_terminated_job_details[:10]:  # Limiter à 10 pour éviter trop de logs
                self.stdout.write(
                    f'    - JobDetail {jd.reference} (location: {jd.location.location_reference if jd.location else "N/A"}, statut: {jd.status})'
                )
            if non_terminated_count > 10:
                self.stdout.write(f'    ... et {non_terminated_count - 10} autre(s)')
            return result
        
        # 1. Détecter et gérer les JobDetail dupliqués
        # Un doublon = même job, même location, même counting
        all_job_details = all_job_details.order_by('-id')
        
        # Grouper par (location_id, counting_id)
        job_details_by_key = defaultdict(list)
        for jd in all_job_details:
            key = (jd.location_id, jd.counting_id if jd.counting_id else None)
            job_details_by_key[key].append(jd)
        
        # Identifier les doublons (plus d'un JobDetail pour la même combinaison)
        duplicates_to_remove = []
        unique_job_details = []
        
        for key, job_details_list in job_details_by_key.items():
            if len(job_details_list) > 1:
                # Doublons détectés : garder le plus récent (dernier créé, donc id le plus élevé)
                # Trier par id décroissant pour avoir le plus récent en premier
                job_details_list.sort(key=lambda x: x.id, reverse=True)
                kept = job_details_list[0]
                duplicates = job_details_list[1:]
                
                result['duplicates_found'] += len(duplicates)
                unique_job_details.append(kept)
                
                for dup in duplicates:
                    duplicates_to_remove.append(dup)
                    self.stdout.write(
                        f'  ⚠️  Doublon détecté: JobDetail {dup.reference} '
                        f'(location_id={dup.location_id}, counting_id={dup.counting_id}) '
                        f'- sera supprimé (gardé: {kept.reference})'
                    )
            else:
                unique_job_details.append(job_details_list[0])
        
        # Supprimer les doublons
        if duplicates_to_remove:
            if not dry_run:
                JobDetail.objects.filter(id__in=[d.id for d in duplicates_to_remove]).delete()
                result['duplicates_removed'] = len(duplicates_to_remove)
                self.stdout.write(f'  ✅ {len(duplicates_to_remove)} doublon(s) supprimé(s)')
            else:
                result['duplicates_removed'] = len(duplicates_to_remove)
                self.stdout.write(f'  [DRY-RUN] {len(duplicates_to_remove)} doublon(s) seraient supprimé(s)')
        
        # 2. Mettre tous les JobDetail uniques à TERMINE
        now = timezone.now()
        job_details_to_update = []
        
        for jd in unique_job_details:
            if jd.status != 'TERMINE':
                jd.status = 'TERMINE'
                jd.termine_date = now
                job_details_to_update.append(jd)
                result['job_details_closed'] += 1
                self.stdout.write(f'  ✅ JobDetail {jd.reference} → TERMINE')
        
        if job_details_to_update and not dry_run:
            JobDetail.objects.bulk_update(job_details_to_update, ['status', 'termine_date'])
        
        # 3. Clôturer les Assignments et synchroniser les CountingDetail
        assignments = Assigment.objects.filter(job=job).select_related('counting')
        
        for assignment in assignments:
            if assignment.status != 'TERMINE':
                # Clôturer l'assignment
                if not dry_run:
                    assignment.status = 'TERMINE'
                    assignment.termine_date = now
                    assignment.save()  # Sauvegarder immédiatement pour la synchronisation
                result['assignments_closed'] += 1
                self.stdout.write(f'  ✅ Assignment {assignment.reference} → TERMINE')
                
                # Synchroniser les CountingDetail si nécessaire (même logique que l'API)
                if not dry_run:
                    sync_result = self._synchronize_counting_details_if_needed(job, assignment)
                    if sync_result.get('synchronized', False):
                        result['counting_details_synced'] += sync_result.get('created_count', 0)
                        self.stdout.write(
                            f'  🔄 {sync_result.get("created_count", 0)} CountingDetail synchronisé(s) '
                            f'entre les countings {sync_result.get("counting_1_order")} et {sync_result.get("counting_2_order")}'
                        )
        
        # 4. Vérifier si tous les assignments sont terminés
        all_assignments = Assigment.objects.filter(job=job)
        all_assignments_terminated = all_assignments.exclude(status='TERMINE').count() == 0
        
        # 5. Vérifier si tous les EcartComptage de l'inventaire ont un final_result non null
        ecart_comptages = EcartComptage.objects.filter(inventory=job.inventory)
        if ecart_comptages.exists():
            all_ecarts_have_final_result = ecart_comptages.filter(final_result__isnull=True).count() == 0
        else:
            all_ecarts_have_final_result = True
        
        # 6. Clôturer le job seulement si toutes les conditions sont remplies
        if all_assignments_terminated and all_ecarts_have_final_result:
            if job.status != 'TERMINE':
                if not dry_run:
                    job.status = 'TERMINE'
                    job.termine_date = now
                    job.save()
                result['job_closed'] = 1
                self.stdout.write(f'  ✅ Job {job.reference} → TERMINE')
        else:
            result['job_skipped_ecarts'] = 1
            reasons = []
            if not all_assignments_terminated:
                reasons.append('assignments non terminés')
            if not all_ecarts_have_final_result:
                reasons.append('écarts sans final_result')
            self.stdout.write(
                self.style.WARNING(
                    f'  ⚠️  Job {job.reference} non clôturé: {", ".join(reasons)}'
                )
            )
        
        return result
    
    def _synchronize_counting_details_if_needed(
        self, 
        job: Job, 
        assignment: Assigment
    ) -> dict:
        """
        Synchronise les CountingDetail entre les deux countings (order 1 et 2) si nécessaire.
        Même logique que l'API.
        
        Args:
            job: Le job concerné
            assignment: L'assignment qui vient d'être clôturé
            
        Returns:
            Dictionnaire avec les informations de synchronisation
        """
        counting_order = assignment.counting.order
        
        # Ne synchroniser que si counting_order est 1 ou 2
        if counting_order not in [1, 2]:
            return {
                'synchronized': False,
                'reason': f'Counting order {counting_order} ne nécessite pas de synchronisation'
            }
        
        # Déterminer l'autre counting_order
        other_counting_order = 2 if counting_order == 1 else 1
        
        # Récupérer l'assignment de l'autre counting pour ce job
        try:
            other_assignment = Assigment.objects.get(
                job=job,
                counting__order=other_counting_order
            )
        except Assigment.DoesNotExist:
            return {
                'synchronized': False,
                'reason': f'Assignment avec counting order {other_counting_order} non trouvé pour ce job'
            }
        
        # Vérifier si l'autre assignment est TERMINE
        if other_assignment.status != 'TERMINE':
            return {
                'synchronized': False,
                'reason': f'Assignment avec counting order {other_counting_order} n\'est pas encore TERMINE'
            }
        
        # Les deux assignments sont TERMINE, on peut synchroniser
        counting_1 = assignment.counting if counting_order == 1 else other_assignment.counting
        counting_2 = other_assignment.counting if counting_order == 1 else assignment.counting
        
        # Récupérer tous les CountingDetail des deux countings pour ce job
        counting_details_1 = CountingDetail.objects.filter(
            job=job,
            counting=counting_1
        ).select_related('location', 'product')
        
        counting_details_2 = CountingDetail.objects.filter(
            job=job,
            counting=counting_2
        ).select_related('location', 'product')
        
        # Créer des sets pour la comparaison rapide
        # Clé de comparaison : (location_id, product_id, dlc, n_lot)
        def get_comparison_key(cd: CountingDetail) -> Tuple[int, Optional[int], Optional[str], Optional[str]]:
            return (
                cd.location_id,
                cd.product_id if cd.product else None,
                cd.dlc.isoformat() if cd.dlc else None,
                cd.n_lot or None
            )
        
        # Créer des dictionnaires pour accès rapide
        details_1_dict = {get_comparison_key(cd): cd for cd in counting_details_1}
        details_2_dict = {get_comparison_key(cd): cd for cd in counting_details_2}
        
        # Trouver les lignes manquantes dans counting_2
        missing_in_2 = []
        for key, cd_1 in details_1_dict.items():
            if key not in details_2_dict:
                missing_in_2.append({
                    'location': cd_1.location,
                    'product': cd_1.product,
                    'dlc': cd_1.dlc,
                    'n_lot': cd_1.n_lot,
                    'counting': counting_2,
                    'job': job
                })
        
        # Trouver les lignes manquantes dans counting_1
        missing_in_1 = []
        for key, cd_2 in details_2_dict.items():
            if key not in details_1_dict:
                missing_in_1.append({
                    'location': cd_2.location,
                    'product': cd_2.product,
                    'dlc': cd_2.dlc,
                    'n_lot': cd_2.n_lot,
                    'counting': counting_1,
                    'job': job
                })
        
        # Créer les CountingDetail manquants avec quantité 0
        created_count = 0
        counting_details_to_create = []
        
        for item in missing_in_2:
            counting_detail = CountingDetail(
                quantity_inventoried=0,
                product=item['product'],
                dlc=item['dlc'],
                n_lot=item['n_lot'],
                location=item['location'],
                counting=item['counting'],
                job=item['job'],
                last_synced_at=timezone.now()
            )
            counting_detail.reference = counting_detail.generate_reference(CountingDetail.REFERENCE_PREFIX)
            counting_details_to_create.append(counting_detail)
            created_count += 1
        
        for item in missing_in_1:
            counting_detail = CountingDetail(
                quantity_inventoried=0,
                product=item['product'],
                dlc=item['dlc'],
                n_lot=item['n_lot'],
                location=item['location'],
                counting=item['counting'],
                job=item['job'],
                last_synced_at=timezone.now()
            )
            counting_detail.reference = counting_detail.generate_reference(CountingDetail.REFERENCE_PREFIX)
            counting_details_to_create.append(counting_detail)
            created_count += 1
        
        # Bulk create si nécessaire
        if counting_details_to_create:
            CountingDetail.objects.bulk_create(counting_details_to_create)
            
            # Régénérer les références avec les IDs réels
            for cd in counting_details_to_create:
                cd.reference = cd.generate_reference(CountingDetail.REFERENCE_PREFIX)
            CountingDetail.objects.bulk_update(counting_details_to_create, ['reference'])
        
        return {
            'synchronized': True,
            'created_count': created_count,
            'counting_1_order': counting_1.order,
            'counting_2_order': counting_2.order
        }
    
    def read_excel_file(self, file_path: str, sheet_name_or_index=0) -> List[str]:
        """
        Lit un fichier Excel et extrait les références de jobs depuis la première colonne.
        
        Args:
            file_path: Chemin vers le fichier Excel
            sheet_name_or_index: Nom ou index de la feuille (défaut: première feuille)
            
        Returns:
            Liste des références de jobs
        """
        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                "openpyxl n'est pas installé. Installez-le avec: pip install openpyxl"
            )
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Obtenir la feuille
            try:
                if isinstance(sheet_name_or_index, int):
                    if sheet_name_or_index >= len(wb.worksheets):
                        raise CommandError(
                            f"Index de feuille {sheet_name_or_index} invalide. "
                            f"Le fichier contient {len(wb.worksheets)} feuille(s)."
                        )
                    sheet = wb.worksheets[sheet_name_or_index]
                else:
                    if sheet_name_or_index not in wb.sheetnames:
                        raise CommandError(
                            f"Feuille '{sheet_name_or_index}' non trouvée. "
                            f"Feuilles disponibles: {', '.join(wb.sheetnames)}"
                        )
                    sheet = wb[sheet_name_or_index]
            except (IndexError, KeyError) as e:
                raise CommandError(f"Erreur lors de l'accès à la feuille: {str(e)}")
            
            # Vérifier que la feuille a au moins une ligne
            if sheet.max_row < 1:
                raise CommandError("La feuille Excel est vide (aucune ligne trouvée)")
            
            # Lire les données depuis la première colonne (colonne A, index 0)
            # On lit toutes les lignes, en ignorant les cellules vides
            job_refs = []
            for row in sheet.iter_rows(min_row=1, values_only=False):
                if len(row) > 0:
                    cell_value = row[0].value  # Première colonne (index 0)
                    if cell_value is not None:
                        job_ref = str(cell_value).strip()
                        if job_ref:
                            job_refs.append(job_ref)
            
            # Toujours retourner une liste, même si vide
            return job_refs if job_refs else []
            
        except CommandError:
            # Re-raise les CommandError telles quelles
            raise
        except Exception as e:
            # Logger l'erreur pour le débogage
            import traceback
            logger.error(f"Erreur lors de la lecture du fichier Excel {file_path}: {str(e)}\n{traceback.format_exc()}")
            raise CommandError(f"Erreur lors de la lecture du fichier Excel: {str(e)}")

