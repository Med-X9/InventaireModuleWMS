"""
Commande Django pour exporter les résultats d'inventaire en Excel.

Cette commande :
- Récupère les résultats d'un inventaire pour un entrepôt spécifique
- Génère un fichier Excel avec les résultats
- Utilise la même logique que InventoryResultExportExcelView

Exemple d'utilisation:
    python manage.py export_inventory_results_to_excel --inventory-id 2 --warehouse-id 1
    python manage.py export_inventory_results_to_excel --inventory-id 2 --warehouse-id 1 --file ./resultats.xlsx
"""
from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
import os
import logging
from pathlib import Path
from io import BytesIO

from apps.inventory.models import Inventory
from apps.inventory.services.inventory_result_service import InventoryResultService
from apps.inventory.exceptions.inventory_exceptions import (
    InventoryNotFoundError,
    InventoryValidationError
)

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Exporte les résultats d\'inventaire en Excel pour un entrepôt spécifique'

    def add_arguments(self, parser):
        parser.add_argument(
            '--inventory-id',
            type=int,
            required=True,
            help='ID de l\'inventaire',
        )
        parser.add_argument(
            '--warehouse-id',
            type=int,
            required=True,
            help='ID de l\'entrepôt',
        )
        parser.add_argument(
            '--file',
            type=str,
            help='Chemin vers le fichier Excel de sortie (.xlsx). Si non fourni, un nom sera généré automatiquement',
        )

    def handle(self, *args, **options):
        inventory_id = options['inventory_id']
        warehouse_id = options['warehouse_id']
        file_path = options.get('file')
        
        try:
            # Récupérer les résultats via le service
            service = InventoryResultService()
            results = service.get_inventory_results_for_warehouse(
                inventory_id=inventory_id,
                warehouse_id=warehouse_id,
            )
            
            if not results:
                self.stdout.write(
                    self.style.WARNING(
                        '⚠️  Aucun résultat trouvé pour cet inventaire et cet entrepôt'
                    )
                )
                return
            
            self.stdout.write(
                self.style.SUCCESS(
                    f'📋 {len(results)} résultat(s) trouvé(s) pour l\'inventaire {inventory_id} et l\'entrepôt {warehouse_id}'
                )
            )
            
            # Générer le fichier Excel
            excel_buffer = self._generate_excel(results, inventory_id, warehouse_id)
            
            # Générer le nom de fichier si non fourni
            if not file_path:
                try:
                    inventory = Inventory.objects.get(id=inventory_id)
                    inventory_ref = inventory.reference.replace(' ', '_')
                except Inventory.DoesNotExist:
                    inventory_ref = f"inventaire_{inventory_id}"
                
                # Créer un nom de fichier avec timestamp
                timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                filename = f"resultats_{inventory_ref}_warehouse_{warehouse_id}_{timestamp}.xlsx"
                file_path = filename
            
            # S'assurer que le chemin est absolu
            if not os.path.isabs(file_path):
                file_path = os.path.abspath(file_path)
            
            # Créer le dossier parent s'il n'existe pas
            parent_dir = os.path.dirname(file_path)
            if parent_dir:
                Path(parent_dir).mkdir(parents=True, exist_ok=True)
            
            # Sauvegarder le fichier Excel
            with open(file_path, 'wb') as f:
                f.write(excel_buffer.getvalue())
            
            self.stdout.write(
                self.style.SUCCESS(
                    f'✅ Fichier Excel généré avec succès: {file_path}'
                )
            )
            
        except InventoryNotFoundError as e:
            raise CommandError(f"Inventaire introuvable: {str(e)}")
        except InventoryValidationError as e:
            raise CommandError(f"Erreur de validation: {str(e)}")
        except ValueError as e:
            raise CommandError(f"Erreur de valeur: {str(e)}")
        except ImportError as e:
            raise CommandError(
                f"Dépendance manquante pour l'export Excel: {str(e)}. "
                "Installez pandas et openpyxl avec: pip install pandas openpyxl"
            )
        except Exception as e:
            logger.error(
                f"Erreur lors de l'export Excel (inventory_id={inventory_id}, warehouse_id={warehouse_id}): {str(e)}",
                exc_info=True
            )
            raise CommandError(f"Une erreur inattendue est survenue: {str(e)}")
    
    def _generate_excel(self, results, inventory_id: int, warehouse_id: int):
        """
        Génère un fichier Excel à partir des résultats.
        Utilise la même logique que InventoryResultExportExcelView._generate_excel
        
        Args:
            results: Liste des résultats à exporter
            inventory_id: ID de l'inventaire
            warehouse_id: ID de l'entrepôt
            
        Returns:
            BytesIO: Buffer contenant le fichier Excel
        """
        try:
            import pandas as pd
        except ImportError:
            raise ImportError(
                "pandas est requis pour l'export Excel. "
                "Installez-le avec: pip install pandas"
            )
        
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl est requis pour l'export Excel. "
                "Installez-le avec: pip install openpyxl"
            )
        
        import io
        
        # Vérifier que les résultats ne sont pas vides
        if not results:
            raise ValueError("Aucune donnée à exporter")
        
        # Mettre le code interne dans la clé product si présent
        transformed_results = []
        for result in results:
            transformed_result = dict(result)
            if 'product_internal_code' in transformed_result:
                transformed_result['product'] = transformed_result.pop('product_internal_code')
            transformed_results.append(transformed_result)
        
        # Convertir les résultats en DataFrame
        try:
            df = pd.DataFrame(transformed_results)
        except Exception as e:
            logger.error(f"Erreur lors de la conversion en DataFrame: {e}", exc_info=True)
            raise ValueError(f"Impossible de convertir les résultats en DataFrame: {str(e)}")
        
        # Vérifier que le DataFrame n'est pas vide
        if df.empty:
            raise ValueError("Le DataFrame est vide après conversion")
        
        # Colonnes à exclure de l'export
        excluded_columns = ['location_id', 'job_id', 'ecart_comptage_id']
        
        # Réorganiser les colonnes pour un meilleur affichage
        column_order = []
        
        # Colonnes de base
        base_columns = ['location']
        
        # Ajouter job_reference si présent
        if 'job_reference' in df.columns:
            base_columns.append('job_reference')
        
        # Colonnes produit (si présentes)
        product_columns = []
        if 'product' in df.columns:
            product_columns.append('product')
        if 'product_description' in df.columns:
            product_columns.append('product_description')
        
        # Colonnes de comptage (triées par ordre)
        counting_columns = sorted([col for col in df.columns if col.endswith(' comptage')], 
                                 key=lambda x: int(x.split()[0]) if x.split()[0].isdigit() else 999)
        
        # Colonnes d'écart (triées)
        ecart_columns = sorted([col for col in df.columns if col.startswith('ecart_')],
                              key=lambda x: tuple(map(int, x.split('_')[1:])) if all(part.isdigit() for part in x.split('_')[1:]) else (999, 999))
        
        # Colonnes finales
        final_columns = []
        if 'final_result' in df.columns:
            final_columns.append('final_result')
        if 'manual_result' in df.columns:
            final_columns.append('manual_result')
        if 'resolved' in df.columns:
            final_columns.append('resolved')
        
        # Construire l'ordre final
        column_order = base_columns + product_columns + counting_columns + ecart_columns + final_columns
        
        # Filtrer les colonnes exclues
        column_order = [col for col in column_order if col not in excluded_columns]
        
        # Réorganiser le DataFrame
        existing_columns = [col for col in column_order if col in df.columns]
        existing_columns = [col for col in existing_columns if col not in excluded_columns]
        
        if not existing_columns:
            existing_columns = [col for col in df.columns if col not in excluded_columns]
        
        df = df[existing_columns]
        
        # Mapping des noms de colonnes
        column_name_mapping = {
            'location': 'Emplacement',
            'job_reference': 'Référence Job',
            'product': 'Code Interne Article',
            'product_description': 'Description Article',
            'final_result': 'Résultat Final',
            'manual_result': 'Résultat Manuel',
            'resolved': 'Résolu',
        }
        
        # Ajouter le mapping pour les colonnes de comptage dynamiquement
        for col in df.columns:
            if col.endswith(' comptage'):
                order_num = col.split()[0]
                column_name_mapping[col] = f'{order_num.capitalize()} Comptage'
            elif col.startswith('ecart_'):
                parts = col.split('_')
                if len(parts) >= 3 and parts[1].isdigit() and parts[2].isdigit():
                    column_name_mapping[col] = f'Écart Comptage {parts[1]}-{parts[2]}'
        
        # Renommer les colonnes
        df = df.rename(columns=column_name_mapping)
        
        # Créer le fichier Excel en mémoire
        excel_buffer = io.BytesIO()
        try:
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Résultats')
                
                # Récupérer la feuille pour appliquer des formats
                worksheet = writer.sheets['Résultats']
                
                # Ajuster la largeur des colonnes
                from openpyxl.utils import get_column_letter
                for idx, col in enumerate(df.columns, start=1):
                    try:
                        col_data = df[col].fillna('').astype(str)
                        if len(col_data) > 0:
                            max_data_length = col_data.map(len).max()
                        else:
                            max_data_length = 0
                        max_length = max(max_data_length, len(str(col)))
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
                    except Exception as e:
                        logger.warning(f"Impossible d'ajuster la largeur de la colonne {col}: {e}")
                        worksheet.column_dimensions[get_column_letter(idx)].width = 15
        except Exception as e:
            logger.error(f"Erreur lors de la génération du fichier Excel: {e}", exc_info=True)
            raise ValueError(f"Impossible de générer le fichier Excel: {str(e)}")
        
        excel_buffer.seek(0)
        return excel_buffer

