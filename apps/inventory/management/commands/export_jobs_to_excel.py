"""
Commande Django pour exporter tous les jobs de la base de données dans un fichier Excel.

Format Excel généré:
- job_reference: Référence du job (ex: JOB-0001)
- session_comptage_1: Username de la session affectée au 1er comptage (ou vide si aucune)
- session_comptage_2: Username de la session affectée au 2ème comptage (ou vide si aucune)

Exemple d'utilisation:
    python manage.py export_jobs_to_excel --file path/to/export.xlsx
"""
from django.core.management.base import BaseCommand, CommandError
from django.db.models import Prefetch
import logging

from apps.inventory.models import Job, Assigment, Counting

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Exporte tous les jobs avec leurs sessions affectées pour les comptages 1 et 2 dans un fichier Excel'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            required=True,
            help='Chemin vers le fichier Excel de sortie (.xlsx)',
        )

    def handle(self, *args, **options):
        file_path = options['file']
        
        self.stdout.write(self.style.SUCCESS(f'📊 Export des jobs vers Excel: {file_path}'))
        
        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                "openpyxl n'est pas installé. Installez-le avec: pip install openpyxl"
            )
        
        # Récupérer tous les jobs avec leurs assignments optimisés
        # Précharger les relations pour éviter les requêtes N+1
        jobs = Job.objects.prefetch_related(
            Prefetch(
                'assigment_set',
                queryset=Assigment.objects.select_related(
                    'counting',
                    'session'
                ).filter(
                    counting__order__in=[1, 2]
                ).order_by('counting__order')
            )
        ).order_by('reference').all()
        
        total_jobs = jobs.count()
        self.stdout.write(f'  📦 Nombre de jobs à exporter: {total_jobs}')
        
        if total_jobs == 0:
            self.stdout.write(self.style.WARNING('⚠️  Aucun job trouvé dans la base de données'))
            return
        
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jobs Export"
        
        # En-têtes
        headers = ['job_reference', 'session_comptage_1', 'session_comptage_2']
        ws.append(headers)
        
        # Style des en-têtes
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Parcourir les jobs et remplir les données
        jobs_exported = 0
        jobs_with_sessions = 0
        
        for job in jobs:
            # Récupérer les assignments pour les comptages 1 et 2
            assignments = job.assigment_set.all()
            
            session_1 = None
            session_2 = None
            
            for assignment in assignments:
                if assignment.counting and assignment.counting.order == 1:
                    if assignment.session:
                        session_1 = assignment.session.username
                elif assignment.counting and assignment.counting.order == 2:
                    if assignment.session:
                        session_2 = assignment.session.username
            
            # Ajouter une ligne au fichier Excel
            row_data = [
                job.reference,
                session_1 or '',
                session_2 or ''
            ]
            ws.append(row_data)
            
            jobs_exported += 1
            if session_1 or session_2:
                jobs_with_sessions += 1
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 20  # job_reference
        ws.column_dimensions['B'].width = 30  # session_comptage_1
        ws.column_dimensions['C'].width = 30  # session_comptage_2
        
        # Sauvegarder le fichier
        try:
            wb.save(file_path)
            self.stdout.write(self.style.SUCCESS(f'✅ Export terminé avec succès!'))
            self.stdout.write(f'  📄 Fichier: {file_path}')
            self.stdout.write(f'  📊 Jobs exportés: {jobs_exported}')
            self.stdout.write(f'  👥 Jobs avec sessions: {jobs_with_sessions}')
        except Exception as e:
            raise CommandError(f"Erreur lors de la sauvegarde du fichier Excel: {str(e)}")

