"""
Commande Django pour exporter les résultats d'inventaire en Excel avec les statuts des jobs et assignments.

Cette commande :
- Récupère les résultats d'un inventaire pour un entrepôt spécifique
- Enrichit les données avec les statuts des jobs et des assignments
- Génère un fichier Excel avec toutes ces informations

Exemple d'utilisation:
    python manage.py export_inventory_results_with_statuses_to_excel --inventory-id 2 --warehouse-id 1
    python manage.py export_inventory_results_with_statuses_to_excel --inventory-id 2 --warehouse-id 1 --file ./resultats_avec_statuts.xlsx
"""
from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from django.db.models import Q
import os
import logging
from pathlib import Path
from io import BytesIO
from collections import defaultdict

from apps.inventory.models import Inventory, Job, Assigment, Counting
from apps.inventory.services.inventory_result_service import InventoryResultService
from apps.inventory.exceptions.inventory_exceptions import (
    InventoryNotFoundError,
    InventoryValidationError
)

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Exporte les résultats d\'inventaire en Excel avec les statuts des jobs et assignments'

    def add_arguments(self, parser):
        parser.add_argument(
            '--inventory-id',
            type=int,
            required=True,
            help='ID de l\'inventaire',
        )
        parser.add_argument(
            '--warehouse-id',
            type=int,
            required=True,
            help='ID de l\'entrepôt',
        )
        parser.add_argument(
            '--file',
            type=str,
            help='Chemin vers le fichier Excel de sortie (.xlsx). Si non fourni, un nom sera généré automatiquement',
        )

    def handle(self, *args, **options):
        inventory_id = options['inventory_id']
        warehouse_id = options['warehouse_id']
        file_path = options.get('file')
        
        try:
            # Récupérer les résultats via le service
            service = InventoryResultService()
            results = service.get_inventory_results_for_warehouse(
                inventory_id=inventory_id,
                warehouse_id=warehouse_id,
            )
            
            if not results:
                self.stdout.write(
                    self.style.WARNING(
                        '⚠️  Aucun résultat trouvé pour cet inventaire et cet entrepôt'
                    )
                )
                return
            
            self.stdout.write(
                self.style.SUCCESS(
                    f'📋 {len(results)} résultat(s) trouvé(s) pour l\'inventaire {inventory_id} et l\'entrepôt {warehouse_id}'
                )
            )
            
            # Enrichir les résultats avec les statuts des jobs et assignments
            self.stdout.write('  🔄 Enrichissement des données avec les statuts des jobs et assignments...')
            enriched_results = self._enrich_results_with_statuses(results, inventory_id, warehouse_id)
            self.stdout.write(f'  ✅ {len(enriched_results)} ligne(s) enrichie(s)')
            
            # Générer le fichier Excel
            excel_buffer = self._generate_excel(enriched_results, inventory_id, warehouse_id)
            
            # Générer le nom de fichier si non fourni
            if not file_path:
                try:
                    inventory = Inventory.objects.get(id=inventory_id)
                    inventory_ref = inventory.reference.replace(' ', '_')
                except Inventory.DoesNotExist:
                    inventory_ref = f"inventaire_{inventory_id}"
                
                # Créer un nom de fichier avec timestamp
                timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
                filename = f"resultats_avec_statuts_{inventory_ref}_warehouse_{warehouse_id}_{timestamp}.xlsx"
                file_path = filename
            
            # S'assurer que le chemin est absolu
            if not os.path.isabs(file_path):
                file_path = os.path.abspath(file_path)
            
            # Créer le dossier parent s'il n'existe pas
            parent_dir = os.path.dirname(file_path)
            if parent_dir:
                Path(parent_dir).mkdir(parents=True, exist_ok=True)
            
            # Sauvegarder le fichier Excel
            with open(file_path, 'wb') as f:
                f.write(excel_buffer.getvalue())
            
            self.stdout.write(
                self.style.SUCCESS(
                    f'✅ Fichier Excel généré avec succès: {file_path}'
                )
            )
            
        except InventoryNotFoundError as e:
            raise CommandError(f"Inventaire introuvable: {str(e)}")
        except InventoryValidationError as e:
            raise CommandError(f"Erreur de validation: {str(e)}")
        except ValueError as e:
            raise CommandError(f"Erreur de valeur: {str(e)}")
        except ImportError as e:
            raise CommandError(
                f"Dépendance manquante pour l'export Excel: {str(e)}. "
                "Installez pandas et openpyxl avec: pip install pandas openpyxl"
            )
        except Exception as e:
            logger.error(
                f"Erreur lors de l'export Excel (inventory_id={inventory_id}, warehouse_id={warehouse_id}): {str(e)}",
                exc_info=True
            )
            raise CommandError(f"Une erreur inattendue est survenue: {str(e)}")
    
    def _enrich_results_with_statuses(
        self, 
        results: list, 
        inventory_id: int, 
        warehouse_id: int
    ) -> list:
        """
        Enrichit les résultats avec les statuts des jobs et des assignments.
        
        Les résultats peuvent déjà contenir les statuts des assignments pour les comptages 1 et 2
        dans les colonnes 'statut_1er_comptage' et 'statut_2er_comptage'.
        On ajoute :
        - Le statut du job
        - Les statuts des assignments pour tous les comptages (standardisation)
        
        Args:
            results: Liste des résultats d'inventaire
            inventory_id: ID de l'inventaire
            warehouse_id: ID de l'entrepôt
            
        Returns:
            Liste des résultats enrichis avec les statuts
        """
        # Récupérer tous les job_ids uniques des résultats
        job_ids = set()
        for result in results:
            if result.get('job_id'):
                job_ids.add(result['job_id'])
        
        if not job_ids:
            # Si aucun job_id, retourner les résultats tels quels
            return results
        
        # Récupérer tous les jobs avec leurs assignments en une seule requête
        jobs = Job.objects.filter(
            id__in=job_ids,
            inventory_id=inventory_id,
            warehouse_id=warehouse_id
        ).select_related(
            'inventory',
            'warehouse'
        ).prefetch_related(
            'assigment_set__counting',
            'assigment_set__session'
        )
        
        # Créer un dictionnaire pour accès rapide : job_id -> job
        jobs_dict = {job.id: job for job in jobs}
        
        # Créer un dictionnaire pour les assignments par job et counting order
        # Structure: {job_id: {counting_order: assignment}}
        assignments_dict = defaultdict(dict)
        
        for job in jobs:
            for assignment in job.assigment_set.all():
                if assignment.counting:
                    counting_order = assignment.counting.order
                    assignments_dict[job.id][counting_order] = assignment
        
        # Enrichir les résultats
        enriched_results = []
        for result in results:
            enriched_result = dict(result)
            job_id = result.get('job_id')
            
            if job_id and job_id in jobs_dict:
                job = jobs_dict[job_id]
                
                # Ajouter le statut du job
                enriched_result['job_status'] = job.status
                
                # Ajouter les statuts des assignments pour chaque counting order
                job_assignments = assignments_dict.get(job_id, {})
                
                # Pour tous les comptages, ajouter le statut de l'assignment
                # On utilise un format standardisé : assignment_status_counting_X
                for counting_order, assignment in job_assignments.items():
                    enriched_result[f'assignment_status_counting_{counting_order}'] = assignment.status
                
                # Si les colonnes 'statut_1er_comptage' et 'statut_2er_comptage' existent déjà,
                # on les garde aussi pour compatibilité, mais on ajoute aussi les colonnes standardisées
                # (elles seront peut-être déjà présentes, donc on ne les écrase pas)
            
            enriched_results.append(enriched_result)
        
        return enriched_results
    
    def _generate_excel(self, results, inventory_id: int, warehouse_id: int):
        """
        Génère un fichier Excel à partir des résultats enrichis.
        
        Args:
            results: Liste des résultats enrichis à exporter
            inventory_id: ID de l'inventaire
            warehouse_id: ID de l'entrepôt
            
        Returns:
            BytesIO: Buffer contenant le fichier Excel
        """
        try:
            import pandas as pd
        except ImportError:
            raise ImportError(
                "pandas est requis pour l'export Excel. "
                "Installez-le avec: pip install pandas"
            )
        
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl est requis pour l'export Excel. "
                "Installez-le avec: pip install openpyxl"
            )
        
        import io
        
        # Vérifier que les résultats ne sont pas vides
        if not results:
            raise ValueError("Aucune donnée à exporter")
        
        # Mettre le code interne dans la clé product si présent
        transformed_results = []
        for result in results:
            transformed_result = dict(result)
            if 'product_internal_code' in transformed_result:
                transformed_result['product'] = transformed_result.pop('product_internal_code')
            transformed_results.append(transformed_result)
        
        # Convertir les résultats en DataFrame
        try:
            df = pd.DataFrame(transformed_results)
        except Exception as e:
            logger.error(f"Erreur lors de la conversion en DataFrame: {e}", exc_info=True)
            raise ValueError(f"Impossible de convertir les résultats en DataFrame: {str(e)}")
        
        # Vérifier que le DataFrame n'est pas vide
        if df.empty:
            raise ValueError("Le DataFrame est vide après conversion")
        
        # Colonnes à exclure de l'export
        excluded_columns = ['location_id', 'job_id', 'ecart_comptage_id']
        
        # Réorganiser les colonnes pour un meilleur affichage
        column_order = []
        
        # Colonnes de base
        base_columns = ['location']
        
        # Ajouter job_reference si présent
        if 'job_reference' in df.columns:
            base_columns.append('job_reference')
        
        # Ajouter job_status si présent
        if 'job_status' in df.columns:
            base_columns.append('job_status')
        
        # Colonnes produit (si présentes)
        product_columns = []
        if 'product' in df.columns:
            product_columns.append('product')
        if 'product_description' in df.columns:
            product_columns.append('product_description')
        
        # Colonnes de comptage (triées par ordre)
        counting_columns = sorted([col for col in df.columns if col.endswith(' comptage')], 
                                 key=lambda x: int(x.split()[0]) if x.split()[0].isdigit() else 999)
        
        # Colonnes de statut d'assignment (triées par ordre)
        # Inclure à la fois les colonnes existantes (statut_1er_comptage, statut_2er_comptage)
        # et les nouvelles colonnes standardisées (assignment_status_counting_X)
        assignment_status_columns = []
        
        # Colonnes existantes du service (statut_1er_comptage, statut_2er_comptage)
        existing_status_columns = sorted(
            [col for col in df.columns if col.startswith('statut_') and col.endswith('_comptage')],
            key=lambda x: int(x.split('_')[1]) if x.split('_')[1].isdigit() else 999
        )
        assignment_status_columns.extend(existing_status_columns)
        
        # Nouvelles colonnes standardisées (assignment_status_counting_X)
        standardized_status_columns = sorted(
            [col for col in df.columns if col.startswith('assignment_status_counting_')],
            key=lambda x: int(x.split('_')[-1]) if x.split('_')[-1].isdigit() else 999
        )
        assignment_status_columns.extend(standardized_status_columns)
        
        # Supprimer les doublons tout en gardant l'ordre
        assignment_status_columns = list(dict.fromkeys(assignment_status_columns))
        
        # Colonnes d'écart (triées)
        ecart_columns = sorted([col for col in df.columns if col.startswith('ecart_')],
                              key=lambda x: tuple(map(int, x.split('_')[1:])) if all(part.isdigit() for part in x.split('_')[1:]) else (999, 999))
        
        # Colonnes finales
        final_columns = []
        if 'final_result' in df.columns:
            final_columns.append('final_result')
        if 'resolved' in df.columns:
            final_columns.append('resolved')
        
        # Construire l'ordre final
        column_order = base_columns + product_columns + counting_columns + assignment_status_columns + ecart_columns + final_columns
        
        # Filtrer les colonnes exclues
        column_order = [col for col in column_order if col not in excluded_columns]
        
        # Réorganiser le DataFrame
        existing_columns = [col for col in column_order if col in df.columns]
        existing_columns = [col for col in existing_columns if col not in excluded_columns]
        
        if not existing_columns:
            existing_columns = [col for col in df.columns if col not in excluded_columns]
        
        df = df[existing_columns]
        
        # Mapping des noms de colonnes
        column_name_mapping = {
            'location': 'Emplacement',
            'job_reference': 'Référence Job',
            'job_status': 'Statut Job',
            'product': 'Code Interne Article',
            'product_description': 'Description Article',
            'final_result': 'Résultat Final',
            'resolved': 'Résolu',
        }
        
        # Ajouter le mapping pour les colonnes de comptage dynamiquement
        for col in df.columns:
            if col.endswith(' comptage'):
                order_num = col.split()[0]
                column_name_mapping[col] = f'{order_num.capitalize()} Comptage'
            elif col.startswith('statut_') and col.endswith('_comptage'):
                # Colonnes existantes : statut_1er_comptage, statut_2er_comptage
                order_num = col.split('_')[1]
                column_name_mapping[col] = f'Statut Assignment {order_num.capitalize()} Comptage'
            elif col.startswith('assignment_status_counting_'):
                # Nouvelles colonnes standardisées : assignment_status_counting_X
                order_num = col.split('_')[-1]
                column_name_mapping[col] = f'Statut Assignment Comptage {order_num}'
            elif col.startswith('ecart_'):
                parts = col.split('_')
                if len(parts) >= 3 and parts[1].isdigit() and parts[2].isdigit():
                    column_name_mapping[col] = f'Écart Comptage {parts[1]}-{parts[2]}'
        
        # Renommer les colonnes
        df = df.rename(columns=column_name_mapping)
        
        # Créer le fichier Excel en mémoire
        excel_buffer = io.BytesIO()
        try:
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Résultats')
                
                # Récupérer la feuille pour appliquer des formats
                worksheet = writer.sheets['Résultats']
                
                # Ajuster la largeur des colonnes
                from openpyxl.utils import get_column_letter
                for idx, col in enumerate(df.columns, start=1):
                    try:
                        col_data = df[col].fillna('').astype(str)
                        if len(col_data) > 0:
                            max_data_length = col_data.map(len).max()
                        else:
                            max_data_length = 0
                        max_length = max(max_data_length, len(str(col)))
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
                    except Exception as e:
                        logger.warning(f"Impossible d'ajuster la largeur de la colonne {col}: {e}")
                        worksheet.column_dimensions[get_column_letter(idx)].width = 15
        except Exception as e:
            logger.error(f"Erreur lors de la génération du fichier Excel: {e}", exc_info=True)
            raise ValueError(f"Impossible de générer le fichier Excel: {str(e)}")
        
        excel_buffer.seek(0)
        return excel_buffer

