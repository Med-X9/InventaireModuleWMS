"""
Commande Django pour générer les PDFs de tous les assignments (peu importe le statut du job).

Cette commande :
- Récupère tous les assignments (sans filtre sur le statut)
- Génère un PDF pour chaque assignment en utilisant la même configuration que JobAssignmentPdfView
- Sauvegarde les PDFs dans un dossier spécifié

Exemple d'utilisation:
    python manage.py generate_pdfs_for_all_assignments --output-dir ./pdfs
    python manage.py generate_pdfs_for_all_assignments --output-dir ./pdfs --job-id 123
    python manage.py generate_pdfs_for_all_assignments --output-dir ./pdfs --inventory-id 456
    python manage.py generate_pdfs_for_all_assignments --output-dir ./pdfs --job-refs JOB-0001 JOB-0002
    python manage.py generate_pdfs_for_all_assignments --output-dir ./pdfs --file jobs_list.txt
    python manage.py generate_pdfs_for_all_assignments --output-dir ./pdfs --warehouse-id 1
"""
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.conf import settings
import os
import logging
from pathlib import Path
from typing import Optional, List

from apps.inventory.models import Assigment, Job, Inventory
from apps.inventory.usecases.job_assignment_pdf import JobAssignmentPdfUseCase

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Génère les PDFs de tous les assignments (peu importe le statut du job)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output-dir',
            type=str,
            required=True,
            help='Dossier de sortie pour sauvegarder les PDFs',
        )
        parser.add_argument(
            '--job-id',
            type=int,
            help='ID du job spécifique (optionnel)',
        )
        parser.add_argument(
            '--inventory-id',
            type=int,
            help='ID de l\'inventaire spécifique (optionnel)',
        )
        parser.add_argument(
            '--warehouse-id',
            type=int,
            help='ID de l\'entrepôt spécifique (optionnel)',
        )
        parser.add_argument(
            '--assignment-id',
            type=int,
            help='ID de l\'assignment spécifique (optionnel)',
        )
        parser.add_argument(
            '--job-refs',
            nargs='+',
            type=str,
            help='Liste des références de jobs (ex: JOB-0001 JOB-0002)',
        )
        parser.add_argument(
            '--file',
            type=str,
            help='Chemin vers un fichier Excel (.xlsx) ou texte (.txt) contenant les références de jobs',
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default=0,
            help='Nom ou index de la feuille Excel à lire (défaut: première feuille)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Mode test : affiche ce qui sera fait sans générer les PDFs',
        )

    def handle(self, *args, **options):
        output_dir = options.get('output_dir')
        job_id = options.get('job_id')
        inventory_id = options.get('inventory_id')
        warehouse_id = options.get('warehouse_id')
        assignment_id = options.get('assignment_id')
        job_refs = options.get('job_refs', [])
        file_path = options.get('file')
        sheet = options.get('sheet', 0)
        dry_run = options.get('dry_run', False)
        
        # Récupérer les références depuis le fichier si fourni
        if file_path:
            file_refs = []
            try:
                if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                    # Fichier Excel
                    file_refs = self.read_excel_file(file_path, sheet)
                    if file_refs is None or not isinstance(file_refs, list):
                        file_refs = []
                else:
                    # Fichier texte
                    with open(file_path, 'r', encoding='utf-8') as f:
                        file_refs = [line.strip() for line in f if line.strip()]
                        if file_refs is None or not isinstance(file_refs, list):
                            file_refs = []
            except FileNotFoundError:
                raise CommandError(f"Fichier non trouvé: {file_path}")
            except CommandError:
                raise
            except Exception as e:
                file_refs = []
                raise CommandError(f"Erreur lors de la lecture du fichier: {str(e)}")
            finally:
                if file_refs is None or not isinstance(file_refs, list):
                    file_refs = []
                if file_refs:
                    job_refs.extend(file_refs)
        
        # Supprimer les doublons
        if job_refs:
            job_refs = list(set(job_refs))
        
        # Afficher les jobs qui seront traités
        if job_refs:
            self.stdout.write(f'  📝 Jobs à traiter: {", ".join(sorted(job_refs))}')
        
        # Créer le dossier de sortie s'il n'existe pas
        if not dry_run:
            try:
                Path(output_dir).mkdir(parents=True, exist_ok=True)
                self.stdout.write(self.style.SUCCESS(f'📁 Dossier de sortie: {output_dir}'))
            except Exception as e:
                raise CommandError(f"Impossible de créer le dossier de sortie: {str(e)}")
        
        # Récupérer tous les assignments (sans filtre sur le statut)
        assignments = self._get_all_assignments(
            job_id=None if job_refs else job_id,  # Ignorer job_id si on a des job_refs
            inventory_id=None if job_refs else inventory_id,  # Ignorer inventory_id si on a des job_refs
            warehouse_id=None if job_refs else warehouse_id,  # Ignorer warehouse_id si on a des job_refs
            assignment_id=assignment_id,
            job_refs=job_refs if job_refs else None
        )
        
        if not assignments.exists():
            self.stdout.write(self.style.WARNING('⚠️  Aucun assignment trouvé'))
            return
        
        self.stdout.write(self.style.SUCCESS(f'📋 {assignments.count()} assignment(s) trouvé(s)'))
        if dry_run:
            self.stdout.write(self.style.WARNING('⚠️  MODE DRY-RUN : Aucun PDF ne sera généré'))
        
        # Statistiques
        stats = {
            'total': 0,
            'success': 0,
            'errors': []
        }
        
        # Générer les PDFs
        use_case = JobAssignmentPdfUseCase()
        
        for assignment in assignments:
            stats['total'] += 1
            try:
                result = self._generate_pdf_for_assignment(
                    assignment, 
                    use_case, 
                    output_dir, 
                    dry_run
                )
                if result['success']:
                    stats['success'] += 1
                    job_status = assignment.job.status if assignment.job else "N/A"
                    assignment_status = assignment.status
                    self.stdout.write(
                        self.style.SUCCESS(
                            f'  ✅ PDF généré: {result["filename"]} '
                            f'(Job: {assignment.job.reference} [{job_status}], '
                            f'Assignment: {assignment.reference} [{assignment_status}])'
                        )
                    )
                else:
                    error_msg = f"Erreur pour assignment {assignment.reference}: {result.get('message', 'Erreur inconnue')}"
                    stats['errors'].append(error_msg)
                    self.stdout.write(self.style.ERROR(f'  ❌ {error_msg}'))
            except Exception as e:
                error_msg = f"Erreur lors de la génération du PDF pour assignment {assignment.reference}: {str(e)}"
                stats['errors'].append(error_msg)
                self.stdout.write(self.style.ERROR(f'  ❌ {error_msg}'))
                logger.error(error_msg, exc_info=True)
        
        # Afficher le résumé
        self.stdout.write('\n' + '='*60)
        self.stdout.write(self.style.SUCCESS('📊 RÉSUMÉ'))
        self.stdout.write('='*60)
        self.stdout.write(f"  Total d'assignments traités: {stats['total']}")
        self.stdout.write(f"  PDFs générés avec succès: {stats['success']}")
        self.stdout.write(f"  Erreurs: {len(stats['errors'])}")
        
        if stats['errors']:
            self.stdout.write(self.style.ERROR(f"\n  ❌ Détails des erreurs:"))
            for error in stats['errors']:
                self.stdout.write(self.style.ERROR(f"    - {error}"))
        
        if dry_run:
            self.stdout.write(self.style.WARNING('\n⚠️  MODE DRY-RUN : Aucun PDF n\'a été généré'))
        else:
            self.stdout.write(self.style.SUCCESS(f'\n✅ Traitement terminé! PDFs sauvegardés dans: {output_dir}'))
    
    def _get_all_assignments(
        self, 
        job_id: Optional[int] = None,
        inventory_id: Optional[int] = None,
        warehouse_id: Optional[int] = None,
        assignment_id: Optional[int] = None,
        job_refs: Optional[List[str]] = None
    ):
        """
        Récupère tous les assignments selon les filtres fournis (sans filtre sur le statut).
        
        Args:
            job_id: ID du job spécifique (optionnel)
            inventory_id: ID de l'inventaire spécifique (optionnel)
            warehouse_id: ID de l'entrepôt spécifique (optionnel)
            assignment_id: ID de l'assignment spécifique (optionnel)
            job_refs: Liste des références de jobs (optionnel)
            
        Returns:
            QuerySet de tous les assignments
        """
        # Récupérer TOUS les assignments (sans filtre sur le statut)
        queryset = Assigment.objects.all().select_related(
            'job',
            'job__inventory',
            'job__warehouse',
            'counting',
            'personne',
            'personne_two'
        ).order_by('job__reference', 'counting__order')
        
        # Appliquer les filtres
        if assignment_id:
            queryset = queryset.filter(id=assignment_id)
        
        if job_id:
            queryset = queryset.filter(job_id=job_id)
        
        if inventory_id:
            queryset = queryset.filter(job__inventory_id=inventory_id)
        
        if warehouse_id:
            queryset = queryset.filter(job__warehouse_id=warehouse_id)
        
        if job_refs:
            queryset = queryset.filter(job__reference__in=job_refs)
        
        return queryset
    
    def _generate_pdf_for_assignment(
        self,
        assignment: Assigment,
        use_case: JobAssignmentPdfUseCase,
        output_dir: str,
        dry_run: bool = False
    ) -> dict:
        """
        Génère le PDF pour un assignment spécifique.
        
        Args:
            assignment: L'assignment pour lequel générer le PDF
            use_case: Le use case pour générer le PDF
            output_dir: Le dossier de sortie
            dry_run: Si True, ne génère pas réellement le PDF
            
        Returns:
            Dictionnaire avec le résultat de la génération
        """
        try:
            job = assignment.job
            if not job:
                return {
                    'success': False,
                    'message': f'Job non trouvé pour l\'assignment {assignment.id}'
                }
            
            # Générer le PDF en utilisant le même use case que la vue
            result = use_case.execute(
                job_id=job.id,
                assignment_id=assignment.id,
                equipe_id=None  # Pas de filtre par équipe pour cette commande
            )
            
            if not result.get('success', False):
                return {
                    'success': False,
                    'message': result.get('message', 'Erreur lors de la génération du PDF')
                }
            
            # Récupérer le buffer PDF
            pdf_buffer = result.get('pdf_buffer')
            if not pdf_buffer:
                return {
                    'success': False,
                    'message': 'Buffer PDF vide'
                }
            
            # Générer le nom de fichier (même format que la vue)
            job_reference = job.reference
            assignment_reference = assignment.reference
            filename = f"FICHE DE COMPTAGE : {job_reference} - {assignment_reference}.pdf"
            
            # Chemin complet du fichier
            file_path = os.path.join(output_dir, filename)
            
            # Sauvegarder le PDF
            if not dry_run:
                with open(file_path, 'wb') as f:
                    f.write(pdf_buffer.getvalue())
            
            return {
                'success': True,
                'filename': filename,
                'file_path': file_path,
                'job_id': job.id,
                'assignment_id': assignment.id
            }
            
        except ValueError as e:
            logger.error(f"Erreur de validation: {str(e)}")
            return {
                'success': False,
                'message': f'Erreur de validation: {str(e)}'
            }
        except Exception as e:
            logger.error(f"Erreur lors de la génération du PDF: {str(e)}", exc_info=True)
            return {
                'success': False,
                'message': f'Erreur: {str(e)}'
            }
    
    def read_excel_file(self, file_path: str, sheet_name_or_index=0) -> List[str]:
        """
        Lit un fichier Excel et extrait les références de jobs depuis la première colonne.
        
        Args:
            file_path: Chemin vers le fichier Excel
            sheet_name_or_index: Nom ou index de la feuille (défaut: première feuille)
            
        Returns:
            Liste des références de jobs
        """
        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                "openpyxl n'est pas installé. Installez-le avec: pip install openpyxl"
            )
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Obtenir la feuille
            try:
                if isinstance(sheet_name_or_index, int):
                    if sheet_name_or_index >= len(wb.worksheets):
                        raise CommandError(
                            f"Index de feuille {sheet_name_or_index} invalide. "
                            f"Le fichier contient {len(wb.worksheets)} feuille(s)."
                        )
                    sheet = wb.worksheets[sheet_name_or_index]
                else:
                    if sheet_name_or_index not in wb.sheetnames:
                        raise CommandError(
                            f"Feuille '{sheet_name_or_index}' non trouvée. "
                            f"Feuilles disponibles: {', '.join(wb.sheetnames)}"
                        )
                    sheet = wb[sheet_name_or_index]
            except (IndexError, KeyError) as e:
                raise CommandError(f"Erreur lors de l'accès à la feuille: {str(e)}")
            
            # Vérifier que la feuille a au moins une ligne
            if sheet.max_row < 1:
                raise CommandError("La feuille Excel est vide (aucune ligne trouvée)")
            
            # Lire les données depuis la première colonne (colonne A, index 0)
            # On lit toutes les lignes, en ignorant les cellules vides
            job_refs = []
            for row in sheet.iter_rows(min_row=1, values_only=False):
                if len(row) > 0:
                    cell_value = row[0].value  # Première colonne (index 0)
                    if cell_value is not None:
                        job_ref = str(cell_value).strip()
                        if job_ref:
                            job_refs.append(job_ref)
            
            # Toujours retourner une liste, même si vide
            return job_refs if job_refs else []
            
        except CommandError:
            # Re-raise les CommandError telles quelles
            raise
        except Exception as e:
            # Logger l'erreur pour le débogage
            import traceback
            logger.error(f"Erreur lors de la lecture du fichier Excel {file_path}: {str(e)}\n{traceback.format_exc()}")
            raise CommandError(f"Erreur lors de la lecture du fichier Excel: {str(e)}")

