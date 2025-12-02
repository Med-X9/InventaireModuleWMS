"""
Vues pour la gestion des inventaires.
"""
import logging
from rest_framework import status, filters
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from ..models import Inventory
from ..services.inventory_result_service import InventoryResultService
from ..services.inventory_service import InventoryService
from ..serializers.inventory_serializer import (
    InventoryCreateSerializer,
    InventoryDuplicateSerializer,
    InventoryDetailSerializer,
    InventoryGetByIdSerializer,
    InventoryTeamSerializer,
    InventoryWarehouseStatsSerializer,
    InventoryUpdateSerializer,
    InventoryDetailModeFieldsSerializer,
    InventoryDetailWithWarehouseSerializer,
)
from ..serializers import InventoryWarehouseResultSerializer, InventoryWarehouseResultEntrySerializer
from ..exceptions import InventoryValidationError, InventoryNotFoundError, StockValidationError
from ..filters import InventoryFilter
from ..repositories import InventoryRepository
from ..interfaces import IInventoryRepository
from ..utils.response_utils import success_response, error_response, validation_error_response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from apps.inventory.usecases.inventory_launch_validation import InventoryLaunchValidationUseCase
from apps.core.datatables.mixins import QueryModelMixin, ServerSideDataTableView

# Configuration du logger
logger = logging.getLogger(__name__)


class InventoryListView(ServerSideDataTableView):
    """
    Vue ultra-simplifiée - Configuration minimale absolue
    
    Supporte automatiquement :
    - Format QueryModel (POST JSON ou GET query params)
    - Format DataTable standard (GET query params)
    - Format REST API (GET query params)
    
    Le package détecte automatiquement :
    - Champs de recherche depuis le serializer
    - Champs de tri depuis le serializer
    - Champs de date depuis le modèle (DateTimeField, DateField)
    - Champs de statut depuis le modèle (CharField avec choices)
    - Optimisations de requêtes
    
    ENDPOINTS:
    - GET  /web/api/inventory/ (tous formats supportés)
    - POST /web/api/inventory/ (format QueryModel uniquement)
    
    DOCUMENTATION: Voir INVENTORY_QUERYMODEL_API.md
    """
    
    # Configuration minimale
    model = Inventory
    serializer_class = InventoryDetailSerializer
    
    # Champs de recherche - tous les champs disponibles dans le JSON
    search_fields = [
        'reference', 'label', 'status', 'inventory_type', 'date',
        'en_preparation_status_date', 'en_realisation_status_date',
        'termine_status_date', 'cloture_status_date', 'created_at',
        # Relations pour account_name, account_reference, warehouse_name
        'awi_links__account__account_name',
        'awi_links__account__reference',
        'awi_links__warehouse__warehouse_name',
        'awi_links__warehouse__reference'
    ]
    
    # Champs de tri - tous les champs disponibles
    order_fields = [
        'id', 'reference', 'label', 'date', 'status', 'inventory_type',
        'en_preparation_status_date', 'en_realisation_status_date',
        'termine_status_date', 'cloture_status_date', 'created_at',
        'awi_links__account__account_name',
        'awi_links__warehouse__warehouse_name'
    ]
    
    # Mapping frontend -> backend pour le filtrage par colonnes (utilisé par QueryModel)
    filter_aliases = {
        'reference': 'reference',
        'label': 'label',
        'date': 'date',
        'status': 'status',
        'inventory_type': 'inventory_type',
        'en_preparation_status_date': 'en_preparation_status_date',
        'en_realisation_status_date': 'en_realisation_status_date',
        'termine_status_date': 'termine_status_date',
        'cloture_status_date': 'cloture_status_date',
        'created_at': 'created_at',
        'account_name': 'awi_links__account__account_name',
        'account_reference': 'awi_links__account__reference',
        'warehouse_name': 'awi_links__warehouse__warehouse_name',
        'warehouse_references': 'awi_links__warehouse__reference'
    }
    
    # Mapping explicite pour QueryModel (colId -> field_name)
    # Utilisé par _get_column_field_mapping_for_querymodel() pour le tri et filtrage
    column_field_mapping = {
        'id': 'id',
        'reference': 'reference',
        'label': 'label',
        'date': 'date',
        'status': 'status',
        'inventory_type': 'inventory_type',
        'en_preparation_status_date': 'en_preparation_status_date',
        'en_realisation_status_date': 'en_realisation_status_date',
        'termine_status_date': 'termine_status_date',
        'cloture_status_date': 'cloture_status_date',
        'created_at': 'created_at',
        'updated_at': 'updated_at',
        'account_name': 'awi_links__account__account_name',
        'account_reference': 'awi_links__account__reference',
        'warehouse_name': 'awi_links__warehouse__warehouse_name',
        'warehouse_references': 'awi_links__warehouse__reference'
    }
    
    # Optionnel : seulement si vous voulez personnaliser
    default_order = '-created_at'
    page_size = 20
    export_filename = 'inventaires'

    def get_datatable_queryset(self):
        """Optimisation automatique des requêtes"""
        queryset = super().get_datatable_queryset()
        # awi_links est une relation inverse (reverse relation), utiliser prefetch_related
        return queryset.prefetch_related(
            'awi_links__account',
            'awi_links__warehouse',
            'countings'
        )



# Exemples d'utilisation de ServerSideDataTableView pour différents cas d'usage

class SimpleInventoryListView(ServerSideDataTableView):
    """
    Exemple simple - configuration minimale
    Supporte automatiquement : tri, recherche, pagination
    """
    model = Inventory
    serializer_class = InventoryDetailSerializer
    search_fields = ['label', 'reference']
    order_fields = ['id', 'label', 'created_at']
    export_filename = 'inventaires_simples'

class AdvancedInventoryListView(ServerSideDataTableView):
    """
    Exemple avancé - avec tous les types de filtres
    Supporte : tri, recherche, filtres django-filter, filtres de date, filtres de statut
    """
    model = Inventory
    serializer_class = InventoryDetailSerializer
    filterset_class = InventoryFilter
    search_fields = ['label', 'reference', 'status', 'inventory_type']
    order_fields = ['id', 'reference', 'label', 'date', 'status', 'inventory_type', 'created_at']
    default_order = '-created_at'
    page_size = 50
    filter_fields = ['status', 'inventory_type']
    date_fields = ['date', 'created_at']
    status_fields = ['status']
    export_filename = 'inventaires_avances'

class CustomInventoryListView(ServerSideDataTableView):
    """
    Exemple avec personnalisation avancée
    Surcharge des méthodes pour comportement personnalisé
    """
    model = Inventory
    serializer_class = InventoryDetailSerializer
    search_fields = ['label', 'reference', 'status']
    order_fields = ['id', 'label', 'created_at']
    
    def get_datatable_queryset(self):
        """Queryset personnalisé - seulement les inventaires non supprimés"""
        return Inventory.objects.filter(is_deleted=False)
    
    def get_datatable_config(self):
        """Configuration personnalisée"""
        config = super().get_datatable_config()
        # Personnaliser la configuration si nécessaire
        return config
    
    export_filename = 'inventaires_personnalises'

class InventoryOrderingTestView(APIView):
    """Vue de test pour vérifier le tri"""
    
    def get(self, request):
        """Teste différents paramètres de tri"""
        from apps.core.datatables.base import DataTableProcessor
        
        # Créer une instance de InventoryListView pour récupérer la configuration
        inventory_view = InventoryListView()
        queryset = inventory_view.get_datatable_queryset()
        config = inventory_view.get_datatable_config()
        
        # Test avec différents paramètres de tri
        test_params = [
            ('ordering=label', 'Tri par label'),
            ('ordering=-label', 'Tri par label décroissant'),
            ('order[0][column]=2&order[0][dir]=asc', 'Tri DataTable colonne 2'),
            ('order[0][column]=2&order[0][dir]=desc', 'Tri DataTable colonne 2 décroissant'),
        ]
        
        results = []
        for param, description in test_params:
            # Créer une requête de test
            from django.test import RequestFactory
            factory = RequestFactory()
            test_request = factory.get(f'/test/?{param}')
            
            # Appliquer le tri
            processor = DataTableProcessor(
                config=config,
                filter_handler=inventory_view.get_datatable_filter(),
                serializer_handler=inventory_view.get_datatable_serializer()
            )
            
            # Tester le tri
            try:
                response = processor.process(test_request, queryset)
                results.append({
                    'param': param,
                    'description': description,
                    'success': True,
                    'response': response.content.decode()[:200] + '...'
                })
            except Exception as e:
                results.append({
                    'param': param,
                    'description': description,
                    'success': False,
                    'error': str(e)
                })
        
        return Response({
            'test_results': results,
            'order_fields': config.get_order_fields(),
            'default_order': config.get_default_order()
        })

class InventoryCreateView(APIView):
    """
    Vue pour créer un inventaire en utilisant la structure Interface, Service, Serializer, Repository, Use Cases.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        from ..services.inventory_management_service import InventoryManagementService
        from ..serializers.inventory_serializer import InventoryCreateSerializer
        self.service = InventoryManagementService()
        self.serializer_class = InventoryCreateSerializer

    def post(self, request, *args, **kwargs):
        """
        Crée un inventaire en utilisant la structure complète.
        """
        try:
            print(request.data)
            # ÉTAPE 1: Validation des données avec le serializer
            serializer = self.serializer_class(data=request.data)
            if not serializer.is_valid():
                logger.warning(f"Données invalides lors de la création d'un inventaire: {serializer.errors}")
                return validation_error_response(
                    serializer.errors,
                    message="Erreur de validation lors de la création de l'inventaire"
                )
            
            # ÉTAPE 2: Validation métier avec le service
            try:
                self.service.validate_create_data(serializer.validated_data)
            except InventoryValidationError as e:
                logger.warning(f"Erreur de validation métier lors de la création: {str(e)}")
                return error_response(
                    message=str(e),
                    status_code=status.HTTP_400_BAD_REQUEST
                )
            
            # ÉTAPE 3: Création avec le service (qui utilise le use case)
            result = self.service.create_inventory(serializer.validated_data)
            
            # ÉTAPE 4: Retour de la réponse standardisée
            return success_response(
                data=result,
                message="Inventaire créé avec succès",
                status_code=status.HTTP_201_CREATED
            )
            
        except InventoryValidationError as e:
            logger.warning(f"Erreur de validation lors de la création: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Erreur lors de la création d'un inventaire: {str(e)}", exc_info=True)
            return error_response(
                message="Une erreur inattendue s'est produite lors de la création de l'inventaire",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class InventoryDuplicateView(APIView):
    """
    Vue pour dupliquer un inventaire existant en utilisant la configuration de comptages source.
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        from ..services.inventory_duplication_service import InventoryDuplicationService
        self.service = InventoryDuplicationService()
        self.serializer_class = InventoryDuplicateSerializer

    def post(self, request, pk, *args, **kwargs):
        """
        Duplique un inventaire en conservant la configuration des comptages.
        """
        serializer = self.serializer_class(data=request.data)
        if not serializer.is_valid():
            logger.warning(
                "Données invalides lors de la duplication d'un inventaire: %s",
                serializer.errors
            )
            return validation_error_response(
                serializer.errors,
                message="Erreur de validation lors de la duplication de l'inventaire"
            )

        try:
            result = self.service.duplicate_inventory(pk, serializer.validated_data)
            return success_response(
                data=result,
                message="Inventaire dupliqué avec succès",
                status_code=status.HTTP_201_CREATED
            )
        except InventoryNotFoundError as e:
            logger.warning(
                "Inventaire source introuvable pour la duplication (id=%s): %s",
                pk,
                str(e)
            )
            return error_response(
                message=str(e),
                status_code=status.HTTP_404_NOT_FOUND
            )
        except InventoryValidationError as e:
            logger.warning(
                "Erreur métier lors de la duplication d'inventaire (id=%s): %s",
                pk,
                str(e)
            )
            return error_response(
                message=str(e),
                status_code=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(
                "Erreur inattendue lors de la duplication d'inventaire (id=%s): %s",
                pk,
                str(e),
                exc_info=True
            )
            return error_response(
                message="Une erreur inattendue s'est produite lors de la duplication de l'inventaire",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class InventoryDetailView(APIView):
    """
    Vue pour récupérer les détails d'un inventaire avec informations complètes des warehouses.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.repository = InventoryRepository()

    def get(self, request, pk, *args, **kwargs):
        """
        Récupère les détails d'un inventaire avec informations complètes des warehouses.
        """
        try:
            inventory = self.repository.get_with_related_data(pk)
            serializer = InventoryDetailWithWarehouseSerializer(inventory)
            return success_response(
                data=serializer.data,
                message="Détails de l'inventaire récupérés avec succès"
            )
        except InventoryNotFoundError as e:
            logger.warning(f"Inventaire non trouvé: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des détails d'un inventaire: {str(e)}", exc_info=True)
            return error_response(
                message="Une erreur inattendue s'est produite lors de la récupération des détails de l'inventaire",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class InventoryDetailByReferenceView(APIView):
    """
    Vue pour récupérer les détails d'un inventaire par sa référence avec informations complètes des warehouses.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.service = InventoryService()

    def get(self, request, reference, *args, **kwargs):
        """
        Récupère les détails d'un inventaire par sa référence avec informations complètes des warehouses.
        """
        try:
            inventory = self.service.get_inventory_with_related_data_by_reference(reference)
            serializer = InventoryDetailWithWarehouseSerializer(inventory)
            return success_response(
                data=serializer.data,
                message="Détails de l'inventaire récupérés avec succès"
            )
        except InventoryNotFoundError as e:
            logger.warning(f"Inventaire non trouvé: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des détails d'un inventaire: {str(e)}", exc_info=True)
            return error_response(
                message="Une erreur inattendue s'est produite lors de la récupération des détails de l'inventaire",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class InventoryUpdateView(APIView):
    """
    Vue pour mettre à jour un inventaire en utilisant la structure Interface, Service, Serializer, Repository, Use Cases.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        from ..services.inventory_management_service import InventoryManagementService
        self.service = InventoryManagementService()
        self.serializer_class = InventoryUpdateSerializer

    def put(self, request, pk, *args, **kwargs):
        """
        Met à jour un inventaire en utilisant la structure complète.
        """
        try:
            # ÉTAPE 1: Validation des données avec le serializer
            serializer = self.serializer_class(data=request.data)
            if not serializer.is_valid():
                logger.warning(f"Données invalides lors de la mise à jour d'un inventaire: {serializer.errors}")
                return validation_error_response(
                    serializer.errors,
                    message="Erreur de validation lors de la mise à jour de l'inventaire"
                )
            
            # ÉTAPE 2: Validation métier avec le service
            try:
                self.service.validate_update_data(serializer.validated_data)
            except InventoryValidationError as e:
                logger.warning(f"Erreur de validation métier lors de la mise à jour: {str(e)}")
                return error_response(
                    message=str(e),
                    status_code=status.HTTP_400_BAD_REQUEST
                )
            
            # ÉTAPE 3: Mise à jour avec le service (qui utilise le use case)
            result = self.service.update_inventory(pk, serializer.validated_data)
            
            # ÉTAPE 4: Retour de la réponse standardisée
            return success_response(
                data=result,
                message="Inventaire mis à jour avec succès"
            )
            
        except InventoryNotFoundError as e:
            logger.warning(f"Inventaire non trouvé lors de la mise à jour: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_404_NOT_FOUND
            )
        except InventoryValidationError as e:
            logger.warning(f"Erreur de validation lors de la mise à jour: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Erreur lors de la mise à jour d'un inventaire: {str(e)}", exc_info=True)
            return error_response(
                message="Une erreur inattendue s'est produite lors de la mise à jour de l'inventaire",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class InventoryDeleteView(APIView):
    """
    Vue pour effectuer un soft delete d'un inventaire.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.repository = InventoryRepository()

    def delete(self, request, pk, *args, **kwargs):
        """
        Effectue un soft delete d'un inventaire si son statut est en attente.
        """
        try:
            service = InventoryService()
            service.delete_inventory(pk)
            return success_response(
                message="L'inventaire a été supprimé avec succès"
            )
        except InventoryNotFoundError as e:
            logger.warning(f"Inventaire non trouvé lors de la suppression: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_404_NOT_FOUND
            )
        except InventoryValidationError as e:
            logger.warning(f"Erreur de validation lors de la suppression: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Erreur lors de la suppression d'un inventaire: {str(e)}", exc_info=True)
            return error_response(
                message="Une erreur inattendue s'est produite lors de la suppression de l'inventaire",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class InventoryLaunchView(APIView):
    """
    Vue pour lancer un inventaire.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.service = InventoryService()

    def post(self, request, pk, *args, **kwargs):
        """
        Lance un inventaire en vérifiant le mode de comptage et en remplissant les détails si nécessaire.
        """
        try:
            # Validation métier avant lancement via le service
            validation_result = self.service.validate_launch_inventory(pk)
            self.service.launch_inventory(pk)
            
            # Préparer la réponse avec les informations de validation
            extra_data = {}
            if validation_result and 'infos' in validation_result:
                extra_data['infos'] = validation_result['infos']
            
            return success_response(
                message="L'inventaire a été lancé avec succès",
                **extra_data
            )
        except InventoryNotFoundError as e:
            logger.warning(f"Inventaire non trouvé lors du lancement: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_404_NOT_FOUND
            )
        except InventoryValidationError as e:
            logger.warning(f"Erreur de validation lors du lancement: {str(e)}")
            # Séparer les erreurs multiples si elles sont séparées par " | "
            error_message = str(e)
            if " | " in error_message:
                errors = error_message.split(" | ")
                return error_response(
                    message="Plusieurs erreurs de validation ont été détectées",
                    errors=errors,
                    status_code=status.HTTP_400_BAD_REQUEST
                )
            else:
                return error_response(
                    message=error_message,
                    status_code=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            logger.error(f"Erreur lors du lancement de l'inventaire: {str(e)}", exc_info=True)
            return error_response(
                message="Une erreur inattendue s'est produite lors du lancement de l'inventaire",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class InventoryCancelView(APIView):
    """
    Vue pour annuler le lancement d'un inventaire.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.service = InventoryService()

    def post(self, request, pk, *args, **kwargs):
        """
        Annule le lancement d'un inventaire.
        """
        try:
            self.service.cancel_inventory(pk)
            return success_response(
                message="L'inventaire a été annulé avec succès"
            )
        except InventoryNotFoundError as e:
            logger.warning(f"Inventaire non trouvé lors de l'annulation: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_404_NOT_FOUND
            )
        except InventoryValidationError as e:
            logger.warning(f"Erreur de validation lors de l'annulation: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Erreur lors de l'annulation de l'inventaire: {str(e)}", exc_info=True)
            return error_response(
                message="Une erreur inattendue s'est produite lors de l'annulation de l'inventaire",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class InventoryTeamView(APIView):
    """
    Vue pour récupérer l'équipe d'un inventaire.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.repository = InventoryRepository()

    def get(self, request, pk, *args, **kwargs):
        """
        Récupère l'équipe d'un inventaire.
        """
        try:
            inventory = self.repository.get_with_related_data(pk)
            serializer = InventoryDetailModeFieldsSerializer(inventory)
            return success_response(
                data=serializer.data,
                message="Détails de l'inventaire récupérés avec succès"
            )
        except InventoryNotFoundError as e:
            logger.warning(f"Inventaire non trouvé: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des détails de l'inventaire: {str(e)}", exc_info=True)
            return error_response(
                message="Une erreur inattendue s'est produite lors de la récupération des détails de l'inventaire",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class InventoryWarehouseStatsView(APIView):
    """
    Vue pour récupérer les statistiques des warehouses d'un inventaire.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, inventory_id):
        """
        Récupère les statistiques des warehouses pour un inventaire.
        
        Args:
            request: La requête HTTP
            inventory_id: L'ID de l'inventaire
            
        Returns:
            Response: La réponse HTTP avec les statistiques des warehouses
        """
        try:
            # Récupérer les statistiques via le service
            inventory_service = InventoryService()
            stats_data = inventory_service.get_warehouse_stats_for_inventory(inventory_id)
            
            # Sérialiser les données
            serializer = InventoryWarehouseStatsSerializer(stats_data, many=True)
            
            return success_response(
                data=serializer.data,
                message='Statistiques des warehouses récupérées avec succès',
                warehouses_count=len(stats_data)
            )
            
        except InventoryNotFoundError as e:
            return error_response(
                message='Inventaire non trouvé',
                status_code=status.HTTP_404_NOT_FOUND
            )
            
        except InventoryValidationError as e:
            return error_response(
                message='Erreur de validation',
                errors=[str(e)],
                status_code=status.HTTP_400_BAD_REQUEST
            )
            
        except Exception as e:
            logger.error(f"Erreur lors de la récupération des statistiques des warehouses: {str(e)}")
            return error_response(
                message='Erreur lors de la récupération des statistiques des warehouses',
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class InventoryResultByWarehouseView(ServerSideDataTableView):
    """
    Vue permettant de récupérer les résultats d'un inventaire agrégés par entrepôt.
    Utilise QueryModel (format AG-Grid) pour gérer automatiquement :
    - Tri multi-colonnes (sortModel)
    - Filtres complexes (filterModel)
    - Pagination (startRow/endRow)
    - Support des listes de dictionnaires via get_data_source()
    """
    
    serializer_class = InventoryWarehouseResultSerializer
    
    # Configuration de pagination
    default_page_size = 20
    max_page_size = 1000
    export_filename = 'resultats_inventaire_par_warehouse'
    
    # Champs pour la recherche globale
    search_fields = [
        'location', 'location_id', 'product', 'product_description', 'product_internal_code',
        'job_id', 'job_reference', 'final_result', 'resolved',
        # Comptages dynamiques
        '1er comptage', '2er comptage', '3er comptage', '4er comptage', '5er comptage',
        # Écarts dynamiques
        'ecart_1_2', 'ecart_2_3', 'ecart_3_4', 'ecart_4_5', 'ecart_1_3', 'ecart_2_4', 'ecart_3_5',
        # Statuts d'assignment par comptage
        'statut_1er_comptage', 'statut_2er_comptage'
    ]
    
    # Mapping des colonnes frontend -> backend (utilisé par QueryModel)
    column_field_mapping = {
        'id': 'location_id',
        'location_id': 'location_id',
        'article': 'product',
        'product': 'product',
        'product_description': 'product_description',
        'product_internal_code': 'product_internal_code',
        'emplacement': 'location',
        'location': 'location',
        'job_id': 'job_id',
        'job_reference': 'job_reference',
        'job': 'job_reference',  # Alias
        'contage_1': '1er comptage',
        'contage_2': '2er comptage',
        'contage_3': '3er comptage',
        'contage_4': '4er comptage',
        'contage_5': '5er comptage',
        'ecart_1_2': 'ecart_1_2',
        'ecart_2_3': 'ecart_2_3',
        'ecart_3_4': 'ecart_3_4',
        'ecart_4_5': 'ecart_4_5',
        'ecart_1_3': 'ecart_1_3',
        'ecart_2_4': 'ecart_2_4',
        'ecart_3_5': 'ecart_3_5',
        'resultats': 'final_result',
        'final_result': 'final_result',
        'resolved': 'resolved',
        # Statuts d'assignment par comptage (seulement 1er et 2ème)
        'statut_1er_comptage': 'statut_1er_comptage',
        'statut_2er_comptage': 'statut_2er_comptage',
        # Alias pour compatibilité
        'statut_comptage_1': 'statut_1er_comptage',
        'statut_comptage_2': 'statut_2er_comptage',
    }
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.service = InventoryResultService()
    
    def get_data_source(self):
        """Retourne la source de données depuis le service."""
        from apps.core.datatables.datasource import DataSourceFactory
        
        inventory_id = self.kwargs.get('inventory_id')
        warehouse_id = self.kwargs.get('warehouse_id')
        
        if not inventory_id or not warehouse_id:
            raise InventoryValidationError("inventory_id et warehouse_id sont requis")
        
        # Récupérer les données depuis le service
        data_list = self.service.get_inventory_results_for_warehouse(inventory_id, warehouse_id)
        
        return DataSourceFactory.create(data_list)
    
    def serialize_data(self, data):
        """
        Sérialise les données avec InventoryWarehouseResultEntrySerializer.
        
        Surcharge pour utiliser le serializer d'entrée au lieu du serializer racine
        qui attend un format avec 'success' et 'data'.
        """
        if isinstance(data, list):
            # Liste de dicts -> sérialiser chaque élément avec EntrySerializer
            serializer = InventoryWarehouseResultEntrySerializer(data, many=True)
            return serializer.data
        else:
            # Convertir en liste si nécessaire
            data_list = list(data) if hasattr(data, '__iter__') else [data]
            serializer = InventoryWarehouseResultEntrySerializer(data_list, many=True)
            return serializer.data


class InventoryResultExportExcelView(APIView):
    """
    Vue pour exporter les résultats d'inventaire en Excel.
    """
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.service = InventoryResultService()
    
    def get(self, request, inventory_id: int, warehouse_id: int, *args, **kwargs):
        """
        Exporte les résultats d'un inventaire pour un entrepôt en fichier Excel.
        
        Args:
            inventory_id: ID de l'inventaire
            warehouse_id: ID de l'entrepôt
            
        Returns:
            Fichier Excel avec les résultats
        """
        try:
            # Récupérer les résultats via le service
            results = self.service.get_inventory_results_for_warehouse(
                inventory_id=inventory_id,
                warehouse_id=warehouse_id,
            )
            
            if not results:
                return error_response(
                    message="Aucun résultat trouvé pour cet inventaire et cet entrepôt",
                    status_code=status.HTTP_404_NOT_FOUND
                )
            
            # Générer le fichier Excel
            excel_buffer = self._generate_excel(results, inventory_id, warehouse_id)
            
            # Récupérer les informations de l'inventaire pour le nom du fichier
            try:
                inventory = Inventory.objects.get(id=inventory_id)
                inventory_ref = inventory.reference.replace(' ', '_')
            except Inventory.DoesNotExist:
                inventory_ref = f"inventaire_{inventory_id}"
            
            # Définir le nom du fichier
            filename = f"resultats_{inventory_ref}_warehouse_{warehouse_id}.xlsx"
            
            # Créer la réponse HTTP avec le fichier Excel
            from django.http import HttpResponse
            response = HttpResponse(
                excel_buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except InventoryNotFoundError as error:
            logger.warning(
                "Inventaire introuvable lors de l'export Excel (id=%s, warehouse_id=%s).",
                inventory_id,
                warehouse_id,
            )
            return error_response(
                message=str(error),
                status_code=status.HTTP_404_NOT_FOUND
            )
        except InventoryValidationError as error:
            logger.warning(
                "Erreur de validation lors de l'export Excel (id=%s, warehouse_id=%s): %s",
                inventory_id,
                warehouse_id,
                error,
            )
            return error_response(
                message=str(error),
                status_code=status.HTTP_400_BAD_REQUEST
            )
        except ValueError as error:
            logger.warning(
                "Erreur de valeur lors de l'export Excel (id=%s, warehouse_id=%s): %s",
                inventory_id,
                warehouse_id,
                error,
            )
            return error_response(
                message=str(error),
                status_code=status.HTTP_400_BAD_REQUEST
            )
        except ImportError as error:
            logger.error(
                "Dépendance manquante pour l'export Excel (id=%s, warehouse_id=%s): %s",
                inventory_id,
                warehouse_id,
                error,
            )
            return error_response(
                message=f"Configuration manquante pour l'export Excel: {str(error)}",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as error:
            logger.error(
                "Erreur inattendue lors de l'export Excel (id=%s, warehouse_id=%s): %s",
                inventory_id,
                warehouse_id,
                error,
                exc_info=True,
            )
            return error_response(
                message=f"Une erreur inattendue est survenue lors de l'export Excel: {str(error)}",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _generate_excel(self, results, inventory_id: int, warehouse_id: int):
        """
        Génère un fichier Excel à partir des résultats.
        
        Les colonnes suivantes sont exclues de l'export :
        - location_id
        - job_id
        - ecart_comptage_id
        
        Args:
            results: Liste des résultats à exporter
            inventory_id: ID de l'inventaire
            warehouse_id: ID de l'entrepôt
            
        Returns:
            BytesIO: Buffer contenant le fichier Excel
            
        Raises:
            ValueError: Si les résultats sont vides ou invalides
            ImportError: Si pandas ou openpyxl ne sont pas installés
        """
        try:
            import pandas as pd
        except ImportError:
            raise ImportError(
                "pandas est requis pour l'export Excel. "
                "Installez-le avec: pip install pandas"
            )
        
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl est requis pour l'export Excel. "
                "Installez-le avec: pip install openpyxl"
            )
        
        import io
        
        # Vérifier que les résultats ne sont pas vides
        if not results:
            raise ValueError("Aucune donnée à exporter")
        
        # Mettre le code interne dans la clé product si présent
        transformed_results = []
        for result in results:
            transformed_result = dict(result)
            if 'product_internal_code' in transformed_result:
                transformed_result['product'] = transformed_result.pop('product_internal_code')
            transformed_results.append(transformed_result)
        
        # Convertir les résultats en DataFrame
        try:
            df = pd.DataFrame(transformed_results)
        except Exception as e:
            logger.error(f"Erreur lors de la conversion en DataFrame: {e}", exc_info=True)
            raise ValueError(f"Impossible de convertir les résultats en DataFrame: {str(e)}")
        
        # Vérifier que le DataFrame n'est pas vide
        if df.empty:
            raise ValueError("Le DataFrame est vide après conversion")
        
        # Vérifier que le DataFrame a des colonnes
        if len(df.columns) == 0:
            raise ValueError("Le DataFrame n'a aucune colonne")
        
        # Colonnes à exclure de l'export
        excluded_columns = ['location_id', 'job_id', 'ecart_comptage_id']
        
        # Réorganiser les colonnes pour un meilleur affichage
        # Ordre souhaité : location, job_reference (si présent), product (si présent), 
        # product_description (si présent), puis tous les comptages, puis final_result, resolved
        column_order = []
        
        # Colonnes de base (sans location_id et job_id)
        base_columns = ['location']  # location_id exclu
        
        # Ajouter job_reference si présent (mais pas job_id)
        if 'job_reference' in df.columns:
            base_columns.append('job_reference')
        
        # Colonnes produit (si présentes)
        product_columns = []
        if 'product' in df.columns:
            product_columns.append('product')
        if 'product_description' in df.columns:
            product_columns.append('product_description')
        
        # Colonnes de comptage (triées par ordre)
        counting_columns = sorted([col for col in df.columns if col.endswith(' comptage')], 
                                 key=lambda x: int(x.split()[0]) if x.split()[0].isdigit() else 999)
        
        # Colonnes d'écart (triées)
        ecart_columns = sorted([col for col in df.columns if col.startswith('ecart_')],
                              key=lambda x: tuple(map(int, x.split('_')[1:])) if all(part.isdigit() for part in x.split('_')[1:]) else (999, 999))
        
        # Colonnes finales (sans ecart_comptage_id)
        final_columns = []
        if 'final_result' in df.columns:
            final_columns.append('final_result')
        # ecart_comptage_id exclu
        if 'resolved' in df.columns:
            final_columns.append('resolved')
        
        # Construire l'ordre final
        column_order = base_columns + product_columns + counting_columns + ecart_columns + final_columns
        
        # Filtrer les colonnes exclues
        column_order = [col for col in column_order if col not in excluded_columns]
        
        # Réorganiser le DataFrame en excluant les colonnes non désirées
        existing_columns = [col for col in column_order if col in df.columns]
        
        # Exclure également toutes les colonnes de la liste excluded_columns qui pourraient être présentes
        existing_columns = [col for col in existing_columns if col not in excluded_columns]
        
        if not existing_columns:
            # Si aucune colonne n'est trouvée, utiliser toutes les colonnes disponibles sauf celles exclues
            existing_columns = [col for col in df.columns if col not in excluded_columns]
        
        df = df[existing_columns]
        
        # Mapping des noms de colonnes techniques vers des noms conviviaux en français
        column_name_mapping = {
            'location': 'Emplacement',
            'job_reference': 'Référence Job',
            'product': 'Code Interne Article',
            'product_description': 'Description Article',
            'final_result': 'Résultat Final',
            'resolved': 'Résolu',
        }
        
        # Ajouter le mapping pour les colonnes de comptage dynamiquement
        for col in df.columns:
            if col.endswith(' comptage'):
                # Extraire le numéro du comptage (ex: "1er comptage" -> "1er Comptage")
                order_num = col.split()[0]
                # Capitaliser la première lettre pour uniformiser
                column_name_mapping[col] = f'{order_num.capitalize()} Comptage'
            elif col.startswith('ecart_'):
                # Extraire les numéros des comptages (ex: "ecart_1_2" -> "Écart Comptage 1-2")
                parts = col.split('_')
                if len(parts) >= 3 and parts[1].isdigit() and parts[2].isdigit():
                    column_name_mapping[col] = f'Écart Comptage {parts[1]}-{parts[2]}'
        
        # Renommer les colonnes avec les noms conviviaux
        df = df.rename(columns=column_name_mapping)
        
        # Créer le fichier Excel en mémoire
        excel_buffer = io.BytesIO()
        try:
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Résultats')
                
                # Récupérer la feuille pour appliquer des formats
                worksheet = writer.sheets['Résultats']
                
                # Ajuster la largeur des colonnes
                from openpyxl.utils import get_column_letter
                for idx, col in enumerate(df.columns, start=1):
                    try:
                        # Remplacer les valeurs None par des chaînes vides pour éviter les erreurs
                        col_data = df[col].fillna('').astype(str)
                        
                        # Calculer la longueur maximale
                        if len(col_data) > 0:
                            max_data_length = col_data.map(len).max()
                        else:
                            max_data_length = 0
                        
                        max_length = max(max_data_length, len(str(col)))
                        # Limiter la largeur maximale à 50
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
                    except Exception as e:
                        # Si l'ajustement de la largeur échoue, utiliser une largeur par défaut
                        logger.warning(f"Impossible d'ajuster la largeur de la colonne {col}: {e}")
                        worksheet.column_dimensions[get_column_letter(idx)].width = 15
        except Exception as e:
            logger.error(f"Erreur lors de la génération du fichier Excel: {e}", exc_info=True)
            raise ValueError(f"Impossible de générer le fichier Excel: {str(e)}")
        
        excel_buffer.seek(0)
        return excel_buffer


class InventoryImportView(APIView):
    """
    Vue pour importer des inventaires via API.
    Permet l'importation en lot d'inventaires avec validation et création.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.service = InventoryService()

    def post(self, request, *args, **kwargs):
        """
        Importe des inventaires en lot.
        
        Format JSON attendu:
        {
            "inventories": [
                {
                    "label": "Inventaire trimestriel Q1 2023",
                    "date": "2024-03-20",
                    "account_id": 2,
                    "warehouse": [
                        {"id": 1, "date": "12/12/2025"},
                        {"id": 2, "date": "12/12/2025"}
                    ],
                    "comptages": [
                        {
                            "order": 1,
                            "count_mode": "en vrac",
                            "unit_scanned": true,
                            "entry_quantity": false,
                            "is_variant": false,
                            "stock_situation": false,
                            "quantity_show": true,
                            "show_product": true,
                            "dlc": false,
                            "n_serie": false,
                            "n_lot": false
                        }
                    ]
                }
            ],
            "options": {
                "validate_only": false,
                "stop_on_error": true,
                "return_details": true
            }
        }
        """
        try:
            # Validation du format de la requête
            if 'inventories' not in request.data:
                return error_response(
                    message='Le champ "inventories" est requis',
                    status_code=status.HTTP_400_BAD_REQUEST
                )
            
            inventories_data = request.data['inventories']
            options = request.data.get('options', {})
            
            if not isinstance(inventories_data, list):
                return error_response(
                    message='Le champ "inventories" doit être une liste',
                    status_code=status.HTTP_400_BAD_REQUEST
                )
            
            # Options d'importation
            validate_only = options.get('validate_only', False)
            stop_on_error = options.get('stop_on_error', True)
            return_details = options.get('return_details', True)
            
            results = []
            errors = []
            success_count = 0
            error_count = 0
            
            for index, inventory_data in enumerate(inventories_data):
                try:
                    # Validation du serializer pour chaque inventaire
                    serializer = InventoryCreateSerializer(data=inventory_data)
                    if not serializer.is_valid():
                        error_msg = f"Inventaire {index + 1}: Erreurs de validation - {serializer.errors}"
                        errors.append({
                            'index': index,
                            'data': inventory_data,
                            'errors': serializer.errors
                        })
                        error_count += 1
                        
                        if stop_on_error:
                            break
                        continue
                    
                    if validate_only:
                        # Validation uniquement
                        result = {
                            'index': index,
                            'status': 'validated',
                            'message': f'Inventaire {index + 1} validé avec succès'
                        }
                        success_count += 1
                    else:
                        # Création de l'inventaire
                        result = self.service.create_inventory(serializer.validated_data)
                        result['index'] = index
                        result['status'] = 'created'
                        success_count += 1
                    
                    if return_details:
                        results.append(result)
                    
                except InventoryValidationError as e:
                    error_msg = f"Inventaire {index + 1}: Erreur de validation - {str(e)}"
                    errors.append({
                        'index': index,
                        'data': inventory_data,
                        'error': str(e)
                    })
                    error_count += 1
                    
                    if stop_on_error:
                        break
                        
                except Exception as e:
                    error_msg = f"Inventaire {index + 1}: Erreur inattendue - {str(e)}"
                    logger.error(f"Erreur lors de l'importation de l'inventaire {index + 1}: {str(e)}", exc_info=True)
                    errors.append({
                        'index': index,
                        'data': inventory_data,
                        'error': str(e)
                    })
                    error_count += 1
                    
                    if stop_on_error:
                        break
            
            # Préparation de la réponse
            summary = {
                'total': len(inventories_data),
                'success_count': success_count,
                'error_count': error_count
            }
            
            # Détermination du code de statut
            if error_count == 0:
                status_code = status.HTTP_201_CREATED
                message = 'Tous les inventaires ont été importés avec succès'
            elif success_count > 0:
                status_code = status.HTTP_207_MULTI_STATUS
                message = f'{success_count} inventaire(s) importé(s) avec succès, {error_count} erreur(s)'
            else:
                status_code = status.HTTP_400_BAD_REQUEST
                message = 'Aucun inventaire n\'a pu être importé'
            
            response_data = {
                'summary': summary
            }
            
            if return_details:
                if results:
                    response_data['results'] = results
                if errors:
                    response_data['errors'] = errors
            
            if error_count == 0:
                return success_response(
                    data=response_data,
                    message=message,
                    status_code=status_code
                )
            else:
                return error_response(
                    message=message,
                    errors=[err.get('error', 'Erreur inconnue') for err in errors] if errors else [],
                    status_code=status_code,
                    **response_data
                )
            
        except Exception as e:
            logger.error(f"Erreur lors de l'importation des inventaires: {str(e)}", exc_info=True)
            return error_response(
                message='Une erreur inattendue s\'est produite lors de l\'importation',
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class InventoryCompleteView(APIView):
    """
    Vue pour finaliser un inventaire.
    Marque un inventaire comme terminé uniquement si tous ses jobs sont terminés.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.service = InventoryService()

    def post(self, request, pk, *args, **kwargs):
        """
        Finalise un inventaire en vérifiant que tous les jobs sont terminés.
        Retourne la liste des jobs non terminés si l'inventaire ne peut pas être finalisé.
        
        Args:
            pk: L'ID de l'inventaire à finaliser
            
        Returns:
            Response: Réponse HTTP avec le statut de l'opération et la liste des jobs non terminés
        """
        try:
            # Finaliser l'inventaire via le service
            result = self.service.complete_inventory(pk)
            
            # Si l'inventaire a été finalisé avec succès
            if result['success']:
                # Sérialiser l'inventaire mis à jour
                serializer = InventoryDetailSerializer(result['inventory'])
                
                return success_response(
                    data=serializer.data,
                    message=result['message'],
                    jobs_not_completed=[],
                    total_jobs=result['total_jobs'],
                    completed_jobs=result['completed_jobs']
                )
            
            # Si l'inventaire n'a pas pu être finalisé (jobs non terminés)
            else:
                return error_response(
                    message=result['message'],
                    errors=[f"Job {job.get('reference', job.get('id', 'inconnu'))} non terminé" for job in result.get('jobs_not_completed', [])],
                    status_code=status.HTTP_400_BAD_REQUEST,
                    jobs_not_completed=result['jobs_not_completed'],
                    total_jobs=result['total_jobs'],
                    completed_jobs=result['completed_jobs']
                )
            
        except InventoryNotFoundError as e:
            logger.warning(f"Inventaire non trouvé lors de la finalisation: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_404_NOT_FOUND
            )
            
        except InventoryValidationError as e:
            logger.warning(f"Erreur de validation lors de la finalisation: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_400_BAD_REQUEST
            )
            
        except Exception as e:
            logger.error(f"Erreur lors de la finalisation de l'inventaire: {str(e)}", exc_info=True)
            return error_response(
                message="Une erreur inattendue s'est produite lors de la finalisation de l'inventaire",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class InventoryCloseView(APIView):
    """
    Vue pour clôturer un inventaire.
    Marque un inventaire comme clôturé uniquement s'il est déjà terminé.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.service = InventoryService()

    def post(self, request, pk, *args, **kwargs):
        """
        Clôture un inventaire en vérifiant qu'il est déjà terminé.
        
        Args:
            pk: L'ID de l'inventaire à clôturer
            
        Returns:
            Response: Réponse HTTP avec le statut de l'opération
        """
        try:
            # Clôturer l'inventaire via le service
            inventory = self.service.close_inventory(pk)
            
            # Sérialiser l'inventaire mis à jour
            serializer = InventoryDetailSerializer(inventory)
            
            return success_response(
                data=serializer.data,
                message="L'inventaire a été marqué comme clôturé avec succès"
            )
            
        except InventoryNotFoundError as e:
            logger.warning(f"Inventaire non trouvé lors de la clôture: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_404_NOT_FOUND
            )
            
        except InventoryValidationError as e:
            logger.warning(f"Erreur de validation lors de la clôture: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_400_BAD_REQUEST
            )
            
        except Exception as e:
            logger.error(f"Erreur lors de la clôture de l'inventaire: {str(e)}", exc_info=True)
            return error_response(
                message="Une erreur inattendue s'est produite lors de la clôture de l'inventaire",
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class StockImportView(APIView):
    """
    Vue pour importer des stocks via API.
    Permet l'importation de stocks depuis un fichier Excel pour un inventaire spécifique.
    Vérifie le type d'inventaire et applique les règles métier appropriées.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        from ..services.stock_service import StockService
        from ..usecases.stock_import_validation import StockImportValidationUseCase
        self.stock_service = StockService()
        self.validation_use_case = StockImportValidationUseCase()

    def post(self, request, inventory_id, *args, **kwargs):
        """
        Importe des stocks depuis un fichier Excel pour un inventaire spécifique.
        
        Règles métier:
        - Inventaire TOURNANT: Un seul import autorisé (refusé si stocks existants)
        - Inventaire GENERAL: Import autorisé (remplace les stocks existants)
        
        Args:
            inventory_id: L'ID de l'inventaire
            request.FILES['file']: Le fichier Excel à importer
            
        Format attendu du fichier Excel:
        - Colonnes requises: 'article', 'emplacement', 'quantite'
        - 'article': Référence du produit
        - 'emplacement': Référence de l'emplacement
        - 'quantite': Quantité disponible
        """
        try:
            # Vérifier qu'un fichier a été fourni
            if 'file' not in request.FILES:
                return error_response(
                    message='Aucun fichier fourni. Utilisez le champ "file" pour uploader le fichier Excel.',
                    status_code=status.HTTP_400_BAD_REQUEST
                )
            
            excel_file = request.FILES['file']
            
            # Vérifier l'extension du fichier
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                return error_response(
                    message='Le fichier doit être au format Excel (.xlsx ou .xls)',
                    status_code=status.HTTP_400_BAD_REQUEST
                )
            
            # ÉTAPE 1: Validation selon le type d'inventaire via le use case
            validation_result = self.validation_use_case.validate_stock_import(inventory_id)
            
            # Si l'import n'est pas autorisé, retourner l'erreur
            if not validation_result['can_import']:
                return error_response(
                    message=validation_result['message'],
                    status_code=status.HTTP_400_BAD_REQUEST,
                    inventory_type=validation_result['inventory_type'],
                    existing_stocks_count=validation_result['existing_stocks_count'],
                    action_required=validation_result['action_required']
                )
            
            # ÉTAPE 2: Si l'import est autorisé, procéder à l'import
            result = self.stock_service.import_stocks_from_excel(inventory_id, excel_file)
            
            # Préparer la réponse
            summary = {
                'total_rows': result['total_rows'],
                'valid_rows': result['valid_rows'],
                'invalid_rows': result['invalid_rows']
            }
            
            extra_data = {
                'inventory_type': validation_result['inventory_type'],
                'summary': summary
            }
            
            if result.get('imported_stocks'):
                extra_data['imported_stocks'] = result['imported_stocks']
            
            if result['success']:
                return success_response(
                    data=extra_data,
                    message=result['message'],
                    status_code=status.HTTP_201_CREATED
                )
            else:
                return error_response(
                    message=result['message'],
                    errors=result.get('errors', []),
                    status_code=status.HTTP_400_BAD_REQUEST,
                    **extra_data
                )
            
        except InventoryNotFoundError as e:
            logger.warning(f"Inventaire non trouvé lors de l'import de stocks: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_404_NOT_FOUND
            )
            
        except StockValidationError as e:
            logger.warning(f"Erreur de validation lors de l'import de stocks: {str(e)}")
            return error_response(
                message=str(e),
                status_code=status.HTTP_400_BAD_REQUEST
            )
            
        except Exception as e:
            logger.error(f"Erreur lors de l'import de stocks: {str(e)}", exc_info=True)
            return error_response(
                message='Une erreur inattendue s\'est produite lors de l\'import',
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
            ) 

 