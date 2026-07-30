"""
Service pour la génération de fichiers Excel consolidés par article
"""
from io import BytesIO
from typing import Optional, Tuple

from apps.inventory.constants import CountMode, InventoryType
from ..repositories.excel_export_repository import ExcelExportRepository


class ExcelExportService:
    """Service pour la génération d'Excel consolidé par article"""
    
    def __init__(self):
        self.repository = ExcelExportRepository()

    def validate_countings_for_export(
        self,
        inventory_id: int,
        inventory_type: str,
    ) -> Tuple[bool, Optional[str]]:
        """
        Valide les comptages selon le type d'inventaire.

        GENERAL conserve la règle existante : les comptages d'ordre 2 et 3
        doivent exister et être en mode « par article ».
        MAGASIN / TOURNANT sont mono-comptage et ne nécessitent pas ces ordres.
        """
        if inventory_type in InventoryType.SINGLE_COUNTING:
            return True, None

        counting_order_2, counting_order_3 = (
            self.repository.get_countings_orders_2_and_3(inventory_id)
        )

        if not counting_order_2:
            return False, "Le comptage d'ordre 2 n'existe pas pour cet inventaire."

        if not counting_order_3:
            return False, "Le comptage d'ordre 3 n'existe pas pour cet inventaire."

        if counting_order_2.count_mode.lower() != CountMode.BY_ARTICLE:
            return False, "Le mode de comptage de cet inventaire n'est pas 'par article'."

        if counting_order_3.count_mode.lower() != CountMode.BY_ARTICLE:
            return False, "Le mode de comptage de cet inventaire n'est pas 'par article'."

        return True, None
    
    def generate_consolidated_excel(
        self,
        inventory_id: int,
        warehouse_id: int,
    ) -> BytesIO:
        """
        Génère un fichier Excel consolidé par article.
        
        Le fichier contient :
        - Les informations de l'article (référence, code, description, etc.)
        - La quantité consolidée (somme de toutes les quantités)
        
        Args:
            inventory_id: ID de l'inventaire
            warehouse_id: ID du magasin
            
        Returns:
            BytesIO: Le contenu du fichier Excel en mémoire
            
        Raises:
            ValueError: Si l'inventaire n'existe pas, si les comptages d'ordre 2 et 3 
                       n'existent pas ou n'ont pas le mode "par article", ou s'il n'y a pas de données
            ImportError: Si pandas ou openpyxl ne sont pas installés
        """
        # Vérifier que l'inventaire existe
        inventory = self.repository.get_inventory_by_id(inventory_id)
        if not inventory:
            raise ValueError(f"Inventaire avec l'ID {inventory_id} non trouvé")

        warehouse = self.repository.get_warehouse_by_id(warehouse_id)
        if not warehouse:
            raise ValueError(f"Magasin avec l'ID {warehouse_id} non trouvé")

        # GENERAL garde la validation historique ; MAGASIN / TOURNANT sont mono-comptage.
        is_valid, error_message = self.validate_countings_for_export(
            inventory_id,
            inventory.inventory_type,
        )
        if not is_valid:
            raise ValueError(error_message)

        total_ecarts, resolved_ecarts = (
            self.repository.get_ecarts_resolution_counts(
                inventory_id,
                warehouse_id,
            )
        )
        if total_ecarts and resolved_ecarts != total_ecarts:
            raise ValueError(
                "Tous les écarts de comptage de cet inventaire doivent être "
                f"résolus pour ce magasin. Écarts résolus : "
                f"{resolved_ecarts}/{total_ecarts}"
            )

        # Limiter strictement la consolidation au magasin demandé.
        consolidated_data = (
            self.repository.get_consolidated_data_by_inventory_and_warehouse(
                inventory_id,
                warehouse_id,
            )
        )
        
        if not consolidated_data:
            raise ValueError(
                f"Aucune donnée trouvée pour l'inventaire {inventory_id} "
                f"et le magasin {warehouse_id}"
            )
        
        # Vérifier que pandas et openpyxl sont disponibles
        try:
            import pandas as pd
        except ImportError:
            raise ImportError(
                "pandas est requis pour l'export Excel. "
                "Installez-le avec: pip install pandas"
            )
        
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl est requis pour l'export Excel. "
                "Installez-le avec: pip install openpyxl"
            )
        
        # Construire les données pour le DataFrame
        rows = []
        
        for product_data in consolidated_data:
            row = {
                'Référence': product_data['product_reference'],
                'Code Produit': product_data['product_code'],
                'Désignation': product_data['product_description'],
                'Code-barres': product_data['product_barcode'],
                'Unité': product_data['product_unit'],
                'Famille': product_data['product_family'],
                'Quantité Consolidée': product_data['total_quantity'],
            }
            
            rows.append(row)
        
        # Créer le DataFrame
        df = pd.DataFrame(rows)
        
        # Générer le fichier Excel
        excel_buffer = BytesIO()
        
        try:
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Articles Consolidés')
                
                # Récupérer la feuille pour appliquer des formats
                worksheet = writer.sheets['Articles Consolidés']
                
                # Appliquer un style aux en-têtes
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Ajuster la largeur des colonnes
                for idx, col in enumerate(df.columns, start=1):
                    try:
                        # Remplacer les valeurs None par des chaînes vides
                        col_data = df[col].fillna('').astype(str)
                        
                        # Calculer la longueur maximale
                        if len(col_data) > 0:
                            max_data_length = col_data.map(len).max()
                        else:
                            max_data_length = 0
                        
                        max_length = max(max_data_length, len(str(col)))
                        # Limiter la largeur maximale à 50
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
                    except Exception as e:
                        # Si l'ajustement de la largeur échoue, utiliser une largeur par défaut
                        worksheet.column_dimensions[get_column_letter(idx)].width = 15
                
                # Figer la première ligne (en-têtes)
                worksheet.freeze_panes = 'A2'
                
        except Exception as e:
            raise ValueError(f"Impossible de générer le fichier Excel: {str(e)}")
        
        excel_buffer.seek(0)
        return excel_buffer

