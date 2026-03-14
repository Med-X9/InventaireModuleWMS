from typing import List, Dict, Any
import logging
from io import BytesIO
from ..repositories.job_export_repository import JobExportRepository
from ..exceptions.job_exceptions import JobCreationError
from ..exceptions.inventory_exceptions import InventoryNotFoundError
from ..exceptions.warehouse_exceptions import WarehouseNotFoundError

logger = logging.getLogger(__name__)


class JobExportService:
    """
    Service pour l'export des jobs prêts
    Contient la logique métier
    """
    
    def __init__(self):
        self.repository = JobExportRepository()
    
    def export_ready_jobs(
        self,
        inventory_id: int,
        warehouse_id: int
    ) -> List[Dict[str, Any]]:
        """
        Exporte les jobs avec assignments PRET ou TRANSFERT avec leurs détails
        Utilise la même logique que l'API PDF pour la récupération des jobs
        
        Args:
            inventory_id: ID de l'inventaire
            warehouse_id: ID de l'entrepôt
            
        Returns:
            Liste des données d'export formatées
            
        Raises:
            InventoryNotFoundError: Si l'inventaire n'existe pas
            WarehouseNotFoundError: Si l'entrepôt n'existe pas
            JobCreationError: Si une erreur métier survient
        """
        try:
            # Valider que l'inventaire existe
            from ..models import Inventory
            try:
                inventory = Inventory.objects.get(id=inventory_id)
                logger.debug(f"Inventaire trouvé: {inventory.reference} (ID: {inventory_id})")
            except Inventory.DoesNotExist:
                logger.warning(f"Tentative d'export avec inventaire inexistant: ID {inventory_id}")
                raise InventoryNotFoundError(f"Inventaire avec l'ID {inventory_id} non trouvé")
            
            # Valider que le warehouse existe
            from apps.masterdata.models import Warehouse
            try:
                warehouse = Warehouse.objects.get(id=warehouse_id)
                logger.debug(f"Entrepôt trouvé: {warehouse.warehouse_name} (ID: {warehouse_id})")
            except Warehouse.DoesNotExist:
                logger.warning(f"Tentative d'export avec entrepôt inexistant: ID {warehouse_id}")
                raise WarehouseNotFoundError(f"Entrepôt avec l'ID {warehouse_id} non trouvé")
            
            # Récupérer les jobs prêts
            jobs = self.repository.get_ready_jobs_by_inventory_and_warehouse(
                inventory_id,
                warehouse_id
            )
            
            if not jobs:
                # Log pour indiquer qu'il n'y a pas de jobs avec assignments PRET ou TRANSFERT
                logger.info(f"Aucun job avec assignments PRET ou TRANSFERT trouvé pour inventory_id={inventory_id}, warehouse_id={warehouse_id}")
                
                # Retourner une liste vide
                return []
            
            # Préparer les données d'export
            export_data = []
            
            for job in jobs:
                job_reference = job.reference
                
                # Récupérer les JobDetail pour ce job
                # Les JobDetail sont déjà triés par location_id et -id dans le prefetch
                job_details = job.jobdetail_set.all()
                
                # Récupérer les assignments pour ce job (déjà filtrés PRET/TRANSFERT dans le repository)
                # Trier les assignments par ordre de comptage pour garantir un ordre cohérent
                # IMPORTANT: Utiliser sorted() pour forcer le tri même si le prefetch a déjà un ordre
                assignments_queryset = job.assigment_set.all()
                assignments = sorted(assignments_queryset, key=lambda a: (a.counting.order if a.counting else 0, a.id))
                
                # Éliminer les doublons d'emplacements pour ce job
                # Utiliser un set pour tracker les locations déjà traitées
                seen_locations = set()
                unique_job_details = []
                
                for job_detail in job_details:
                    location_id = job_detail.location_id
                    if location_id not in seen_locations:
                        seen_locations.add(location_id)
                        unique_job_details.append(job_detail)
                
                # Pour chaque assignment, créer une ligne par emplacement
                # IMPORTANT: Le compteur de ligne est réinitialisé à 1 pour CHAQUE assignment
                for assignment in assignments:
                    counting = assignment.counting
                    session = assignment.session
                    
                    # Compteur de ligne pour cet assignment (RÉINITIALISÉ à 0 pour chaque assignment)
                    # Il sera incrémenté à 1 pour le premier emplacement, 2 pour le deuxième, etc.
                    ligne_numero = 0
                    
                    # Pour chaque JobDetail unique (location), créer une ligne d'export
                    # Le numéro de ligne va de 1 jusqu'au nombre d'emplacements pour CET assignment
                    for job_detail in unique_job_details:
                        location = job_detail.location
                        
                        # Incrémenter le compteur pour cet assignment (1, 2, 3, ..., nombre_emplacements)
                        ligne_numero += 1
                        
                        # Préparer les données de la ligne
                        ligne_data = {
                            'reference_job': job_reference,
                            'location': location.location_reference,
                            'code_article': '',  # Vide comme demandé
                            'code_barre': '',  # Vide comme demandé
                            'designation': '',  # Vide comme demandé
                            'quantite': '',  # Vide comme demandé
                            'session': session.username if session else '',
                            'ordre_comptage': counting.order if counting else '',
                            'ligne_numero': ligne_numero
                        }
                        
                        export_data.append(ligne_data)
            
            logger.info(f"Export réussi: {len(export_data)} lignes exportées pour inventory_id={inventory_id}, warehouse_id={warehouse_id}")
            return export_data
            
        except (InventoryNotFoundError, WarehouseNotFoundError, JobCreationError):
            # Re-lever les exceptions métier sans modification
            raise
        except Exception as e:
            # Encapsuler les autres erreurs dans JobCreationError
            logger.error(f"Erreur inattendue lors de l'export des jobs prêts (inventory_id={inventory_id}, warehouse_id={warehouse_id}): {str(e)}", exc_info=True)
            raise JobCreationError(
                f"Erreur lors de l'export des jobs prêts: {str(e)}"
            )
    
    def generate_excel_export(
        self,
        inventory_id: int,
        warehouse_id: int
    ) -> BytesIO:
        """
        Génère un fichier Excel pour l'export des jobs avec assignments PRET ou TRANSFERT
        Utilise la même logique que l'API PDF pour la récupération des jobs
        
        Args:
            inventory_id: ID de l'inventaire
            warehouse_id: ID de l'entrepôt
            
        Returns:
            BytesIO: Le contenu du fichier Excel en mémoire
            
        Raises:
            InventoryNotFoundError: Si l'inventaire n'existe pas
            WarehouseNotFoundError: Si l'entrepôt n'existe pas
            JobCreationError: Si une erreur métier survient
        """
        # Vérifier que pandas et openpyxl sont disponibles
        try:
            import pandas as pd
        except ImportError:
            raise JobCreationError(
                "pandas est requis pour l'export Excel. "
                "Installez-le avec: pip install pandas"
            )
        
        try:
            import openpyxl
        except ImportError:
            raise JobCreationError(
                "openpyxl est requis pour l'export Excel. "
                "Installez-le avec: pip install openpyxl"
            )
        
        # Récupérer les données d'export
        export_data = self.export_ready_jobs(inventory_id, warehouse_id)
        
        if not export_data:
            raise JobCreationError("Aucune donnée à exporter")
        
        # Construire les données pour le DataFrame
        rows = []
        for row_data in export_data:
            row = {
                'Référence Job': row_data['reference_job'],
                'Emplacement': row_data['location'],
                'Code Article': row_data['code_article'],
                'Code à Barre': row_data['code_barre'],
                'Désignation': row_data['designation'],
                'Quantité': row_data['quantite'],
                'Session': row_data['session'],
                'Ordre Comptage': row_data['ordre_comptage'],
                'Ligne N°': row_data['ligne_numero'],
            }
            rows.append(row)
        
        # Créer le DataFrame
        df = pd.DataFrame(rows)
        
        # Générer le fichier Excel
        excel_buffer = BytesIO()
        
        try:
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Jobs Prêts')
                
                # Récupérer la feuille pour appliquer des formats
                worksheet = writer.sheets['Jobs Prêts']
                
                # Appliquer un style aux en-têtes
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Ajuster la largeur des colonnes
                for idx, col in enumerate(df.columns, start=1):
                    try:
                        # Remplacer les valeurs None par des chaînes vides
                        col_data = df[col].fillna('').astype(str)
                        
                        # Calculer la longueur maximale
                        if len(col_data) > 0:
                            max_data_length = col_data.map(len).max()
                        else:
                            max_data_length = 0
                        
                        max_length = max(max_data_length, len(str(col)))
                        # Limiter la largeur maximale à 50
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
                    except Exception as e:
                        # Si l'ajustement de la largeur échoue, utiliser une largeur par défaut
                        worksheet.column_dimensions[get_column_letter(idx)].width = 15
                
                # Figer la première ligne (en-têtes)
                worksheet.freeze_panes = 'A2'
                
        except Exception as e:
            logger.error(f"Erreur lors de la génération du fichier Excel: {str(e)}", exc_info=True)
            raise JobCreationError(f"Impossible de générer le fichier Excel: {str(e)}")
        
        excel_buffer.seek(0)
        logger.info(f"Fichier Excel généré avec succès: {len(export_data)} lignes")
        return excel_buffer

